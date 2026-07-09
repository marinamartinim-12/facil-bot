"""
Fácil Financiamentos — Servidor principal v2
FastAPI + Webhook Z-API + Dashboard com login
"""

import asyncio
import base64
import json
import os
import re
import subprocess
import tempfile
import uuid
import httpx
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo

# Fuso horário de Brasília (UTC-3)
_TZ_BR = ZoneInfo("America/Sao_Paulo")

def _fmt_br(dt: datetime | None, fmt: str = "%d/%m/%Y %H:%M") -> str | None:
    """Converte datetime UTC para horário de Brasília e formata."""
    if not dt:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(_TZ_BR).strftime(fmt)

def _agora_br() -> datetime:
    """Retorna o datetime atual no fuso de Brasília."""
    return datetime.now(_TZ_BR)

def _data_br_para_utc(data_str: str):
    """'DD/MM/YYYY' (data BR) → datetime UTC naive (meio-dia BR). None se inválida."""
    try:
        d = datetime.strptime((data_str or "").strip(), "%d/%m/%Y")
        return d.replace(hour=12, tzinfo=_TZ_BR).astimezone(timezone.utc).replace(tzinfo=None)
    except Exception:
        return None
from fastapi import FastAPI, Request, Depends, HTTPException, Query, Response, UploadFile, File, Form
from fastapi.responses import HTMLResponse, JSONResponse, PlainTextResponse, FileResponse
import secrets
from pathlib import Path
from sqlalchemy.orm import Session

from config import get_settings
from models import Lead, MensagemConversa, Usuario, Configuracao, Contrato, Parceiro, ContatoParceiro, SessaoUsuario, AusenciaFuncionaria, RegistroPonto, JustificativaPonto, CorrecaoPonto, AtividadePing, Agendamento, MidiaArquivo, DocumentoCliente, HistoricoLead, criar_tabelas, get_db, StatusLeadEnum, ModalidadeEnum, RoleEnum, EstadoConversaEnum
from bot import processar_mensagem, obter_resumo_lead, _proximo_horario_atendimento, diagnostico_ia
from auth import verificar_senha, hash_senha, criar_token, obter_usuario_atual, requer_admin, requer_gestao, role_do_token

settings = get_settings()
app = FastAPI(title="Fácil Financiamentos", version="2.0.0")


# ── Guarda global: o perfil DONO é somente leitura ──────────────────────────────
# Bloqueia qualquer escrita (POST/PUT/DELETE/PATCH) vinda de um token de dono.
# Segurança de verdade no servidor — não depende de esconder botão no front.
_DONO_ESCRITA_LIBERADA = {"/auth/login", "/auth/logout", "/api/logout", "/api/heartbeat"}
_DONO_GET_BLOQUEADO_PREFIX = ("/api/admin/", "/api/debug/", "/api/placa-debug")
_DONO_GET_BLOQUEADO_EXATO = {"/api/diagnostico-ia", "/api/config"}

@app.middleware("http")
async def _guarda_dono_somente_leitura(request: Request, call_next):
    metodo = request.method
    path = request.url.path
    if metodo not in ("GET", "HEAD", "OPTIONS"):
        # Escrita: bloqueia TUDO pro dono (exceto login/logout/heartbeat)
        if path not in _DONO_ESCRITA_LIBERADA and role_do_token(request.cookies.get("access_token")) == "dono":
            return JSONResponse(
                {"detail": "Perfil do dono é somente leitura — sem permissão para alterar."},
                status_code=403,
            )
    else:
        # Leitura: bloqueia telas sensíveis/de ação (migração, limpar ponto, backup, debug, config…)
        if (path.startswith(_DONO_GET_BLOQUEADO_PREFIX) or path in _DONO_GET_BLOQUEADO_EXATO) \
                and role_do_token(request.cookies.get("access_token")) == "dono":
            return JSONResponse({"detail": "Sem permissão para esta área."}, status_code=403)
    return await call_next(request)

# Cache para deduplicar mensagens enviadas pelo painel vs. webhook fromMe
# Chave: (telefone, texto_normalizado) → timestamp do envio
_msgs_painel_recentes: dict[tuple, float] = {}
_TTL_DEDUP = 30  # segundos

def _registrar_msg_painel(telefone: str, texto: str):
    """Registra mensagem enviada pelo painel para evitar duplicata do webhook fromMe."""
    import time
    chave = (telefone, texto.strip().lower()[:100])
    _msgs_painel_recentes[chave] = time.time()
    # Limpa entradas antigas
    agora = time.time()
    expiradas = [k for k, t in _msgs_painel_recentes.items() if agora - t > _TTL_DEDUP]
    for k in expiradas:
        del _msgs_painel_recentes[k]

def _e_duplicata_painel(telefone: str, texto: str) -> bool:
    """Retorna True se essa mensagem foi enviada recentemente pelo painel."""
    import time
    chave = (telefone, texto.strip().lower()[:100])
    ts = _msgs_painel_recentes.get(chave)
    if ts and time.time() - ts < _TTL_DEDUP:
        del _msgs_painel_recentes[chave]  # consome a entrada
        return True
    return False


# ─── Follow-up automático ───────────────────────────────────────────────────────

# Estados do bot onde o lead ainda não completou os dados para o consultor
_ESTADOS_BOT_ATIVO = [
    EstadoConversaEnum.inicio,
    EstadoConversaEnum.aguardando_nome,
    EstadoConversaEnum.aguardando_modalidade,
    EstadoConversaEnum.coletando_cidade,
    EstadoConversaEnum.coletando_cpf,
    EstadoConversaEnum.coletando_data_nasc,
    EstadoConversaEnum.coletando_carro,
]


def _dentro_horario_atendimento() -> bool:
    """Retorna True se estiver dentro do horário de funcionamento (horário de Brasília)."""
    from zoneinfo import ZoneInfo
    agora = datetime.now(ZoneInfo("America/Sao_Paulo"))
    dia = agora.weekday()       # 0=seg … 6=dom
    hora_dec = agora.hour + agora.minute / 60
    if dia < 5 and 9 <= hora_dec < 18:   # segunda a sexta
        return True
    if dia == 5 and 9 <= hora_dec < 13:  # sábado
        return True
    return False


async def _enviar_followups():
    from models import SessionLocal
    db = SessionLocal()
    try:
        # ── 1. Só envia dentro do horário de atendimento ──────────────────────
        if not _dentro_horario_atendimento():
            print("⏰ Follow-up ignorado: fora do horário de atendimento")
            return

        # ── 2. Carrega configurações ──────────────────────────────────────────
        def _cfg(chave, padrao=""):
            c = db.query(Configuracao).filter(Configuracao.chave == chave).first()
            return c.valor if c else padrao

        # Interruptor geral — permite desligar todos os follow-ups
        if _cfg("followup_ativo", "1").strip().lower() in ("0", "false", "nao", "não", "off"):
            print("⏸️ Follow-up desativado nas configurações")
            return

        horas = int(h) if (h := _cfg("followup_horas", "4")).isdigit() else 4
        limite_1 = datetime.utcnow() - timedelta(hours=horas)   # 1º: X h sem resposta

        # Segurança anti-banimento: teto por rodada e idade máxima do lead
        max_rodada = int(m) if (m := _cfg("followup_max_rodada", "20")).isdigit() else 20
        max_dias   = int(d) if (d := _cfg("followup_max_dias", "15")).isdigit() else 15

        msgs = {
            0: _cfg("mensagem_followup",
                    "Oi! 😊 Vi que nossa conversa ficou parada...\n"
                    "Quando quiser continuar, estou aqui! Gostaria de retomar?"),
            1: _cfg("mensagem_followup_2",
                    "Olá! 👋 Passando para saber se ainda tem interesse em financiar ou refinanciar seu veículo.\n"
                    "Estamos com ótimas condições! Ficou alguma dúvida?"),
            2: _cfg("mensagem_followup_3",
                    "Oi! Última tentativa de contato por aqui. 😊\n"
                    "Se mudar de ideia, pode nos chamar a qualquer momento! Ficamos à disposição. 🤝"),
        }

        # ── 3. Leads ainda no fluxo do bot ───────────────────────────────────
        leads = db.query(Lead).filter(
            Lead.estado_conversa.in_([e.value for e in _ESTADOS_BOT_ATIVO]),
            Lead.status.in_([
                StatusLeadEnum.em_atendimento.value,
                StatusLeadEnum.qualificado.value,
            ]),
        ).all()

        enviados = 0
        optout_on = _cfg("followup_optout", "1").strip().lower() not in ("0", "false", "off")
        optout_txt = _cfg("followup_optout_texto", "\n\n_Se não quiser mais receber, responda SAIR._")

        for lead in leads:
            try:
                # Pula leads manuais sem telefone real
                if lead.telefone.startswith("_manual_"):
                    continue
                # Pula quem pediu para não receber mais (opt-out)
                if lead.descadastrado:
                    continue

                # Última mensagem do usuário
                ultima_user = (
                    db.query(MensagemConversa)
                    .filter(
                        MensagemConversa.telefone == lead.telefone,
                        MensagemConversa.role == "user",
                    )
                    .order_by(MensagemConversa.id.desc())
                    .first()
                )

                if not ultima_user:
                    continue

                tentativa = lead.followup_tentativa or 0
                agora = datetime.utcnow()

                # Se o cliente respondeu após o último follow-up → zera a sequência
                if tentativa > 0 and lead.followup_em and ultima_user.criado_em > lead.followup_em:
                    lead.followup_tentativa = 0
                    db.commit()
                    tentativa = 0

                # ── Decide se é hora de disparar ─────────────────────────────
                if tentativa == 0:
                    # Não começa a cutucar lead que ficou frio há muito tempo (anti-spam)
                    if ultima_user.criado_em < (agora - timedelta(days=max_dias)):
                        continue
                    # 1º: cliente ficou X horas sem responder
                    if ultima_user.criado_em > limite_1:
                        continue   # ainda recente
                elif tentativa == 1:
                    # 2º: 24h após o 1º sem resposta
                    if not lead.followup_em or (agora - lead.followup_em) < timedelta(hours=24):
                        continue
                elif tentativa == 2:
                    # 3º: 48h após o 2º sem resposta
                    if not lead.followup_em or (agora - lead.followup_em) < timedelta(hours=48):
                        continue
                else:
                    continue   # já esgotou as 3 tentativas

                # ── Monta e envia ─────────────────────────────────────────────
                nome = f" {lead.nome}" if lead.nome else ""
                texto_base = msgs[tentativa]
                texto = texto_base.replace("{nome}", nome.strip()).replace("Oi!", f"Oi{nome}!")
                if optout_on and optout_txt:
                    texto = texto + optout_txt

                await enviar_zapi(lead.telefone, texto)
                _salvar_msg_webhook(db, lead.telefone, texto, role="assistant")

                lead.followup_em = agora
                lead.followup_tentativa = tentativa + 1

                # 3º follow-up enviado → marca como Perdido automaticamente
                if tentativa == 2:
                    lead.status = StatusLeadEnum.perdido
                    print(f"🔴 Lead #{lead.id} marcado como Perdido após 3 follow-ups sem resposta")

                db.commit()
                enviados += 1
                print(f"📨 Follow-up #{tentativa + 1} enviado para {lead.telefone} (lead #{lead.id})")

                # Teto por rodada — evita rajada de mensagens (parece spam → risco de ban)
                if enviados >= max_rodada:
                    print(f"🛑 Teto de {max_rodada} follow-ups por rodada atingido")
                    break
                # Espaça os envios para não disparar tudo no mesmo segundo
                await asyncio.sleep(6)

            except Exception as e_lead:
                print(f"⚠️ Erro ao enviar follow-up para lead #{lead.id}: {e_lead}")
                db.rollback()

        if enviados:
            print(f"✅ Follow-ups enviados nesta rodada: {enviados}")
        else:
            print("ℹ️ Nenhum lead precisava de follow-up agora")

    except Exception as e:
        print(f"❌ Erro geral no follow-up: {e}")
    finally:
        db.close()


async def _loop_followup():
    """Roda a cada 30 minutos verificando leads parados."""
    await asyncio.sleep(60)  # aguarda 1 min após startup
    while True:
        try:
            print("🔍 Verificando leads para follow-up…")
            await _enviar_followups()
        except Exception as e:
            print(f"❌ Erro inesperado no loop de follow-up: {e}")
        await asyncio.sleep(30 * 60)  # a cada 30 minutos


async def _loop_ocultar_inativos():
    """Roda 1x por dia: oculta do funil leads sem atividade há 30+ dias."""
    await asyncio.sleep(120)  # aguarda 2 min após startup
    while True:
        try:
            from models import SessionLocal
            db = SessionLocal()
            limite = datetime.utcnow() - timedelta(days=30)
            inativos = db.query(Lead).filter(
                Lead.atualizado_em < limite,
                Lead.oculto_funil == False,
                Lead.status.notin_([
                    StatusLeadEnum.desqualificado.value,
                ]),
            ).all()
            total = 0
            for lead in inativos:
                lead.oculto_funil = True
                total += 1
            if total:
                db.commit()
                print(f"📦 {total} lead(s) ocultados do funil por inatividade (30 dias)")

            # ── Limpeza mensal: no virar do mês, esconde do funil os CONTRATOS
            #    FECHADOS do mês anterior (1x por mês). Os dados continuam salvos. ──
            try:
                agora_br = _agora_br()
                mes_atual = agora_br.strftime("%Y-%m")
                cfg = db.query(Configuracao).filter(Configuracao.chave == "ultima_limpeza_funil").first()
                if not cfg or cfg.valor != mes_atual:
                    inicio_mes_utc = agora_br.replace(day=1, hour=0, minute=0, second=0,
                                                      microsecond=0).astimezone(timezone.utc).replace(tzinfo=None)
                    fechados = db.query(Lead).filter(
                        Lead.status == StatusLeadEnum.fechado,
                        Lead.oculto_funil == False,
                        Lead.atualizado_em < inicio_mes_utc,
                    ).all()
                    for l in fechados:
                        l.oculto_funil = True
                    if cfg:
                        cfg.valor = mes_atual
                    else:
                        db.add(Configuracao(chave="ultima_limpeza_funil", valor=mes_atual,
                                            descricao="Último mês em que o funil foi limpo dos contratos fechados"))
                    db.commit()
                    print(f"🗓️ Limpeza mensal do funil: {len(fechados)} contrato(s) fechado(s) ocultado(s) ({mes_atual})")
            except Exception as e:
                print(f"⚠️ Erro na limpeza mensal do funil: {e}")

            db.close()
        except Exception as e:
            print(f"❌ Erro ao ocultar leads inativos: {e}")
        await asyncio.sleep(24 * 60 * 60)  # 1x por dia


# ─── Startup ────────────────────────────────────────────────────────────────────

@app.on_event("startup")
async def startup():
    # ── LIMPEZA EMERGENCIAL: remove backups locais que encheram o disco ──
    # (o backup automático foi removido; isso libera o espaço que ele ocupou)
    try:
        from models import engine as _eng
        _dbp = _eng.url.database
        if _dbp:
            _bkp_dir = os.path.join(os.path.dirname(_dbp) or ".", "backups")
            if os.path.isdir(_bkp_dir):
                import shutil as _shutil
                _shutil.rmtree(_bkp_dir, ignore_errors=True)
                print(f"🧹 Backups locais removidos para liberar disco: {_bkp_dir}")
    except Exception as e:
        print(f"⚠️ limpeza de backups: {e}")

    # Garante que o diretório do banco existe (volume /data ou local)
    db_url = settings.DATABASE_URL
    if "sqlite" in db_url:
        db_path = db_url.replace("sqlite:///", "").replace("sqlite://", "")
        db_dir = os.path.dirname(db_path)
        if db_dir:
            try:
                os.makedirs(db_dir, exist_ok=True)
                print(f"📁 Diretório do banco: {db_dir}")
            except Exception as e:
                print(f"⚠️ Não foi possível criar diretório {db_dir}: {e}")

    try:
        criar_tabelas()
    except Exception as e:
        print(f"⚠️ Erro ao criar tabelas: {e}")
        raise

    # Conexão usada nas migrações e no setup
    db_startup = next(get_db())
    db = db_startup

    # Migrações de schema PRIMEIRO — antes de qualquer consulta ORM, senão o
    # ORM tenta ler colunas novas que ainda não existem e o app quebra no boot.
    from sqlalchemy import text
    _migracoes = [
        ("leads",     "followup_em",           "DATETIME"),
        ("contratos", "dados_contrato",        "TEXT"),
        ("contratos", "selfie_path",           "VARCHAR(300)"),
        ("contratos", "assinatura_path",       "VARCHAR(300)"),
        ("contratos", "ip_cliente",            "VARCHAR(50)"),
        ("contratos", "geolocalizacao",        "VARCHAR(200)"),
        ("contratos", "pdf_assinado",          "VARCHAR(300)"),
        ("contratos", "assinado_em",           "DATETIME"),
        ("contratos", "doc_frente_req_path",   "VARCHAR(300)"),
        ("contratos", "doc_verso_req_path",    "VARCHAR(300)"),
        ("contratos", "token_prop",            "VARCHAR(64)"),
        ("contratos", "status_prop",           "VARCHAR(20)"),
        ("contratos", "selfie_prop_path",      "VARCHAR(300)"),
        ("contratos", "assinatura_prop_path",  "VARCHAR(300)"),
        ("contratos", "doc_frente_prop_path",  "VARCHAR(300)"),
        ("contratos", "doc_verso_prop_path",   "VARCHAR(300)"),
        ("contratos", "ip_prop",               "VARCHAR(50)"),
        ("contratos", "geo_prop",              "VARCHAR(200)"),
        ("contratos", "assinado_prop_em",      "DATETIME"),
        ("contratos", "codigo_req",            "VARCHAR(10)"),
        ("contratos", "codigo_req_expira",     "DATETIME"),
        ("contratos", "codigo_prop",           "VARCHAR(10)"),
        ("contratos", "codigo_prop_expira",    "DATETIME"),
        ("leads",     "origem",                "VARCHAR(50)"),
        ("leads",     "origem_detalhe",          "VARCHAR(100)"),
        ("leads",     "parceiro_id",             "INTEGER"),
        ("parceiros", "telefones_extras",         "TEXT"),
        ("sessoes_usuario", "tempo_ativo_s",      "INTEGER DEFAULT 0"),
        ("leads",           "followup_tentativa", "INTEGER DEFAULT 0"),
        ("leads",           "deal_data",          "VARCHAR(10)"),
        ("leads",           "deal_veiculo",       "VARCHAR(200)"),
        ("leads",           "deal_placa",         "VARCHAR(10)"),
        ("leads",           "fechado_em",         "DATETIME"),
        ("leads",           "deal_retorno",       "VARCHAR(5)"),
        ("leads",           "deal_valor",         "VARCHAR(20)"),
        ("leads",           "deal_comissao",      "VARCHAR(20)"),
        ("leads",           "deal_banco",         "VARCHAR(30)"),
        ("leads",           "deal_conta_pg",      "VARCHAR(50)"),
        ("leads",           "deal_operadora",     "VARCHAR(150)"),
        ("leads",           "dados_contrato",     "TEXT"),
        ("leads",           "cidade",             "VARCHAR(100)"),
        ("leads",           "renda",              "VARCHAR(30)"),
        ("leads",           "profissao",          "VARCHAR(100)"),
        ("leads",           "tem_cnh",            "BOOLEAN"),
        ("leads",           "oculto_funil",       "BOOLEAN DEFAULT 0"),
        ("leads",           "descadastrado",      "BOOLEAN DEFAULT 0"),
        ("leads",           "ignorar_relatorios", "BOOLEAN DEFAULT 0"),
        ("leads",           "carros_proposta",    "TEXT"),
        ("leads",           "email",              "VARCHAR(150)"),
        ("parceiros",       "nome_agenda",        "VARCHAR(200)"),
        ("parceiros",       "operadora_id",       "INTEGER"),
        ("agendamentos",    "resultado",          "TEXT"),
        ("registros_ponto", "foto_filename",      "VARCHAR(64)"),
        ("leads",           "lido_em",            "DATETIME"),
        ("leads",           "nao_lido_manual",    "BOOLEAN DEFAULT FALSE"),
        ("usuarios",        "data_admissao",      "VARCHAR(10)"),
        ("usuarios",        "ferias_ajuste",      "INTEGER DEFAULT 0"),
        ("ausencias_funcionaria", "status",          "VARCHAR(20) DEFAULT 'aprovada'"),
        ("ausencias_funcionaria", "solicitante_id",  "INTEGER"),
        ("ausencias_funcionaria", "desconta_ferias", "BOOLEAN DEFAULT 0"),
        ("ausencias_funcionaria", "obs_admin",       "VARCHAR(400)"),
        ("ausencias_funcionaria", "resolvido_por",   "INTEGER"),
        ("ausencias_funcionaria", "resolvido_em",    "DATETIME"),
    ]
    for tabela, coluna, tipo in _migracoes:
        try:
            with db_startup.bind.connect() as conn:
                conn.execute(text(f"ALTER TABLE {tabela} ADD COLUMN {coluna} {tipo}"))
                conn.commit()
            print(f"✅ Migração: {tabela}.{coluna} adicionada")
        except Exception:
            pass  # coluna já existe — ignorar

    # Cria admin padrão se não existir nenhum usuário (DEPOIS das migrações!)
    try:
        if db.query(Usuario).count() == 0:
            admin = Usuario(
                nome=settings.ADMIN_NOME,
                email=settings.ADMIN_EMAIL,
                senha_hash=hash_senha(settings.ADMIN_PASSWORD),
                role=RoleEnum.admin,
                ativo=True,
            )
            db.add(admin)
            db.commit()
            print(f"✅ Admin criado: {settings.ADMIN_EMAIL} / {settings.ADMIN_PASSWORD}")
    except Exception as e:
        print(f"⚠️ admin: {e}")

    # Backfill: define fechado_em dos leads já fechados a partir da data do negócio
    try:
        pendentes = db_startup.query(Lead).filter(
            Lead.status == StatusLeadEnum.fechado, Lead.fechado_em.is_(None)
        ).all()
        n = 0
        for l in pendentes:
            dt = _data_br_para_utc(l.deal_data) if l.deal_data else None
            if dt:
                l.fechado_em = dt
                n += 1
        if n:
            db_startup.commit()
            print(f"✅ Backfill fechado_em em {n} contrato(s) a partir da data do negócio")
    except Exception as e:
        print(f"⚠️ Erro no backfill fechado_em: {e}")

    # Configurações padrão do bot
    _criar_config_padrao(db_startup)
    db_startup.close()

    # Inicia tarefa de follow-up automático
    asyncio.create_task(_loop_followup())
    asyncio.create_task(_loop_ocultar_inativos())

    print("✅ Fácil Financiamentos Bot v2 iniciado!")
    print(f"🗄️  Banco: {settings.DATABASE_URL}")
    print("📊 Dashboard: http://localhost:8000/dashboard")


# ─── Autenticação ────────────────────────────────────────────────────────────────

async def _geo_por_ip(ip: str) -> str:
    """Retorna 'Cidade, Estado, País' via ip-api.com (grátis, sem chave)."""
    if not ip or ip in ("127.0.0.1", "::1", "testclient"):
        return "Local"
    try:
        async with httpx.AsyncClient(timeout=3.0) as client:
            r = await client.get(
                f"http://ip-api.com/json/{ip}",
                params={"fields": "status,city,regionName,country"},
            )
            d = r.json()
            if d.get("status") == "success":
                partes = [d.get("city",""), d.get("regionName",""), d.get("country","")]
                return ", ".join(p for p in partes if p)
    except Exception:
        pass
    return ip


def _ip_da_requisicao(request: Request) -> str:
    fwd = request.headers.get("X-Forwarded-For")
    if fwd:
        return fwd.split(",")[0].strip()
    return request.client.host if request.client else "desconhecido"


@app.post("/auth/login")
async def login(request: Request, response: Response, db: Session = Depends(get_db)):
    body = await request.json()
    email = body.get("email", "").strip().lower()
    senha = body.get("senha", "")

    usuario = db.query(Usuario).filter(Usuario.email == email).first()
    if not usuario or not verificar_senha(senha, usuario.senha_hash):
        raise HTTPException(status_code=401, detail="E-mail ou senha incorretos")
    if not usuario.ativo:
        raise HTTPException(status_code=403, detail="Usuário desativado")

    token = criar_token({"sub": str(usuario.id), "role": getattr(usuario.role, "value", usuario.role)})
    response.set_cookie(
        key="access_token",
        value=token,
        httponly=True,
        samesite="lax",
        max_age=60 * 60 * 24 * 7,
    )

    # ── Fecha sessões anteriores que ficaram abertas (aba fechada sem clicar Sair) ──
    sessoes_abertas = db.query(SessaoUsuario).filter(
        SessaoUsuario.usuario_id == usuario.id,
        SessaoUsuario.logout_em == None,
    ).all()
    for s in sessoes_abertas:
        # Usa o último ping como horário real de saída
        s.logout_em = s.ultimo_ativo_em or s.login_em
    if sessoes_abertas:
        db.commit()

    # ── Registrar sessão ──────────────────────────────────────────────────────
    ip = _ip_da_requisicao(request)
    geo = await _geo_por_ip(ip)
    sessao = SessaoUsuario(usuario_id=usuario.id, ip=ip, localizacao=geo)
    db.add(sessao)
    db.commit()
    db.refresh(sessao)
    response.set_cookie(
        key="sessao_id",
        value=str(sessao.id),
        httponly=True,
        samesite="lax",
        max_age=60 * 60 * 24 * 7,
    )

    return {"id": usuario.id, "nome": usuario.nome, "email": usuario.email, "role": usuario.role}


@app.post("/auth/logout")
async def logout(request: Request, response: Response, db: Session = Depends(get_db)):
    # Fecha sessão aberta
    sid = request.cookies.get("sessao_id")
    if sid:
        try:
            sessao = db.query(SessaoUsuario).filter(SessaoUsuario.id == int(sid)).first()
            if sessao and not sessao.logout_em:
                sessao.logout_em = datetime.utcnow()
                db.commit()
        except Exception:
            pass
    response.delete_cookie("access_token")
    response.delete_cookie("sessao_id")
    return {"status": "ok"}


@app.post("/api/heartbeat")
async def heartbeat(request: Request, response: Response, db: Session = Depends(get_db),
                    usuario: Usuario = Depends(obter_usuario_atual)):
    """Atualiza último momento ativo. Se ativo=true, incrementa tempo_ativo_s.
    Se a sessão não existir (ex: histórico zerado), cria uma nova automaticamente."""
    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    realmente_ativo = body.get("ativo", False)

    sessao = None
    sid = request.cookies.get("sessao_id")
    if sid:
        try:
            sessao = db.query(SessaoUsuario).filter(SessaoUsuario.id == int(sid)).first()
        except Exception:
            pass

    # Sessão não encontrada (deletada ou cookie antigo) — recria
    if not sessao:
        ip  = _ip_da_requisicao(request)
        geo = await _geo_por_ip(ip)
        sessao = SessaoUsuario(usuario_id=usuario.id, ip=ip, localizacao=geo)
        db.add(sessao)
        db.commit()
        db.refresh(sessao)
        response.set_cookie(
            key="sessao_id", value=str(sessao.id),
            httponly=True, samesite="lax", max_age=60 * 60 * 24 * 7,
        )

    sessao.ultimo_ativo_em = datetime.utcnow()
    if realmente_ativo:
        sessao.tempo_ativo_s = (sessao.tempo_ativo_s or 0) + 60
        # Registra ping com horário para cruzar com a jornada de ponto (admin only)
        db.add(AtividadePing(usuario_id=usuario.id, timestamp=datetime.utcnow()))
    db.commit()
    return {"status": "ok"}


@app.get("/auth/me")
async def me(usuario: Usuario = Depends(obter_usuario_atual)):
    return {"id": usuario.id, "nome": usuario.nome, "email": usuario.email, "role": usuario.role}



# ─── WhatsApp: envio ─────────────────────────────────────────────────────────────

async def enviar_zapi(telefone: str, mensagem: str):
    if not settings.ZAPI_INSTANCE or not settings.ZAPI_TOKEN:
        print(f"[Z-API SIMULADO] {telefone}: {mensagem}")
        return
    url = f"https://api.z-api.io/instances/{settings.ZAPI_INSTANCE}/token/{settings.ZAPI_TOKEN}/send-text"
    headers = {"Client-Token": settings.ZAPI_CLIENT_TOKEN}
    payload = {"phone": telefone, "message": mensagem}
    async with httpx.AsyncClient() as client:
        resp = await client.post(url, headers=headers, json=payload, timeout=10)
        if resp.status_code != 200:
            print(f"⚠️ Erro Z-API: {resp.text}")


async def enviar_meta(telefone: str, mensagem: str):
    if not settings.WHATSAPP_TOKEN or not settings.WHATSAPP_PHONE_ID:
        print(f"[META SIMULADO] {telefone}: {mensagem}")
        return
    url = f"https://graph.facebook.com/v19.0/{settings.WHATSAPP_PHONE_ID}/messages"
    headers = {"Authorization": f"Bearer {settings.WHATSAPP_TOKEN}", "Content-Type": "application/json"}
    payload = {"messaging_product": "whatsapp", "to": telefone, "type": "text", "text": {"body": mensagem}}
    async with httpx.AsyncClient() as client:
        resp = await client.post(url, headers=headers, json=payload, timeout=10)
        if resp.status_code != 200:
            print(f"⚠️ Erro Meta: {resp.text}")


# ─── Webhooks ────────────────────────────────────────────────────────────────────

def _montar_msg_recontato(lead) -> list[str]:
    """Mensagem para um cliente que JÁ existe e volta a falar.
    - Fora do horário: manda só a mensagem de horário (com o nome), igual à preferida.
    - Dentro do horário: boas-vindas personalizada + 'em breve uma consultora'."""
    nome = lead.nome or ""
    saud_nome = f" {nome}" if nome else ""

    # FORA do expediente → a mensagem que o proprietário aprovou, com o nome do cliente
    if _proximo_horario_atendimento():   # != "" → estamos FORA do horário
        return [
            f"Olá{saud_nome}! 😊 No momento estamos fora do horário de atendimento. "
            "Funcionamos seg–sex das 09h às 18h e sábado das 09h às 13h. "
            "Retornaremos seu contato no primeiro horário disponível! 🕘"
        ]

    # DENTRO do expediente → boas-vindas personalizada + dados + consultora
    saudacao = f"Olá{saud_nome}! Que bom ter você de volta! 😊"
    linhas = []
    mod_map = {"financiamento": "Financiamento", "refinanciamento": "Refinanciamento"}
    if lead.modalidade and lead.modalidade != "indefinido":
        linhas.append(f"📋 Modalidade: {mod_map.get(lead.modalidade, lead.modalidade)}")
    if lead.carro_interesse:
        linhas.append(f"🚗 Veículo de interesse: {lead.carro_interesse}")
    if linhas:
        msg1 = f"{saudacao}\n\nEncontrei seus dados cadastrados aqui:\n" + "\n".join(linhas)
    else:
        msg1 = f"{saudacao}\nSeus dados estão registrados no nosso sistema."
    msg2 = "Em breve uma de nossas consultoras entrará em contato. Tem alguma informação que mudou desde nossa última conversa? 😊"
    return [msg1, msg2]


async def _reativar_lead_perdido(lead, texto: str, db, enviar_fn) -> bool:
    """
    Trata lead 'perdido' que voltou a enviar mensagem.
    - Com dados: reativa como qualificado, manda boas-vindas personalizadas.
    - Sem dados: reinicia o fluxo do bot do zero.
    Retorna True se tratou como recontato (chamador não precisa fazer mais nada).
    """
    _salvar_msg_webhook(db, lead.telefone, texto, role="user")

    if lead.nome:
        # ── Tem dados: boas-vindas + passa para equipe ─────────────────────
        msgs = _montar_msg_recontato(lead)
        lead.status = StatusLeadEnum.qualificado
        lead.estado_conversa = EstadoConversaEnum.finalizado
        lead.atribuido_para = None      # libera para qualquer atendente assumir
        lead.assumido_em = None
        lead.followup_em = None
        lead.atualizado_em = datetime.utcnow()
        db.commit()
        for i, msg in enumerate(msgs):
            if i > 0:
                await asyncio.sleep(0.8)
            await enviar_fn(lead.telefone, msg)
            _salvar_msg_webhook(db, lead.telefone, msg, role="assistant")
        print(f"🔄 Lead #{lead.id} ({lead.nome}) reativado — tinha dados, voltou como qualificado")
        return True
    else:
        # ── Sem dados: reinicia o bot do zero ─────────────────────────────
        lead.status = StatusLeadEnum.em_atendimento
        lead.estado_conversa = EstadoConversaEnum.inicio
        lead.followup_em = None
        lead.atualizado_em = datetime.utcnow()
        db.commit()
        print(f"🔄 Lead #{lead.id} reativado — sem dados, reiniciando fluxo")
        return False   # deixa o fluxo normal do bot processar


def _extrair_texto_zapi(body: dict) -> str:
    """Extrai o texto de mensagem de qualquer tipo de payload Z-API."""
    return (
        (body.get("text") or {}).get("message", "")
        or (body.get("extendedTextMessage") or {}).get("text", "")
        or (body.get("listResponseMessage") or {}).get("title", "")
        or (body.get("buttonsResponseMessage") or {}).get("selectedDisplayText", "")
        or body.get("body", "")
        or ""
    ).strip()


def _tel_canonico(t) -> str:
    """Forma comparável de um telefone BR: sem código do país (55) e sem o 9º dígito → DDD + 8 dígitos.
    Faz '5531999998888', '31999998888' e '3199998888' baterem entre si."""
    d = "".join(c for c in str(t or "") if c.isdigit())
    if d.startswith("55") and len(d) >= 12:
        d = d[2:]                       # remove o código do país
    if len(d) == 11 and d[2:3] == "9":  # DDD + 9 + 8 → remove o 9º dígito
        d = d[:2] + d[3:]
    return d


def _nome_so_prefixo(nome) -> bool:
    """True se o nome é só o prefixo 'Pc'/'Pv' (sem nome real), ex.: 'Pc', 'Pv -', 'Pc/'."""
    resto = re.sub(r"^p[cv][\s\-/.]*", "", (nome or "").strip(), flags=re.I).strip()
    return len(resto) < 2


def _buscar_parceiro_por_telefone(telefone: str, db) -> "Parceiro | None":
    """Retorna o Parceiro ativo cujo telefone (principal ou extra) bate com o número.
    Comparação robusta: ignora código do país e o 9º dígito (formatos diferentes do WhatsApp)."""
    from models import Parceiro as _Parceiro
    alvo = _tel_canonico(telefone)
    if not alvo:
        return None
    for p in db.query(_Parceiro).filter(_Parceiro.ativo == True).all():
        if _tel_canonico(p.telefone) == alvo:
            return p
        try:
            if any(_tel_canonico(e) == alvo for e in json.loads(p.telefones_extras or "[]")):
                return p
        except Exception:
            pass
    return None


def _extrair_audio_url_zapi(body: dict) -> str | None:
    """Extrai a URL do áudio de um webhook Z-API, se houver."""
    audio = body.get("audio") or {}
    if audio.get("audioUrl"):
        return audio["audioUrl"]
    ptt = body.get("ptt") or {}
    if ptt.get("audioUrl"):
        return ptt["audioUrl"]
    return None


def _extrair_imagem_zapi(body: dict) -> dict | None:
    """Extrai URL e legenda de imagem do webhook Z-API."""
    img = body.get("image") or {}
    if img.get("imageUrl"):
        return {"url": img["imageUrl"], "caption": img.get("caption", "")}
    return None


def _extrair_documento_zapi(body: dict) -> dict | None:
    """Extrai URL, nome e mime de documento/PDF do webhook Z-API."""
    doc = body.get("document") or {}
    if doc.get("documentUrl"):
        return {
            "url": doc["documentUrl"],
            "nome": doc.get("fileName", "documento"),
            "mime": doc.get("mimeType", "application/octet-stream"),
        }
    return None


def _guardar_blob(db, filename: str, tipo: str, dados: bytes,
                  nome_original: str = None, mime: str = None, subdir: str = None):
    """Guarda os bytes do arquivo NO BANCO (durável) e também em disco (cache local).
    O banco é a fonte de verdade — o disco é reciclado pelo Railway a cada deploy."""
    # 1) Banco — persistente
    try:
        if not db.query(MidiaArquivo).filter(MidiaArquivo.filename == filename).first():
            db.add(MidiaArquivo(
                filename=filename, tipo=tipo, nome_original=(nome_original or "")[:200],
                mime=(mime or "")[:120], dados=dados, tamanho=len(dados),
            ))
            db.flush()  # garante INSERT; o commit ocorre no fluxo do webhook
    except Exception as e:
        print(f"⚠️ Erro ao guardar mídia no banco ({filename}): {e}")
    # 2) Disco — cache rápido enquanto o container vive
    if subdir:
        try:
            os.makedirs(f"/app/{subdir}", exist_ok=True)
            with open(f"/app/{subdir}/{filename}", "wb") as f:
                f.write(dados)
        except Exception as e:
            print(f"⚠️ Erro ao gravar mídia em disco ({filename}): {e}")


def _tamanho_legivel(n: int) -> str:
    """Converte bytes em texto legível: 1.2 MB, 340 KB, etc."""
    n = n or 0
    if n < 1024:
        return f"{n} B"
    if n < 1024 * 1024:
        return f"{n/1024:.0f} KB"
    return f"{n/(1024*1024):.1f} MB"


def _transcode_audio_sync(audio_bytes: bytes, fmt: str) -> bytes | None:
    """Converte bytes de áudio para 'ogg' (voz WhatsApp) ou 'mp3' (player universal) via ffmpeg.
    Grava a SAÍDA em arquivo temporário (não em pipe): assim o MP3 sai com o cabeçalho de
    duração (Xing/LAME). Sem ele o player do navegador não sabe o tamanho do áudio, a barra
    de progresso não funciona e parece que não toca completo — a operadora precisava baixar.
    Entrada também em arquivo (mp4 do iPhone não vai por pipe)."""
    fd_in, inpath = tempfile.mkstemp(suffix=".bin")
    fd_out, outpath = tempfile.mkstemp(suffix="." + fmt)
    os.close(fd_out)   # o ffmpeg escreve o arquivo de saída
    try:
        with os.fdopen(fd_in, "wb") as f:
            f.write(audio_bytes)
        if fmt == "ogg":
            args = ["ffmpeg", "-y", "-i", inpath, "-c:a", "libopus",
                    "-b:a", "32k", "-ac", "1", "-ar", "48000",
                    "-application", "voip", "-f", "ogg", outpath]
        else:  # mp3 — -write_xing 1 grava a duração no cabeçalho (player toca completo)
            args = ["ffmpeg", "-y", "-i", inpath, "-c:a", "libmp3lame",
                    "-b:a", "64k", "-ac", "1", "-write_xing", "1", "-f", "mp3", outpath]
        proc = subprocess.run(args, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=60)
        if proc.returncode == 0:
            with open(outpath, "rb") as f:
                dados = f.read()
            if dados:
                return dados
        print(f"⚠️ ffmpeg falhou ({fmt}) rc={proc.returncode}: {proc.stderr[:300]}")
    except FileNotFoundError:
        print("⚠️ ffmpeg não está instalado — usando áudio original")
    except Exception as e:
        print(f"⚠️ Erro no transcode de áudio ({fmt}): {e}")
    finally:
        for p in (inpath, outpath):
            try:
                os.remove(p)
            except Exception:
                pass
    return None


async def _transcode_audio(audio_bytes: bytes, fmt: str) -> bytes | None:
    """Versão async — roda o ffmpeg (bloqueante) numa thread para não travar o servidor."""
    return await asyncio.to_thread(_transcode_audio_sync, audio_bytes, fmt)


async def _salvar_imagem(url: str, db=None) -> str | None:
    """Baixa imagem com retries, guarda no banco (+disco) e retorna marcador [IMAGE:filename].
    Retorna None em caso de falha (caller deve salvar placeholder)."""
    for tentativa in range(3):
        try:
            async with httpx.AsyncClient(timeout=60) as client:
                resp = await client.get(url)
            if resp.status_code != 200 or len(resp.content) < 50:
                if tentativa < 2:
                    await asyncio.sleep(2)
                    continue
                return None
            ct = resp.headers.get("content-type", "image/jpeg")
            if "png" in ct:
                ext = "png"
            elif "webp" in ct:
                ext = "webp"
            elif "gif" in ct:
                ext = "gif"
            else:
                ext = "jpg"
            filename = f"{uuid.uuid4().hex}.{ext}"
            if db is not None:
                _guardar_blob(db, filename, "imagem", resp.content, mime=ct, subdir="imagens")
            else:
                os.makedirs("/app/imagens", exist_ok=True)
                with open(f"/app/imagens/{filename}", "wb") as f:
                    f.write(resp.content)
            return f"[IMAGE:{filename}]"
        except Exception as e:
            print(f"⚠️ Erro ao salvar imagem (tentativa {tentativa+1}): {e}")
            if tentativa < 2:
                await asyncio.sleep(2)
    return None


async def _salvar_documento(url: str, nome_original: str, db=None) -> str | None:
    """Baixa documento/PDF com retries, guarda no banco (+disco) e retorna marcador [DOC:filename|nome].
    Retorna None em caso de falha (caller deve salvar placeholder)."""
    nome_display = re.sub(r"[|]", "_", nome_original)[:100] if nome_original else "documento"
    for tentativa in range(3):
        try:
            async with httpx.AsyncClient(timeout=60) as client:
                resp = await client.get(url)
            if resp.status_code != 200 or len(resp.content) < 10:
                if tentativa < 2:
                    await asyncio.sleep(2)
                    continue
                return None
            ct = resp.headers.get("content-type", "")
            nome_lower = nome_original.lower()
            if "pdf" in ct or nome_lower.endswith(".pdf"):
                ext = "pdf"
            elif "." in nome_original:
                ext = nome_original.rsplit(".", 1)[-1][:5]
            else:
                ext = "bin"
            filename = f"{uuid.uuid4().hex}.{ext}"
            if db is not None:
                _guardar_blob(db, filename, "documento", resp.content,
                              nome_original=nome_display, mime=ct, subdir="documentos")
            else:
                os.makedirs("/app/documentos", exist_ok=True)
                with open(f"/app/documentos/{filename}", "wb") as f:
                    f.write(resp.content)
            return f"[DOC:{filename}|{nome_display}]"
        except Exception as e:
            print(f"⚠️ Erro ao salvar documento (tentativa {tentativa+1}): {e}")
            if tentativa < 2:
                await asyncio.sleep(2)
    return None


async def _salvar_audio_cliente(telefone: str, audio_url: str, db) -> str | None:
    """Baixa o áudio do cliente, salva em /app/audios/ e retorna o conteúdo [AUDIO:filename].
    Tenta até 3 vezes com timeout generoso. Se falhar, retorna marcador de áudio indisponível."""
    for tentativa in range(3):
        try:
            async with httpx.AsyncClient(timeout=60) as client:
                resp = await client.get(audio_url)
            if resp.status_code != 200:
                print(f"⚠️ Áudio cliente HTTP {resp.status_code} (tentativa {tentativa+1})")
                if tentativa < 2:
                    await asyncio.sleep(2)
                    continue
                return "[AUDIO_INDISPONIVEL]"
            if len(resp.content) < 100:
                print(f"⚠️ Áudio cliente vazio/muito pequeno ({len(resp.content)} bytes), tentativa {tentativa+1}")
                if tentativa < 2:
                    await asyncio.sleep(2)
                    continue
                return "[AUDIO_INDISPONIVEL]"
            # Converte para MP3 (toca em qualquer navegador, inclusive iPhone/Safari)
            audio_id = uuid.uuid4().hex
            mp3 = await _transcode_audio(resp.content, "mp3")
            if mp3:
                filename = f"{audio_id}.mp3"
                _guardar_blob(db, filename, "audio", mp3, mime="audio/mpeg", subdir="audios")
            else:
                # Fallback: guarda o original se o ffmpeg não estiver disponível
                ct = resp.headers.get("content-type", "audio/ogg")
                ext = "ogg" if "ogg" in ct else ("mp3" if "mp3" in ct or "mpeg" in ct else "webm")
                filename = f"{audio_id}.{ext}"
                _guardar_blob(db, filename, "audio", resp.content, mime=ct, subdir="audios")
            print(f"🎙️ Áudio do cliente salvo: {filename}")
            return f"[AUDIO:{filename}]"
        except Exception as e:
            print(f"⚠️ Erro ao salvar áudio do cliente (tentativa {tentativa+1}): {e}")
            if tentativa < 2:
                await asyncio.sleep(2)
    return "[AUDIO_INDISPONIVEL]"


_DEBUG_WEBHOOKS = []  # últimos webhooks recebidos (memória) — apenas para depuração


def _e_descadastro(texto: str) -> bool:
    """Detecta se o cliente pediu para parar de receber mensagens (opt-out)."""
    t = re.sub(r"[^\w\s]", "", (texto or "").strip().lower()).strip()
    if not t:
        return False
    exatas = {"sair", "parar", "pare", "para", "cancelar", "cancela", "stop",
              "descadastrar", "sai", "remover", "remove", "desinscrever"}
    if t in exatas:
        return True
    frases = ["nao quero receber", "não quero receber", "nao quero mais mensagem",
              "não quero mais mensagem", "para de mandar", "pare de mandar",
              "parem de mandar", "descadastr", "me tira da lista", "me tirem da lista",
              "nao me mande", "não me mande", "nao me envie", "não me envie",
              "nao quero mais contato", "não quero mais contato"]
    return any(f in t for f in frases)


def _ha_funcionaria_online(db) -> bool:
    """True se alguma funcionária está com sessão ativa (ativa nos últimos 10 min).
    Usado para NÃO mandar a mensagem de 'fora do horário' quando há alguém disponível."""
    limite = datetime.utcnow() - timedelta(seconds=600)
    return db.query(SessaoUsuario).join(Usuario, Usuario.id == SessaoUsuario.usuario_id).filter(
        Usuario.role == RoleEnum.funcionario,
        SessaoUsuario.logout_em.is_(None),
        SessaoUsuario.ultimo_ativo_em >= limite,
    ).first() is not None


def _humano_respondeu_recente(db, telefone: str, horas: int = 1) -> bool:
    """True se um humano (operador/admin) respondeu a ESTE cliente nas últimas `horas`.
    Mensagens de humano são salvas com prefixo '[Nome]: '; as do bot/automáticas não.
    Serve para PARAR de mandar 'fora do horário' quando alguém já está atendendo o cliente
    (ex.: admin trabalhando fora do horário) — por conversa, sem afetar clientes novos."""
    from models import MensagemConversa
    limite = datetime.utcnow() - timedelta(hours=horas)
    return db.query(MensagemConversa).filter(
        MensagemConversa.telefone == telefone,
        MensagemConversa.role == "assistant",
        MensagemConversa.conteudo.like("[%]: %"),
        MensagemConversa.criado_em >= limite,
    ).first() is not None


@app.get("/api/online")
async def usuarios_online(db: Session = Depends(get_db),
                          usuario: Usuario = Depends(obter_usuario_atual)):
    """Funcionárias online agora (ativas nos últimos 10 min). Admins não entram na lista.
    Visível para todos os logados."""
    limite = datetime.utcnow() - timedelta(seconds=600)
    sessoes = (db.query(SessaoUsuario)
               .join(Usuario, Usuario.id == SessaoUsuario.usuario_id)
               .filter(SessaoUsuario.logout_em.is_(None),
                       SessaoUsuario.ultimo_ativo_em >= limite,
                       Usuario.ativo == True,
                       Usuario.role == RoleEnum.funcionario)
               .all())
    vistos = {}
    for s in sessoes:
        u = s.usuario
        if not u:
            continue
        ant = vistos.get(u.id)
        if not ant or (s.ultimo_ativo_em and s.ultimo_ativo_em > ant["_ts"]):
            vistos[u.id] = {"id": u.id, "nome": u.nome, "role": u.role,
                            "_ts": s.ultimo_ativo_em or datetime.min}
    lista = sorted(vistos.values(), key=lambda x: (x["role"] != "admin", x["nome"].lower()))
    for x in lista:
        x.pop("_ts", None)
    return {"online": lista, "total": len(lista)}


@app.post("/webhook/zapi")
async def receber_webhook_zapi(request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    # Diagnóstico: guarda os últimos webhooks em memória (visível em /api/debug/webhooks)
    try:
        _DEBUG_WEBHOOKS.append({
            "em": _fmt_br(datetime.utcnow(), "%d/%m %H:%M:%S"),
            "type": body.get("type"),
            "notification": body.get("notification"),
            "phone": body.get("phone"),
            "fromMe": body.get("fromMe"),
            "body": body,
        })
        del _DEBUG_WEBHOOKS[:-20]
    except Exception:
        pass
    try:
        telefone = body.get("phone", "").replace("+", "").replace(" ", "")
        if not telefone:
            return JSONResponse({"status": "ignored"})

        # ── Mensagem editada pelo cliente ou pelo atendente ──────────────────
        if body.get("isEdit") or body.get("editedMessage"):
            print(f"📝 EDIT webhook recebido: {json.dumps(body, ensure_ascii=False)[:500]}")
            edited = body.get("editedMessage") or {}
            # Z-API aninha o conteúdo em editedMessage.message.*
            edited_msg = edited.get("message") or {}
            texto_editado = (
                # estrutura aninhada: editedMessage.message.conversation
                edited_msg.get("conversation", "")
                or (edited_msg.get("extendedTextMessage") or {}).get("text", "")
                # estrutura plana dentro de editedMessage
                or (edited.get("text") or {}).get("message", "")
                or (edited.get("extendedTextMessage") or {}).get("text", "")
                or edited.get("conversation", "")
                # fallback: texto direto no body (alguns formatos Z-API)
                or _extrair_texto_zapi(body)
            ).strip()
            print(f"📝 EDIT texto extraído: '{texto_editado}'")
            if telefone and texto_editado:
                lead = db.query(Lead).filter(Lead.telefone == telefone).first()
                if lead:
                    role = "assistant" if body.get("fromMe") else "user"
                    _salvar_msg_webhook(db, telefone, f"✏️ {texto_editado}", role=role)
                    print(f"📝 EDIT salvo para {telefone} ({role})")
            return JSONResponse({"status": "edit_saved"})

        # ── Ligação recebida pelo WhatsApp (Z-API: notification CALL_*) ───────
        notif = (body.get("notification") or "").upper()
        if notif.startswith("CALL"):
            perdida = "MISSED" in notif
            video = "VIDEO" in notif
            icone = "📹" if video else "📞"
            tipo_txt = ("Chamada de vídeo" if video else "Ligação")
            estado = "perdida" if perdida else "recebida"
            conteudo = f"{icone} {tipo_txt} {estado} pelo WhatsApp"
            eh_lid = ("@" in telefone) or ("lid" in telefone.lower())
            lead = db.query(Lead).filter(Lead.telefone == telefone).first()
            if not lead:
                # Número oculto (@lid): NÃO cria lead fantasma — evita duplicar o contato
                if eh_lid:
                    print(f"📞 Ligação de número oculto ({telefone}) — lead fantasma NÃO criado")
                    return JSONResponse({"status": "chamada_lid_ignorada"})
                nome_caller = (body.get("senderName") or body.get("chatName") or "").strip() or None
                lead = Lead(telefone=telefone, nome=nome_caller)
                db.add(lead)
                db.commit()
                db.refresh(lead)
            _salvar_msg_webhook(db, telefone, conteudo, role="user")
            lead.atualizado_em = datetime.utcnow()
            if lead.oculto_funil:
                lead.oculto_funil = False
            db.commit()
            print(f"{icone} {tipo_txt} {estado} de {telefone}")
            return JSONResponse({"status": "chamada_registrada"})

        # ── Mensagem enviada do próprio aparelho pelo atendente ──────────────
        if body.get("fromMe"):
            texto = _extrair_texto_zapi(body)
            if texto:
                # Ignora eco de mensagem já enviada pelo painel (evita texto duplicado/errado)
                if _e_duplicata_painel(telefone, texto):
                    return JSONResponse({"status": "fromme_ignored_duplicate"})
                lead = db.query(Lead).filter(Lead.telefone == telefone).first()
                if lead:
                    _salvar_msg_webhook(db, telefone, texto, role="assistant")
            else:
                # Imagem ou documento enviado do celular pelo atendente
                lead = db.query(Lead).filter(Lead.telefone == telefone).first()
                if lead:
                    img_info = _extrair_imagem_zapi(body)
                    if img_info:
                        conteudo = await _salvar_imagem(img_info["url"], db)
                        if conteudo:
                            _salvar_msg_webhook(db, telefone, conteudo, role="assistant")
                    else:
                        doc_info = _extrair_documento_zapi(body)
                        if doc_info:
                            conteudo = await _salvar_documento(doc_info["url"], doc_info["nome"], db)
                            if conteudo:
                                _salvar_msg_webhook(db, telefone, conteudo, role="assistant")
            return JSONResponse({"status": "fromme_saved"})

        # ── Mensagem recebida do cliente ─────────────────────────────────────
        texto = _extrair_texto_zapi(body)
        if not texto:
            lead_midia = db.query(Lead).filter(Lead.telefone == telefone).first()

            # Verifica áudio do cliente
            audio_url = _extrair_audio_url_zapi(body)
            if audio_url and lead_midia:
                conteudo_audio = await _salvar_audio_cliente(telefone, audio_url, db)
                # Sempre salva no histórico — mesmo se falhou, operadora precisa saber que havia um áudio
                msg_audio = conteudo_audio or "[AUDIO_INDISPONIVEL]"
                _salvar_msg_webhook(db, telefone, msg_audio, role="user")
                lead_midia.atualizado_em = datetime.utcnow()
                if lead_midia.oculto_funil:
                    lead_midia.oculto_funil = False
                db.commit()
                return JSONResponse({"status": "audio_salvo" if conteudo_audio else "audio_indisponivel"})

            # Verifica imagem do cliente
            img_info = _extrair_imagem_zapi(body)
            if img_info and lead_midia:
                conteudo_img = await _salvar_imagem(img_info["url"], db)
                # Sempre salva — placeholder se download falhou
                if conteudo_img:
                    if img_info.get("caption"):
                        conteudo_img += f"\n{img_info['caption']}"
                else:
                    legenda = img_info.get("caption", "")
                    conteudo_img = f"[IMAGEM_INDISPONIVEL]{(':' + legenda) if legenda else ''}"
                _salvar_msg_webhook(db, telefone, conteudo_img, role="user")
                lead_midia.atualizado_em = datetime.utcnow()
                if lead_midia.oculto_funil:
                    lead_midia.oculto_funil = False
                db.commit()
                return JSONResponse({"status": "imagem_salva" if "[IMAGE:" in conteudo_img else "imagem_indisponivel"})

            # Verifica documento/PDF do cliente
            doc_info = _extrair_documento_zapi(body)
            if doc_info and lead_midia:
                conteudo_doc = await _salvar_documento(doc_info["url"], doc_info["nome"], db)
                # Sempre salva — placeholder se download falhou
                if not conteudo_doc:
                    nome_safe = re.sub(r"[|]", "_", doc_info.get("nome", "documento"))[:100]
                    conteudo_doc = f"[DOC_INDISPONIVEL:{nome_safe}]"
                _salvar_msg_webhook(db, telefone, conteudo_doc, role="user")
                lead_midia.atualizado_em = datetime.utcnow()
                if lead_midia.oculto_funil:
                    lead_midia.oculto_funil = False
                db.commit()
                return JSONResponse({"status": "documento_salvo" if "[DOC:" in conteudo_doc else "documento_indisponivel"})

            return JSONResponse({"status": "ignored"})

        # ── Opt-out: cliente pediu para parar de receber mensagens ───────────
        if _e_descadastro(texto):
            lead_opt = db.query(Lead).filter(Lead.telefone == telefone).first()
            if lead_opt and not telefone.startswith("_manual_"):
                _salvar_msg_webhook(db, telefone, texto, role="user")
                lead_opt.descadastrado = True
                lead_opt.followup_tentativa = 99   # garante que não entra mais em follow-up
                lead_opt.atualizado_em = datetime.utcnow()
                db.commit()
                conf = ("Pronto! ✅ Você não receberá mais mensagens automáticas nossas. "
                        "Se mudar de ideia, é só nos chamar aqui a qualquer momento. Obrigado! 🙏")
                await enviar_zapi(telefone, conf)
                _salvar_msg_webhook(db, telefone, conf, role="assistant")
                print(f"🚫 Lead {telefone} descadastrado (opt-out)")
                return JSONResponse({"status": "descadastrado"})

        # ── Verifica se é número de parceiro ─────────────────────────────────
        parceiro = _buscar_parceiro_por_telefone(telefone, db)
        if parceiro:
            lead = db.query(Lead).filter(Lead.telefone == telefone).first()
            ja_em_atendimento = lead and lead.status in [
                StatusLeadEnum.assumido,
                StatusLeadEnum.proposta_enviada,
                StatusLeadEnum.proposta_aprovada,
                StatusLeadEnum.fechado,
                StatusLeadEnum.qualificado,
            ]
            if ja_em_atendimento:
                # Parceiro já transferido — só registra a mensagem (SEM resposta automática)
                _salvar_msg_webhook(db, telefone, texto, role="user")
                if lead.oculto_funil:
                    lead.oculto_funil = False
                    db.commit()
                return JSONResponse({"status": "parceiro_aguardando_humano"})

            # Primeiro contato do parceiro → transfere direto pra equipe, SEM mensagem automática
            if not lead:
                lead = Lead(telefone=telefone, nome=parceiro.nome, parceiro_id=parceiro.id)
                db.add(lead)
                db.commit()
                db.refresh(lead)

            _salvar_msg_webhook(db, telefone, texto, role="user")
            lead.estado_conversa = EstadoConversaEnum.transferido
            lead.status = StatusLeadEnum.qualificado
            lead.parceiro_id = parceiro.id
            if not lead.nome:
                lead.nome = parceiro.nome
            lead.atualizado_em = datetime.utcnow()
            lead.oculto_funil = False
            db.commit()

            await _notificar_equipe(telefone, db)
            return JSONResponse({"status": "parceiro_transferido"})

        lead = db.query(Lead).filter(Lead.telefone == telefone).first()

        # Lead perdido voltou a entrar em contato → reativa inteligentemente
        if lead and lead.status == StatusLeadEnum.perdido:
            tratado = await _reativar_lead_perdido(lead, texto, db, enviar_zapi)
            if tratado:
                await _notificar_equipe(telefone, db)
                return JSONResponse({"status": "reativado"})
            # tratado=False → reiniciou do zero, cai no fluxo normal abaixo

        # Se está sendo atendido por humano, salva a mensagem e avisa se fora do horário
        if lead and lead.status in [
            StatusLeadEnum.assumido,
            StatusLeadEnum.pre_analise,
            StatusLeadEnum.proposta_enviada,
            StatusLeadEnum.proposta_aprovada,
            StatusLeadEnum.fechado,
            StatusLeadEnum.parceiro,
        ]:
            # Cliente voltou → reexibe no funil se estava oculto
            if lead.oculto_funil:
                lead.oculto_funil = False
                db.commit()
            _salvar_msg_webhook(db, telefone, texto)
            prox = _proximo_horario_atendimento()
            # Só avisa "fora do horário" se NÃO houver funcionária online E ninguém já
            # estiver respondendo este cliente (admin/operador respondeu na última 1h)
            if prox and not _ha_funcionaria_online(db) and not _humano_respondeu_recente(db, telefone):
                nome = f" {lead.nome}" if lead.nome else ""
                aviso = (
                    f"Olá{nome}! 😊 No momento estamos fora do horário de atendimento. "
                    f"Funcionamos seg–sex das 09h às 18h e sábado das 09h às 13h. "
                    f"Retornaremos seu contato no primeiro horário disponível! 🕘"
                )
                await enviar_zapi(telefone, aviso)
                _salvar_msg_webhook(db, telefone, aviso, role="assistant")
            return JSONResponse({"status": "aguardando_humano"})

        # Bot qualificando — roda em thread separada p/ NÃO travar o servidor
        # (a chamada à IA é bloqueante; sem isso o app cai sob carga / IA lenta)
        respostas = await asyncio.to_thread(processar_mensagem, telefone, texto, db)
        for i, msg in enumerate(respostas):
            if i > 0:
                await asyncio.sleep(0.8)
            await enviar_zapi(telefone, msg)

        lead = db.query(Lead).filter(Lead.telefone == telefone).first()
        if lead and lead.status == StatusLeadEnum.qualificado:
            await _notificar_equipe(telefone, db)

    except Exception as e:
        print(f"⚠️ Erro webhook Z-API: {e}")
    return JSONResponse({"status": "ok"})


@app.get("/webhook/meta")
async def verificar_webhook_meta(
    hub_mode: str = Query(None, alias="hub.mode"),
    hub_verify_token: str = Query(None, alias="hub.verify_token"),
    hub_challenge: str = Query(None, alias="hub.challenge"),
):
    if hub_mode == "subscribe" and hub_verify_token == settings.WHATSAPP_VERIFY_TOKEN:
        return PlainTextResponse(hub_challenge)
    raise HTTPException(status_code=403, detail="Token inválido")


@app.post("/webhook/meta")
async def receber_webhook_meta(request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    try:
        entry = body["entry"][0]["changes"][0]["value"]
        for msg in entry.get("messages", []):
            if msg.get("type") != "text":
                continue
            telefone = msg["from"]
            texto = msg["text"]["body"].strip()
            if not texto:
                continue

            lead = db.query(Lead).filter(Lead.telefone == telefone).first()

            # Lead perdido voltou → reativa inteligentemente
            if lead and lead.status == StatusLeadEnum.perdido:
                tratado = await _reativar_lead_perdido(lead, texto, db, enviar_meta)
                if tratado:
                    await _notificar_equipe(telefone, db)
                    continue
                # tratado=False → reiniciou do zero, cai no fluxo normal

            # Humano atendendo → salva e avisa se fora do horário
            if lead and lead.status in [
                StatusLeadEnum.assumido,
                StatusLeadEnum.pre_analise,
                StatusLeadEnum.proposta_enviada,
                StatusLeadEnum.proposta_aprovada,
                StatusLeadEnum.fechado,
                StatusLeadEnum.parceiro,
            ]:
                _salvar_msg_webhook(db, telefone, texto)
                prox = _proximo_horario_atendimento()
                if prox and not _humano_respondeu_recente(db, telefone):
                    nome = f" {lead.nome}" if lead.nome else ""
                    aviso = (
                        f"Olá{nome}! 😊 No momento estamos fora do horário de atendimento. "
                        f"Funcionamos seg–sex das 09h às 18h e sábado das 09h às 13h. "
                        f"Retornaremos seu contato {prox}! 🕘"
                    )
                    await enviar_meta(telefone, aviso)
                    _salvar_msg_webhook(db, telefone, aviso, role="assistant")
                continue

            respostas = await asyncio.to_thread(processar_mensagem, telefone, texto, db)
            for i, msg in enumerate(respostas):
                if i > 0:
                    await asyncio.sleep(0.8)
                await enviar_meta(telefone, msg)
            lead = db.query(Lead).filter(Lead.telefone == telefone).first()
            if lead and lead.status == StatusLeadEnum.qualificado:
                await _notificar_equipe(telefone, db)
    except (KeyError, IndexError):
        pass
    return JSONResponse({"status": "ok"})


# ─── Teste ───────────────────────────────────────────────────────────────────────

@app.get("/api/diagnostico-ia")
async def diagnostico_ia_endpoint(usuario: Usuario = Depends(requer_admin)):
    """Testa a IA (Maria) e devolve o erro exato, se houver. Só admin."""
    return JSONResponse(diagnostico_ia())


@app.post("/testar")
async def testar_bot(request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    telefone = body.get("telefone", "5500000000000")
    mensagem = body.get("mensagem", "")
    if not mensagem:
        raise HTTPException(status_code=400, detail="Campo 'mensagem' obrigatório")
    respostas = processar_mensagem(telefone, mensagem, db)
    return JSONResponse({"respostas": respostas, "telefone": telefone})


# ─── API Parceiros ───────────────────────────────────────────────────────────────

def _serial_parceiro(p: Parceiro) -> dict:
    try:
        extras = json.loads(p.telefones_extras or "[]")
        if not isinstance(extras, list):
            extras = []
    except Exception:
        extras = []
    return {
        "id": p.id,
        "nome": p.nome,
        "data_nascimento": p.data_nascimento or "",
        "cpf": p.cpf or "",
        "telefone": p.telefone,
        "telefones_extras": extras,
        "email": p.email or "",
        "observacoes": p.observacoes or "",
        "nome_agenda":    p.nome_agenda or "",
        "operadora_id":   p.operadora_id,
        "operadora_nome": p.operadora.nome if p.operadora else "",
        "ativo": p.ativo,
        "criado_em": _fmt_br(p.criado_em, "%d/%m/%Y") or "",
        "contatos": [
            {"id": c.id, "nome": c.nome, "telefone": c.telefone or "",
             "email": c.email or "", "cargo": c.cargo or ""}
            for c in p.contatos
        ],
    }


@app.get("/api/parceiros")
async def listar_parceiros(
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    parceiros = db.query(Parceiro).filter(Parceiro.ativo == True)\
                  .order_by(Parceiro.nome).all()
    return [_serial_parceiro(p) for p in parceiros]


@app.post("/api/parceiros")
async def criar_parceiro(
    request: Request,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    body = await request.json()
    nome     = body.get("nome", "").strip()
    telefone = body.get("telefone", "").strip()
    cpf      = body.get("cpf", "").strip() or None

    if not nome:
        raise HTTPException(status_code=400, detail="O nome do parceiro é obrigatório")
    if not telefone:
        raise HTTPException(status_code=400, detail="O telefone é obrigatório")

    # Verificar duplicidade por CPF
    if cpf and db.query(Parceiro).filter(Parceiro.cpf == cpf).first():
        raise HTTPException(status_code=400, detail="Já existe um parceiro com este CPF")
    # Verificar duplicidade por telefone
    tel_norm = "".join(c for c in telefone if c.isdigit())
    existente = db.query(Parceiro).filter(Parceiro.telefone == tel_norm).first()
    if existente:
        raise HTTPException(status_code=400, detail=f"Já existe um parceiro com este telefone: {existente.nome}")

    extras = [t.strip() for t in body.get("telefones_extras", []) if t.strip()]
    operadora_id = body.get("operadora_id") or None
    if operadora_id:
        operadora_id = int(operadora_id)
    p = Parceiro(
        nome=nome,
        data_nascimento=body.get("data_nascimento", "").strip() or None,
        cpf=cpf,
        telefone=tel_norm or telefone,
        telefones_extras=json.dumps(extras, ensure_ascii=False) if extras else None,
        email=body.get("email", "").strip() or None,
        observacoes=body.get("observacoes", "").strip() or None,
        nome_agenda=body.get("nome_agenda", "").strip() or None,
        operadora_id=operadora_id,
    )
    db.add(p)
    db.flush()

    # Contatos adicionais
    for c in body.get("contatos", []):
        nome_c = c.get("nome", "").strip()
        if nome_c:
            db.add(ContatoParceiro(
                parceiro_id=p.id,
                nome=nome_c,
                telefone=c.get("telefone", "").strip() or None,
                email=c.get("email", "").strip() or None,
                cargo=c.get("cargo", "").strip() or None,
            ))

    db.commit()
    db.refresh(p)
    return _serial_parceiro(p)


@app.post("/api/parceiros/importar")
async def importar_parceiros(
    arquivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin: Usuario = Depends(requer_admin),
):
    """Importa parceiros em massa de um CSV (colunas: nome,telefone).
    Pula quem já existe (telefone comparado de forma robusta) e duplicados internos."""
    import csv as _csv
    import io as _io
    bruto = await arquivo.read()
    try:
        conteudo = bruto.decode("utf-8-sig")
    except UnicodeDecodeError:
        conteudo = bruto.decode("latin-1", errors="replace")
    linhas = list(_csv.reader(_io.StringIO(conteudo)))
    if not linhas:
        raise HTTPException(status_code=400, detail="Arquivo vazio ou ilegível.")

    # Detecta cabeçalho "nome,telefone"
    cab = [(c or "").strip().lower() for c in linhas[0]]
    i_nome, i_tel, inicio = 0, 1, 0
    if "nome" in cab and "telefone" in cab:
        i_nome, i_tel, inicio = cab.index("nome"), cab.index("telefone"), 1

    # Mapa telefone canônico -> parceiro já cadastrado (pra não duplicar e corrigir nomes quebrados)
    por_tel = {}
    for p in db.query(Parceiro).all():
        por_tel.setdefault(_tel_canonico(p.telefone), p)
        try:
            for e in json.loads(p.telefones_extras or "[]"):
                por_tel.setdefault(_tel_canonico(e), p)
        except Exception:
            pass

    criados = ja_existiam = atualizados = dup_arquivo = invalidos = 0
    vistos = set()
    novos = []
    for linha in linhas[inicio:]:
        if not linha or len(linha) <= max(i_nome, i_tel):
            invalidos += 1
            continue
        nome = (linha[i_nome] or "").strip()[:200]
        tel = (linha[i_tel] or "").strip()
        canon = _tel_canonico(tel)
        if not nome or len(canon) < 10:
            invalidos += 1
            continue
        if canon in vistos:
            dup_arquivo += 1
            continue
        vistos.add(canon)
        existente = por_tel.get(canon)
        if existente is not None:
            # Corrige nome quebrado (só "Pc") sem mexer nos nomes bons já curados
            if _nome_so_prefixo(existente.nome) and not _nome_so_prefixo(nome):
                existente.nome = nome
                if _nome_so_prefixo(existente.nome_agenda or ""):
                    existente.nome_agenda = nome
                atualizados += 1
            else:
                ja_existiam += 1
            continue
        tel_digitos = "".join(c for c in tel if c.isdigit())[:20]
        novos.append(Parceiro(
            nome=nome,
            telefone=tel_digitos or tel[:20],
            nome_agenda=nome,
            ativo=True,
        ))
        criados += 1

    if novos:
        db.add_all(novos)
    db.commit()  # commit sempre — pode haver só correções de nome
    return {
        "criados": criados,
        "atualizados": atualizados,
        "ja_existiam": ja_existiam,
        "duplicados_arquivo": dup_arquivo,
        "invalidos": invalidos,
        "total_linhas": len(linhas) - inicio,
    }


@app.post("/api/parceiros/transferir-carteira")
async def transferir_carteira_parceiros(request: Request, db: Session = Depends(get_db), admin: Usuario = Depends(requer_admin)):
    """Passa TODOS os parceiros de uma carteira (operadora de origem, ou 'sem carteira')
    para outra operadora ativa. Usado quando alguém é desligada e a carteira precisa ir
    para outra pessoa (ex.: parceiros da Camila → Luana). Só muda o responsável."""
    body = await request.json()
    de_id = body.get("de_id")      # int (operadora origem) ou None (sem carteira)
    para_id = body.get("para_id")
    if not para_id:
        raise HTTPException(status_code=400, detail="Escolha a operadora de destino.")
    para = db.query(Usuario).filter(Usuario.id == int(para_id), Usuario.ativo == True).first()
    if not para:
        raise HTTPException(status_code=400, detail="Operadora de destino inválida ou inativa.")
    if de_id is not None and int(de_id) == int(para_id):
        raise HTTPException(status_code=400, detail="A carteira de origem e destino são a mesma.")
    if de_id is None:
        parceiros = db.query(Parceiro).filter(Parceiro.operadora_id.is_(None)).all()
    else:
        parceiros = db.query(Parceiro).filter(Parceiro.operadora_id == int(de_id)).all()
    for p in parceiros:
        p.operadora_id = para.id
    db.commit()
    return {"transferidos": len(parceiros), "para": para.nome}


@app.put("/api/parceiros/{pid}")
async def atualizar_parceiro(
    pid: int,
    request: Request,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    p = db.query(Parceiro).filter(Parceiro.id == pid).first()
    if not p:
        raise HTTPException(status_code=404, detail="Parceiro não encontrado")

    body = await request.json()
    nome     = body.get("nome", "").strip()
    telefone = body.get("telefone", "").strip()
    cpf      = body.get("cpf", "").strip() or None

    if not nome:
        raise HTTPException(status_code=400, detail="O nome é obrigatório")
    if not telefone:
        raise HTTPException(status_code=400, detail="O telefone é obrigatório")

    # Duplicidade CPF (outro parceiro)
    if cpf and cpf != p.cpf:
        if db.query(Parceiro).filter(Parceiro.cpf == cpf, Parceiro.id != pid).first():
            raise HTTPException(status_code=400, detail="Já existe um parceiro com este CPF")

    tel_norm = "".join(c for c in telefone if c.isdigit()) or telefone
    if tel_norm != p.telefone:
        if db.query(Parceiro).filter(Parceiro.telefone == tel_norm, Parceiro.id != pid).first():
            raise HTTPException(status_code=400, detail="Já existe um parceiro com este telefone")

    extras = [t.strip() for t in body.get("telefones_extras", []) if t.strip()]
    op_id = body.get("operadora_id") or None
    p.nome             = nome
    p.data_nascimento  = body.get("data_nascimento", "").strip() or None
    p.cpf              = cpf
    p.telefone         = tel_norm
    p.telefones_extras = json.dumps(extras, ensure_ascii=False) if extras else None
    p.email            = body.get("email", "").strip() or None
    p.observacoes      = body.get("observacoes", "").strip() or None
    p.nome_agenda      = body.get("nome_agenda", "").strip() or None
    # Só admin pode alterar a operadora responsável
    if usuario.role == RoleEnum.admin:
        p.operadora_id = int(op_id) if op_id else None

    db.commit()
    db.refresh(p)
    return _serial_parceiro(p)


@app.delete("/api/parceiros/{pid}")
async def desativar_parceiro(
    pid: int,
    db: Session = Depends(get_db),
    admin: Usuario = Depends(requer_admin),
):
    p = db.query(Parceiro).filter(Parceiro.id == pid).first()
    if not p:
        raise HTTPException(status_code=404, detail="Parceiro não encontrado")
    p.ativo = False
    db.commit()
    return {"status": "ok"}


@app.post("/api/parceiros/{pid}/contatos")
async def adicionar_contato(
    pid: int,
    request: Request,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    p = db.query(Parceiro).filter(Parceiro.id == pid).first()
    if not p:
        raise HTTPException(status_code=404, detail="Parceiro não encontrado")
    body = await request.json()
    nome_c = body.get("nome", "").strip()
    if not nome_c:
        raise HTTPException(status_code=400, detail="Nome do contato é obrigatório")
    c = ContatoParceiro(
        parceiro_id=pid,
        nome=nome_c,
        telefone=body.get("telefone", "").strip() or None,
        email=body.get("email", "").strip() or None,
        cargo=body.get("cargo", "").strip() or None,
    )
    db.add(c)
    db.commit()
    db.refresh(p)
    return _serial_parceiro(p)


@app.delete("/api/parceiros/{pid}/contatos/{cid}")
async def remover_contato(
    pid: int,
    cid: int,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    c = db.query(ContatoParceiro).filter(
        ContatoParceiro.id == cid,
        ContatoParceiro.parceiro_id == pid,
    ).first()
    if not c:
        raise HTTPException(status_code=404, detail="Contato não encontrado")
    db.delete(c)
    db.commit()
    return {"status": "ok"}


# ─── API Leads ───────────────────────────────────────────────────────────────────

@app.post("/api/leads")
async def criar_lead_manual(
    request: Request,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    """Cria um lead manualmente (sem passar pelo bot do WhatsApp)."""
    import time
    body = await request.json()

    nome            = body.get("nome", "").strip()
    telefone        = body.get("telefone", "").strip()
    origem          = body.get("origem", "whatsapp").strip() or "whatsapp"
    origem_detalhe  = body.get("origem_detalhe", "").strip() or None
    parceiro_id     = body.get("parceiro_id") or None
    modalidade      = body.get("modalidade", ModalidadeEnum.indefinido)
    obs             = body.get("observacao", "").strip()
    atrib_id        = body.get("atribuido_para")

    if not nome:
        raise HTTPException(status_code=400, detail="O nome do lead é obrigatório")

    # Telefone é único — se não informado, gera placeholder
    if not telefone:
        telefone = f"_manual_{int(time.time() * 1000)}"
    else:
        telefone_norm = "".join(c for c in telefone if c.isdigit() or c == "+")
        if not telefone_norm:
            telefone_norm = telefone
        telefone = telefone_norm
        if db.query(Lead).filter(Lead.telefone == telefone).first():
            raise HTTPException(status_code=400, detail="Já existe um lead com este número de telefone")

    # Responsável
    responsavel = None
    if atrib_id:
        responsavel = db.query(Usuario).filter(Usuario.id == atrib_id, Usuario.ativo == True).first()

    # Parceiro
    if parceiro_id:
        p = db.query(Parceiro).filter(Parceiro.id == parceiro_id, Parceiro.ativo == True).first()
        if not p:
            parceiro_id = None

    lead = Lead(
        nome=nome,
        telefone=telefone,
        origem=origem,
        origem_detalhe=origem_detalhe,
        parceiro_id=int(parceiro_id) if parceiro_id else None,
        modalidade=modalidade,
        status=StatusLeadEnum.assumido,
        estado_conversa=EstadoConversaEnum.transferido,
        atribuido_para=responsavel.id if responsavel else usuario.id,
        assumido_em=datetime.utcnow(),
    )
    db.add(lead)
    db.flush()

    if obs:
        lista = []
        try:
            lista = json.loads(lead.observacoes or "[]")
        except Exception:
            pass
        lista.append({
            "texto": obs,
            "usuario": usuario.nome,
            "em": _agora_br().strftime("%d/%m/%Y %H:%M"),
        })
        lead.observacoes = json.dumps(lista, ensure_ascii=False)

    db.commit()
    db.refresh(lead)
    return _serial_lead(lead, db)


@app.get("/api/leads")
async def listar_leads(
    status: str = None,
    modalidade: str = None,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    query = db.query(Lead)
    if status:
        query = query.filter(Lead.status == status)
    if modalidade:
        query = query.filter(Lead.modalidade == modalidade)
    leads = query.order_by(Lead.criado_em.desc()).all()
    result = [_serial_lead(l, db) for l in leads]
    # Marca "sem próximo passo": lead ativo (assumido→proposta) SEM agendamento pendente.
    # Uma consulta só (não N+1) p/ o conjunto de leads com agendamento em aberto.
    com_pendente = {lid for (lid,) in db.query(Agendamento.lead_id)
                    .filter(Agendamento.concluido.isnot(True)).distinct().all()}
    _ativos = {StatusLeadEnum.assumido.value, StatusLeadEnum.pre_analise.value,
               StatusLeadEnum.proposta_enviada.value, StatusLeadEnum.proposta_aprovada.value}
    for r in result:
        r["sem_proximo_passo"] = (r["status"] in _ativos and r["id"] not in com_pendente)
    return result


@app.get("/api/leads/{lead_id}/conversa")
async def obter_conversa(
    lead_id: int,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead não encontrado")
    msgs = db.query(MensagemConversa).filter(
        MensagemConversa.telefone == lead.telefone
    ).order_by(MensagemConversa.id).all()
    mensagens = []
    for m in msgs:
        try:
            mensagens.append({
                "role": m.role,
                "conteudo": m.conteudo if m.conteudo is not None else "",
                "horario": _fmt_br(m.criado_em) or "",
            })
        except Exception:
            continue
    # Serializa o lead com proteção — nunca derruba a conversa
    try:
        lead_serial = _serial_lead(lead, db)
    except Exception as e:
        print(f"⚠️ Erro ao serializar lead {lead_id} na conversa: {e}")
        lead_serial = {"id": lead.id, "telefone": lead.telefone, "nome": lead.nome or "—",
                       "status": lead.status, "observacoes": [], "dados_contrato": {},
                       "carros_proposta": []}
    return {"lead": lead_serial, "mensagens": mensagens}


def _log_historico(db, lead_id: int, de_status, para_status, usuario_id, quando=None):
    """Registra uma transição de status na linha do tempo do lead (HistoricoLead).
    Best-effort: nunca deve quebrar o fluxo principal. Ignora no-ops (de == para).
    Deve ser chamado ANTES do db.commit() do endpoint p/ gravar na mesma transação.
    'quando' usa a data OFICIAL do evento (ex.: fechado_em = data do negócio) em vez do
    instante do clique — mantém os relatórios coerentes; default = agora."""
    try:
        de = de_status.value if hasattr(de_status, "value") else de_status
        para = para_status.value if hasattr(para_status, "value") else para_status
        if de == para:
            return
        db.add(HistoricoLead(lead_id=lead_id, de_status=de, para_status=para,
                             usuario_id=usuario_id, quando=quando or datetime.utcnow()))
    except Exception:
        pass


@app.post("/api/leads/{lead_id}/assumir")
async def assumir_lead(
    lead_id: int,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead não encontrado")
    if lead.status not in [StatusLeadEnum.qualificado, StatusLeadEnum.assumido]:
        raise HTTPException(status_code=400, detail="Lead não pode ser assumido neste status")
    de_status = lead.status
    lead.atribuido_para = usuario.id
    lead.status = StatusLeadEnum.assumido
    lead.assumido_em = datetime.utcnow()
    _log_historico(db, lead.id, de_status, StatusLeadEnum.assumido, usuario.id)
    db.commit()
    db.refresh(lead)
    return _serial_lead(lead, db)


@app.post("/api/leads/{lead_id}/mover")
async def mover_lead(
    lead_id: int,
    request: Request,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    """Move um lead para outro estágio do funil."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead não encontrado")

    body = await request.json()
    novo_status = body.get("status", "")

    estagios_validos = [
        StatusLeadEnum.em_atendimento,
        StatusLeadEnum.qualificado,
        StatusLeadEnum.assumido,
        StatusLeadEnum.pre_analise,
        StatusLeadEnum.proposta_enviada,
        StatusLeadEnum.proposta_aprovada,
        StatusLeadEnum.fechado,
        StatusLeadEnum.perdido,
        StatusLeadEnum.parceiro,
    ]
    if novo_status not in [s.value for s in estagios_validos]:
        raise HTTPException(status_code=400, detail="Estágio inválido")

    # Apenas administradores podem desqualificar ou devolver para Atendimento IA
    if usuario.role != RoleEnum.admin and novo_status in (
        StatusLeadEnum.desqualificado.value, StatusLeadEnum.em_atendimento.value
    ):
        raise HTTPException(status_code=403, detail="Apenas administradores podem realizar esta ação.")

    era_fechado = lead.status == StatusLeadEnum.fechado.value
    de_status = lead.status
    lead.status = novo_status
    if novo_status == StatusLeadEnum.assumido and not lead.atribuido_para:
        lead.atribuido_para = usuario.id
        lead.assumido_em = datetime.utcnow()
    # Marca a data de fechamento ao virar Fechado (estável p/ relatórios)
    if novo_status == StatusLeadEnum.fechado.value and not era_fechado and not lead.fechado_em:
        lead.fechado_em = datetime.utcnow()
    # Evento de fechamento usa a data OFICIAL (fechado_em); demais usam o instante do clique
    _quando_ev = lead.fechado_em if novo_status == StatusLeadEnum.fechado.value else None
    _log_historico(db, lead.id, de_status, novo_status, usuario.id, quando=_quando_ev)
    lead.atualizado_em = datetime.utcnow()
    db.commit()
    db.refresh(lead)
    return _serial_lead(lead, db)


@app.post("/api/leads/{lead_id}/fechar-contrato")
async def fechar_contrato_lead(
    lead_id: int,
    request: Request,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    """Salva os dados financeiros do contrato fechado e marca o lead como Fechado."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead não encontrado")

    body = await request.json()
    lead.deal_data     = body.get("deal_data", "").strip() or None
    lead.deal_veiculo  = body.get("deal_veiculo", "").strip() or None
    lead.deal_placa    = body.get("deal_placa", "").strip().upper() or None
    lead.deal_retorno  = body.get("deal_retorno", "").strip() or None
    lead.deal_valor    = body.get("deal_valor", "").strip() or None
    lead.deal_comissao = body.get("deal_comissao", "").strip() or None
    lead.deal_banco     = body.get("deal_banco", "").strip() or None
    lead.deal_conta_pg  = body.get("deal_conta_pg", "").strip() or None
    lead.deal_operadora = body.get("deal_operadora", "").strip() or None
    de_status          = lead.status
    lead.status        = StatusLeadEnum.fechado
    # Data estável de fechamento (usa a data do negócio se informada, senão agora)
    lead.fechado_em    = _data_br_para_utc(lead.deal_data) or lead.fechado_em or datetime.utcnow()
    lead.atualizado_em = datetime.utcnow()
    if not lead.atribuido_para:
        lead.atribuido_para = usuario.id
        lead.assumido_em    = datetime.utcnow()
    _log_historico(db, lead.id, de_status, StatusLeadEnum.fechado, usuario.id, quando=lead.fechado_em)
    db.commit()
    db.refresh(lead)
    return _serial_lead(lead, db)


def _transcricao_lead(db, lead, limite_msgs: int = 150) -> str:
    """Monta a transcrição (CLIENTE/ATENDIMENTO) das últimas mensagens do lead, em ordem
    cronológica, p/ alimentar a IA. Limita quantidade e tamanho p/ controlar custo."""
    msgs = (db.query(MensagemConversa)
            .filter(MensagemConversa.telefone == lead.telefone)
            .order_by(MensagemConversa.criado_em.desc())
            .limit(limite_msgs).all())
    linhas = []
    for m in reversed(msgs):   # volta à ordem cronológica
        c = (m.conteudo or "").strip()
        if not c:
            continue
        cu = c.upper()
        if cu.startswith("[IMAGE") or cu.startswith("[IMAGEM"):
            c = "[imagem]"
        elif cu.startswith("[AUDIO"):
            c = "[áudio]"
        elif cu.startswith("[DOC"):
            c = "[documento]"
        quem = "CLIENTE" if m.role == "user" else "ATENDIMENTO"
        linhas.append(f"{quem}: {c}")
    texto = "\n".join(linhas)
    return texto[-12000:] if len(texto) > 12000 else texto


@app.post("/api/leads/{lead_id}/resumo-ia")
async def resumo_ia(lead_id: int, db: Session = Depends(get_db),
                    usuario: Usuario = Depends(obter_usuario_atual)):
    """IA resume a conversa do lead p/ a operadora pegar o contexto em segundos."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(404, "Lead não encontrado")
    transcricao = _transcricao_lead(db, lead)
    if not transcricao.strip():
        return {"resumo": "Ainda não há conversa para resumir."}

    def _chamar():
        from bot import client, MODELO_IA
        prompt = (
            "Você é assistente INTERNO de uma equipe de F&I (financiamento de veículos). "
            "Resuma a conversa abaixo para a OPERADORA entender o caso em segundos. "
            "Máximo 4 linhas curtas, português, objetivo, sem saudação. Cubra: quem é o cliente e o que quer "
            "(veículo/modalidade), em que etapa está, o que já se sabe (renda, entrada, restrição) e qual o "
            "PRÓXIMO PASSO. Se faltar algo importante, diga o que falta.\n\n=== CONVERSA ===\n" + transcricao
        )
        resp = client.messages.create(model=MODELO_IA, max_tokens=350,
                                      messages=[{"role": "user", "content": prompt}])
        return resp.content[0].text.strip()

    try:
        resumo = await asyncio.to_thread(_chamar)
    except Exception as e:
        raise HTTPException(503, f"IA indisponível agora ({type(e).__name__}). Tente de novo.")
    return {"resumo": resumo}


@app.post("/api/leads/{lead_id}/extrair-dados-ia")
async def extrair_dados_ia(lead_id: int, db: Session = Depends(get_db),
                           usuario: Usuario = Depends(obter_usuario_atual)):
    """IA extrai os dados do cliente da conversa p/ a operadora só conferir e salvar."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(404, "Lead não encontrado")
    transcricao = _transcricao_lead(db, lead)
    if not transcricao.strip():
        return {"dados": {}}

    def _chamar():
        from bot import client, MODELO_IA
        prompt = (
            "Extraia os dados do CLIENTE a partir da conversa abaixo. Responda SOMENTE com um JSON válido "
            "(sem texto fora dele) com EXATAMENTE estas chaves, usando null quando o dado NÃO estiver claro "
            "na conversa (NUNCA invente):\n"
            '{"nome": null, "cpf": null, "data_nascimento": null, "carro_interesse": null, '
            '"modalidade": null, "cidade": null, "renda": null, "profissao": null, "email": null}\n'
            "Regras: data_nascimento no formato DD/MM/AAAA; modalidade só pode ser 'financiamento' ou "
            "'refinanciamento'; cpf apenas com os dígitos; renda deve ser EXATAMENTE uma destas faixas "
            "(ou null): 'Até R$2.000', 'R$2.000 a R$5.000', 'R$5.000 a R$10.000', 'Acima de R$10.000'.\n\n"
            "=== CONVERSA ===\n" + transcricao
        )
        resp = client.messages.create(model=MODELO_IA, max_tokens=400,
                                      messages=[{"role": "user", "content": prompt}])
        return resp.content[0].text.strip()

    try:
        raw = await asyncio.to_thread(_chamar)
    except Exception as e:
        raise HTTPException(503, f"IA indisponível agora ({type(e).__name__}). Tente de novo.")

    dados = {}
    try:
        i, j = raw.find("{"), raw.rfind("}")
        if i != -1 and j != -1:
            dados = json.loads(raw[i:j + 1])
    except Exception:
        dados = {}
    permitidas = {"nome", "cpf", "data_nascimento", "carro_interesse", "modalidade",
                  "cidade", "renda", "profissao", "email"}
    # Só aceita valores escalares (string/número), nunca dict/list/bool da IA fora do formato
    dados = {k: v for k, v in dados.items()
             if k in permitidas and isinstance(v, (str, int, float)) and not isinstance(v, bool)
             and str(v).strip() not in ("", "null", "None")}
    return {"dados": dados}


@app.post("/api/leads/{lead_id}/rascunho-ia")
async def rascunho_ia(lead_id: int, tipo: str = "auto", db: Session = Depends(get_db),
                      usuario: Usuario = Depends(obter_usuario_atual)):
    """IA escreve um RASCUNHO de mensagem (follow-up/proposta) p/ a operadora revisar e enviar."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(404, "Lead não encontrado")
    transcricao = _transcricao_lead(db, lead)
    primeiro_nome = ""
    if lead.nome and lead.nome != "—":
        primeiro_nome = lead.nome.strip().split(" ")[0]

    def _chamar():
        from bot import client, MODELO_IA
        instr = {
            "followup": "Escreva um FOLLOW-UP amigável para reengajar o cliente que parou de responder.",
            "proposta": "Escreva uma mensagem retomando/apresentando a PROPOSTA de financiamento de forma clara e convidativa.",
        }.get(tipo, "Escreva a PRÓXIMA mensagem natural para dar sequência ao atendimento — follow-up, "
                    "retomada ou próximo passo, o que fizer mais sentido pela conversa.")
        prompt = (
            "Você é a operadora da Fácil Financiamentos escrevendo uma mensagem de WhatsApp PARA O CLIENTE. "
            + instr + " Tom: cordial, brasileiro, direto e humano (como a equipe fala), sem parecer robô. "
            "1 a 3 frases curtas. "
            + (f"Chame o cliente pelo primeiro nome ('{primeiro_nome}'). " if primeiro_nome else "")
            + "NÃO invente valores, taxas ou condições que não apareceram na conversa. "
            "Responda SOMENTE com o texto da mensagem, sem aspas e sem explicação.\n\n"
            "=== CONVERSA ===\n" + (transcricao or "(ainda sem conversa)")
        )
        resp = client.messages.create(model=MODELO_IA, max_tokens=300,
                                      messages=[{"role": "user", "content": prompt}])
        return resp.content[0].text.strip()

    try:
        rascunho = await asyncio.to_thread(_chamar)
    except Exception as e:
        raise HTTPException(503, f"IA indisponível agora ({type(e).__name__}). Tente de novo.")
    return {"rascunho": rascunho}


# Cache em memória do briefing por (usuario, dia, período): gera no MÁX 2x/dia/operadora
_BRIEFING_CACHE = {}


@app.get("/api/briefing-dia")
async def briefing_dia(db: Session = Depends(get_db),
                       usuario: Usuario = Depends(obter_usuario_atual)):
    """Briefing do dia. Funcionária: o que priorizar (agendamentos + leads parados dela).
    Admin: visão de GESTÃO (panorama da equipe + seus compromissos). Gera no máx 2x/dia/usuário."""
    agora = datetime.utcnow()
    _br = _agora_br()
    hoje_br = _br.strftime("%Y-%m-%d")
    # Seus compromissos (agendamentos) — barato, usado por todos os papéis
    pares = (db.query(Agendamento, Lead).join(Lead, Agendamento.lead_id == Lead.id)
             .filter((Agendamento.criado_por == usuario.id) | (Lead.atribuido_para == usuario.id),
                     Agendamento.concluido.isnot(True)).all())
    atrasados = sum(1 for a, _l in pares if a.quando and a.quando <= agora)
    hoje = sum(1 for a, _l in pares
               if a.quando and a.quando > agora and _fmt_br(a.quando, "%Y-%m-%d") == hoje_br)
    nome = (usuario.nome or "").strip().split(" ")[0] or "tudo bem"

    # Cache por período: manhã (< 13h) e tarde (>= 13h); senão reusa
    periodo = "manha" if _br.hour < 13 else "tarde"
    chave = (usuario.id, hoje_br, periodo)
    if _BRIEFING_CACHE.get(chave):
        return {"briefing": _BRIEFING_CACHE[chave], "atrasados": atrasados, "hoje": hoje, "periodo": periodo}

    _ign = Lead.ignorar_relatorios.isnot(True)
    ATIVOS = [StatusLeadEnum.assumido.value, StatusLeadEnum.pre_analise.value,
              StatusLeadEnum.proposta_enviada.value, StatusLeadEnum.proposta_aprovada.value]
    com_pend = {lid for (lid,) in db.query(Agendamento.lead_id)
                .filter(Agendamento.concluido.isnot(True)).distinct().all()}

    if usuario.role == RoleEnum.admin:
        from sqlalchemy import func
        inicio_hoje = datetime(_br.year, _br.month, _br.day, tzinfo=_TZ_BR).astimezone(timezone.utc).replace(tzinfo=None)
        novos_hoje = db.query(Lead).filter(Lead.criado_em >= inicio_hoje, _ign).count()
        qualificados = db.query(Lead).filter(Lead.status == StatusLeadEnum.qualificado.value, _ign).count()
        em_atend = db.query(Lead).filter(Lead.status.in_(ATIVOS), _ign).count()
        fechados_hoje = db.query(Lead).filter(Lead.status == StatusLeadEnum.fechado.value,
                                              Lead.fechado_em >= inicio_hoje, _ign).count()
        ativos_atrib = db.query(Lead.id).filter(Lead.status.in_(ATIVOS),
                                                Lead.atribuido_para.isnot(None), _ign).all()
        sem_passo_eq = sum(1 for (lid,) in ativos_atrib if lid not in com_pend)
        # clientes esperando resposta (última mensagem é do cliente)
        tels = [t for (t,) in db.query(Lead.telefone)
                .filter(Lead.status.in_([StatusLeadEnum.qualificado.value] + ATIVOS), _ign).all()]
        paradas = 0
        if tels:
            sub = (db.query(MensagemConversa.telefone, func.max(MensagemConversa.id).label("mid"))
                   .filter(MensagemConversa.telefone.in_(tels))
                   .group_by(MensagemConversa.telefone).subquery())
            ultimas = db.query(MensagemConversa.role).join(sub, MensagemConversa.id == sub.c.mid).all()
            paradas = sum(1 for (role,) in ultimas if role == "user")
        fatos = (f"- {novos_hoje} novo(s) lead(s) hoje\n"
                 f"- {qualificados} qualificado(s) aguardando atendimento\n"
                 f"- {em_atend} em atendimento (assumidos/proposta)\n"
                 f"- {paradas} cliente(s) esperando resposta agora\n"
                 f"- {sem_passo_eq} lead(s) da equipe sem próximo passo\n"
                 f"- {fechados_hoje} contrato(s) fechado(s) hoje\n"
                 f"- seus compromissos: {atrasados} atrasado(s), {hoje} para hoje")
        prompt = (
            f"Você é assistente de GESTÃO de uma equipe de F&I (financiamento de veículos). Escreva um "
            f"briefing curto e estratégico para a GESTORA {nome} (dona da operação), com base nos números de "
            f"HOJE. Comece com 'Bom dia, {nome}!'. 3 a 5 linhas, tom de gestão: dê o panorama e aponte o que "
            f"merece atenção (gargalos como clientes esperando e leads sem próximo passo); destaque o que está "
            f"bom. Não use asteriscos. Responda só com o texto.\n\n" + fatos)
        fallback = (f"Bom dia, {nome}! Hoje: {novos_hoje} novos leads, {qualificados} qualificados aguardando, "
                    f"{paradas} clientes esperando resposta, {sem_passo_eq} leads da equipe sem próximo passo "
                    f"e {fechados_hoje} fechado(s).")
    else:
        meus_ativos = db.query(Lead.id).filter(Lead.status.in_(ATIVOS),
                                               Lead.atribuido_para == usuario.id, _ign).all()
        sem_passo = sum(1 for (lid,) in meus_ativos if lid not in com_pend)
        fatos = (f"- {atrasados} agendamento(s) atrasado(s)\n"
                 f"- {hoje} agendamento(s) para hoje\n"
                 f"- {sem_passo} lead(s) sem próximo passo agendado")
        prompt = (
            f"Escreva um BRIEFING matinal curto e motivador para {nome}, operadora de uma equipe de "
            f"financiamento de veículos, com base nestes números. Comece com 'Bom dia, {nome}!'. 2 a 4 linhas, "
            f"tom amigável e prático, português. Aponte o que priorizar hoje. Se estiver tudo zerado, parabenize "
            f"por estar em dia. Não use asteriscos. Responda só com o texto.\n\n" + fatos)
        fallback = (f"Bom dia, {nome}! Hoje você tem {atrasados} agendamento(s) atrasado(s), "
                    f"{hoje} para hoje e {sem_passo} lead(s) sem próximo passo.")

    def _chamar():
        from bot import client, MODELO_IA
        resp = client.messages.create(model=MODELO_IA, max_tokens=300,
                                      messages=[{"role": "user", "content": prompt}])
        return resp.content[0].text.strip()

    try:
        texto = await asyncio.to_thread(_chamar)
        for _k in [k for k in _BRIEFING_CACHE if k[1] != hoje_br]:   # limpa dias antigos
            _BRIEFING_CACHE.pop(_k, None)
        _BRIEFING_CACHE[chave] = texto                                # guarda só quando a IA respondeu
    except Exception:
        texto = fallback
    return {"briefing": texto, "atrasados": atrasados, "hoje": hoje, "periodo": periodo}


@app.get("/api/pendencias")
async def pendencias(db: Session = Depends(get_db),
                     usuario: Usuario = Depends(obter_usuario_atual)):
    """Só os contadores de pendências da operadora (SEM IA) — p/ a faixa compacta no funil."""
    agora = datetime.utcnow()
    hoje_br = _agora_br().strftime("%Y-%m-%d")
    pares = (db.query(Agendamento, Lead).join(Lead, Agendamento.lead_id == Lead.id)
             .filter((Agendamento.criado_por == usuario.id) | (Lead.atribuido_para == usuario.id),
                     Agendamento.concluido.isnot(True)).all())
    atrasados = sum(1 for a, _l in pares if a.quando and a.quando <= agora)
    hoje = sum(1 for a, _l in pares
               if a.quando and a.quando > agora and _fmt_br(a.quando, "%Y-%m-%d") == hoje_br)
    ATIVOS = [StatusLeadEnum.assumido.value, StatusLeadEnum.pre_analise.value,
              StatusLeadEnum.proposta_enviada.value, StatusLeadEnum.proposta_aprovada.value]
    meus_ativos = db.query(Lead.id).filter(Lead.status.in_(ATIVOS),
                                           Lead.atribuido_para == usuario.id,
                                           Lead.ignorar_relatorios.isnot(True)).all()
    com_pend = {lid for (lid,) in db.query(Agendamento.lead_id)
                .filter(Agendamento.concluido.isnot(True)).distinct().all()}
    sem_passo = sum(1 for (lid,) in meus_ativos if lid not in com_pend)
    return {"atrasados": atrasados, "hoje": hoje, "sem_passo": sem_passo}


def _calc_idade(data_nasc_str: str | None) -> str:
    """Calcula idade a partir de DD/MM/YYYY."""
    if not data_nasc_str or data_nasc_str == "—":
        return "—"
    try:
        from datetime import date
        partes = data_nasc_str.strip().split("/")
        if len(partes) != 3:
            return "—"
        nasc = date(int(partes[2]), int(partes[1]), int(partes[0]))
        hoje = date.today()
        idade = hoje.year - nasc.year - ((hoje.month, hoje.day) < (nasc.month, nasc.day))
        return str(idade)
    except Exception:
        return "—"


def _origem_label(origem: str | None, detalhe: str | None) -> str:
    mapa = {
        "rede_social": "Rede Social",
        "parceiro": "Parceiro",
        "ex_cliente": "Ex-cliente",
        "indicacao": "Indicação",
        "whatsapp": "WhatsApp",
    }
    base = mapa.get(origem or "", origem or "WhatsApp")
    if detalhe:
        return f"{base} ({detalhe})"
    return base


def _contratos_periodo(db, periodo: str = "mes") -> list:
    """Retorna leads fechados do período. periodo = 'semana' | 'mes' | 'tudo'.
    Usa fechado_em (data estável) — não atualizado_em, que muda a cada edição."""
    hoje = _agora_br()
    if periodo == "semana":
        dia_semana = hoje.weekday()  # 0=seg
        inicio_br = datetime(hoje.year, hoje.month, hoje.day, tzinfo=_TZ_BR) - timedelta(days=dia_semana)
        inicio = inicio_br.astimezone(timezone.utc).replace(tzinfo=None)
    elif periodo == "mes":
        inicio = datetime(hoje.year, hoje.month, 1, tzinfo=_TZ_BR).astimezone(timezone.utc).replace(tzinfo=None)
    else:
        inicio = None

    q = db.query(Lead).filter(Lead.status == StatusLeadEnum.fechado)
    if inicio:
        q = q.filter(Lead.fechado_em >= inicio)
    return q.order_by(Lead.fechado_em.desc().nullslast()).all()


def _inicio_periodo(periodo: str):
    """Início (UTC naive) do período p/ filtros de relatório, ou None p/ 'tudo'.
    periodo = 'semana' | 'mes' | '7dias' | '30dias' | 'tudo'."""
    hoje = _agora_br()
    if periodo == "semana":
        ini_br = datetime(hoje.year, hoje.month, hoje.day, tzinfo=_TZ_BR) - timedelta(days=hoje.weekday())
        return ini_br.astimezone(timezone.utc).replace(tzinfo=None)
    if periodo == "mes":
        return datetime(hoje.year, hoje.month, 1, tzinfo=_TZ_BR).astimezone(timezone.utc).replace(tzinfo=None)
    if periodo == "7dias":
        return datetime.utcnow() - timedelta(days=7)
    if periodo == "30dias":
        return datetime.utcnow() - timedelta(days=30)
    return None


@app.get("/api/relatorio/contratos-mes")
async def relatorio_contratos_mes(
    periodo: str = "mes",
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    """Retorna contagem e metas do período. Totais financeiros só para admin."""
    leads_fechados = _contratos_periodo(db, periodo)

    hoje = datetime.utcnow()
    total = len(leads_fechados)
    cfg_meta = db.query(Configuracao).filter(Configuracao.chave == "meta_contratos").first()
    meta = int(cfg_meta.valor) if cfg_meta and cfg_meta.valor.isdigit() else 20
    percentual = round(total / meta * 100) if meta > 0 else 0

    resultado: dict = {
        "total":      total,
        "meta":       meta,
        "percentual": min(percentual, 100),
        "mes":        _agora_br().strftime("%B %Y"),
        "periodo":    periodo,
    }

    if usuario.role == RoleEnum.admin:
        def _to_float(v):
            try:
                return float(str(v).replace("R$","").replace(".","").replace(",",".").strip())
            except Exception:
                return 0.0

        total_valor    = sum(_to_float(l.deal_valor)    for l in leads_fechados if l.deal_valor)
        total_comissao = sum(_to_float(l.deal_comissao) for l in leads_fechados if l.deal_comissao)

        resultado["total_valor"]    = total_valor
        resultado["total_comissao"] = total_comissao
        resultado["contratos"] = [
            {
                "id":            l.id,
                "nome":          l.nome or "—",
                "data_nascimento": l.data_nascimento or "—",
                "idade":         _calc_idade(l.data_nascimento),
                "telefone":      l.telefone,
                "deal_data":     l.deal_data or "—",
                "deal_veiculo":  l.deal_veiculo or "—",
                "deal_retorno":  l.deal_retorno or "—",
                "deal_valor":    l.deal_valor or "—",
                "deal_comissao": l.deal_comissao or "—",
                "deal_banco":    l.deal_banco or "—",
                "deal_conta_pg": l.deal_conta_pg or "—",
                "deal_operadora": l.deal_operadora or (l.responsavel.nome if l.responsavel else "—"),
                "responsavel":   l.responsavel.nome if l.responsavel else "—",
                "modalidade":    l.modalidade,
                "origem":        _origem_label(l.origem, l.origem_detalhe),
            }
            for l in leads_fechados
        ]

    return resultado


@app.get("/api/relatorio/contratos/csv")
async def relatorio_contratos_csv(
    periodo: str = "mes",
    db: Session = Depends(get_db),
    admin: Usuario = Depends(requer_admin),
):
    """Exporta contratos fechados do período como CSV (admin only)."""
    from fastapi.responses import StreamingResponse
    import csv, io
    leads = _contratos_periodo(db, periodo)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Cliente", "Idade", "Data", "Veículo", "Retorno",
                     "Valor Financiado", "Comissão", "Banco", "Conta PG",
                     "Origem", "Operadora Responsável"])
    for l in leads:
        writer.writerow([
            l.nome or "—",
            _calc_idade(l.data_nascimento),
            l.deal_data or "—",
            l.deal_veiculo or "—",
            l.deal_retorno or "—",
            l.deal_valor or "—",
            l.deal_comissao or "—",
            l.deal_banco or "—",
            l.deal_conta_pg or "—",
            _origem_label(l.origem, l.origem_detalhe),
            l.deal_operadora or (l.responsavel.nome if l.responsavel else "—"),
        ])
    output.seek(0)
    nome_arquivo = f"contratos_{periodo}.csv"
    headers = {"Content-Disposition": f"attachment; filename={nome_arquivo}"}
    return StreamingResponse(iter([output.getvalue()]),
                             media_type="text/csv; charset=utf-8-sig", headers=headers)


# ─── Analytics / Perfil para Anúncios ────────────────────────────────────────────

@app.get("/api/relatorio/perfil")
async def relatorio_perfil(
    db: Session = Depends(get_db),
    admin: Usuario = Depends(requer_admin),
):
    """Painel de perfil de clientes para direcionar anúncios (admin only)."""
    from collections import Counter
    import re as _re

    fechados  = db.query(Lead).filter(Lead.status == StatusLeadEnum.fechado).all()
    todos     = db.query(Lead).all()

    # ── Faixa etária ─────────────────────────────────────────────────────────
    faixas = {"Até 25": 0, "26-35": 0, "36-45": 0, "46-55": 0, "56+": 0, "N/I": 0}
    for l in fechados:
        idade_str = _calc_idade(l.data_nascimento)
        if idade_str == "—" or not idade_str.isdigit():
            faixas["N/I"] += 1
        else:
            i = int(idade_str)
            if   i <= 25: faixas["Até 25"] += 1
            elif i <= 35: faixas["26-35"]  += 1
            elif i <= 45: faixas["36-45"]  += 1
            elif i <= 55: faixas["46-55"]  += 1
            else:         faixas["56+"]    += 1

    # ── Renda ─────────────────────────────────────────────────────────────────
    rendas = Counter(l.renda for l in fechados if l.renda)

    # ── Profissão (top 8) ─────────────────────────────────────────────────────
    profissoes = Counter(l.profissao.strip().title() for l in fechados if l.profissao).most_common(8)

    # ── Modalidade ────────────────────────────────────────────────────────────
    modalidades = Counter(l.modalidade for l in fechados)

    # ── Origem com taxa de conversão ──────────────────────────────────────────
    origem_total   = Counter(l.origem or "whatsapp" for l in todos)
    origem_fechado = Counter(l.origem or "whatsapp" for l in fechados)
    origens_conv = []
    for orig, total in sorted(origem_total.items(), key=lambda x: -x[1]):
        fechou = origem_fechado.get(orig, 0)
        taxa = round(fechou / total * 100) if total else 0
        origens_conv.append({"origem": orig, "total": total, "fechados": fechou, "taxa": taxa})

    # ── Top veículos (top 8) ──────────────────────────────────────────────────
    veiculos_raw = [l.deal_veiculo or l.carro_interesse for l in fechados if l.deal_veiculo or l.carro_interesse]
    # Extrai marca (primeira palavra)
    marcas = Counter()
    for v in veiculos_raw:
        marca = v.strip().split()[0].upper() if v.strip() else "?"
        marcas[marca] += 1
    top_veiculos = marcas.most_common(8)

    # ── Cidade (top 8) ────────────────────────────────────────────────────────
    cidades = Counter(l.cidade.strip().title() for l in fechados if l.cidade).most_common(8)

    # ── Dia da semana / hora (horário BR, separado por origem) ────────────────
    dias = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
    dia_count = Counter(); hora_count = Counter()
    dia_bot = Counter(); hora_bot = Counter()
    dia_out = Counter(); hora_out = Counter()
    for l in todos:
        if l.criado_em:
            br = l.criado_em.replace(tzinfo=timezone.utc).astimezone(_TZ_BR)
            wd = dias[br.weekday()]; hr = br.hour
            dia_count[wd] += 1; hora_count[hr] += 1
            if (l.origem or "") in ("", "whatsapp"):   # bot
                dia_bot[wd] += 1; hora_bot[hr] += 1
            else:                                       # parceiro / inserido manualmente
                dia_out[wd] += 1; hora_out[hr] += 1

    # ── Valor médio financiado ────────────────────────────────────────────────
    def _to_float(v):
        try:
            return float(str(v).replace("R$","").replace(".","").replace(",",".").strip())
        except Exception:
            return 0.0
    valores = [_to_float(l.deal_valor) for l in fechados if l.deal_valor]
    valor_medio = round(sum(valores) / len(valores)) if valores else 0

    # ── Tempo médio até fechar (dias) ─────────────────────────────────────────
    tempos = []
    for l in fechados:
        if l.criado_em and l.atualizado_em:
            dias_delta = (l.atualizado_em - l.criado_em).days
            if 0 <= dias_delta <= 365:
                tempos.append(dias_delta)
    tempo_medio = round(sum(tempos) / len(tempos), 1) if tempos else 0

    return {
        "total_fechados": len(fechados),
        "total_leads":    len(todos),
        "valor_medio":    valor_medio,
        "tempo_medio_dias": tempo_medio,
        "faixas_etarias": faixas,
        "rendas":         dict(rendas),
        "profissoes":     profissoes,
        "modalidades":    dict(modalidades),
        "origens":        origens_conv,
        "top_veiculos":   top_veiculos,
        "cidades":        cidades,
        "dias_semana":        {d: dia_count.get(d, 0) for d in dias},
        "dias_semana_bot":    {d: dia_bot.get(d, 0) for d in dias},
        "dias_semana_outros": {d: dia_out.get(d, 0) for d in dias},
        "horas":          {str(h): hora_count.get(h, 0) for h in range(0, 24)},
        "horas_bot":      {str(h): hora_bot.get(h, 0) for h in range(0, 24)},
        "horas_outros":   {str(h): hora_out.get(h, 0) for h in range(0, 24)},
    }


@app.post("/api/conversa/iniciar")
async def iniciar_conversa(
    request: Request,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    """Funcionário inicia uma nova conversa com um número de WhatsApp."""
    body = await request.json()
    telefone = re.sub(r"\D", "", body.get("telefone", ""))
    texto = body.get("mensagem", "").strip()

    if not telefone or len(telefone) < 10:
        raise HTTPException(status_code=400, detail="Número de telefone inválido")
    if not texto:
        raise HTTPException(status_code=400, detail="Mensagem não pode ser vazia")

    # Busca ou cria o lead
    lead = db.query(Lead).filter(Lead.telefone == telefone).first()
    if not lead:
        lead = Lead(
            telefone=telefone,
            status=StatusLeadEnum.assumido,
            atribuido_para=usuario.id,
            assumido_em=datetime.utcnow(),
            estado_conversa="transferido",
        )
        db.add(lead)
        db.commit()
        db.refresh(lead)
    elif lead.status not in [StatusLeadEnum.assumido, StatusLeadEnum.pre_analise, StatusLeadEnum.proposta_enviada, StatusLeadEnum.proposta_aprovada]:
        lead.status = StatusLeadEnum.assumido
        lead.estado_conversa = EstadoConversaEnum.transferido
        lead.atribuido_para = usuario.id
        lead.assumido_em = datetime.utcnow()
        lead.atualizado_em = datetime.utcnow()
        db.commit()
    else:
        # Lead já em atendimento humano — garante estado_conversa correto
        if lead.estado_conversa not in (EstadoConversaEnum.transferido, EstadoConversaEnum.finalizado):
            lead.estado_conversa = EstadoConversaEnum.transferido
            lead.atualizado_em = datetime.utcnow()
            db.commit()

    # Salva e envia (registra no cache para ignorar o eco fromMe do Z-API)
    _registrar_msg_painel(telefone, texto)
    _salvar_msg_webhook(db, telefone, f"[{usuario.nome}]: {texto}", role="assistant")
    await enviar_zapi(telefone, texto)

    return _serial_lead(lead, db)


@app.post("/api/leads/{lead_id}/mensagem")
async def enviar_mensagem_funcionario(
    lead_id: int,
    request: Request,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    """Funcionário envia mensagem para o cliente pelo dashboard."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead não encontrado")

    body = await request.json()
    texto = body.get("mensagem", "").strip()
    if not texto:
        raise HTTPException(status_code=400, detail="Mensagem não pode ser vazia")

    # Se funcionário responde, assume o lead automaticamente
    if lead.status in [StatusLeadEnum.em_atendimento, StatusLeadEnum.qualificado]:
        lead.status = StatusLeadEnum.assumido
        lead.atribuido_para = usuario.id
        lead.assumido_em = datetime.utcnow()
        lead.atualizado_em = datetime.utcnow()
        db.commit()

    # Salva no histórico como mensagem do atendente (registra no cache para ignorar eco fromMe)
    _registrar_msg_painel(lead.telefone, texto)
    _salvar_msg_webhook(db, lead.telefone, f"[{usuario.nome}]: {texto}", role="assistant")

    # Envia pelo WhatsApp
    await enviar_zapi(lead.telefone, texto)

    return {"status": "enviado"}


@app.post("/api/leads/{lead_id}/reativar-funil")
async def reativar_funil(
    lead_id: int,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    """Traz o lead de volta ao funil (chamado ao abrir o lead na lista)."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404)
    if lead.oculto_funil:
        lead.oculto_funil = False
        lead.atualizado_em = datetime.utcnow()
        db.commit()
    return {"status": "ok"}


@app.post("/api/leads/{lead_id}/enviar-audio")
async def enviar_audio_gravado(
    lead_id: int,
    request: Request,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    """Atendente envia áudio gravado no navegador para o cliente."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead não encontrado")

    body = await request.json()
    audio_base64_raw = body.get("audio_base64", "")
    if not audio_base64_raw:
        raise HTTPException(status_code=400, detail="Áudio não recebido")

    # Extrai MIME e bytes
    mime = "audio/webm"
    raw_b64 = audio_base64_raw
    if "," in audio_base64_raw:
        header, raw_b64 = audio_base64_raw.split(",", 1)
        mime = header.split(":")[1].split(";")[0] if ":" in header else mime

    audio_bytes = base64.b64decode(raw_b64)
    print(f"🎤 Áudio para {lead.telefone} | mime={mime} | bytes={len(audio_bytes)}")

    # Converte ogg/opus (voz WhatsApp) e mp3 (player do painel) EM PARALELO — não soma o tempo dos dois
    ogg, mp3 = await asyncio.gather(
        _transcode_audio(audio_bytes, "ogg"),
        _transcode_audio(audio_bytes, "mp3"),
    )

    # O que será enviado ao WhatsApp
    if ogg:
        envio_mime, envio_b64 = "audio/ogg", base64.b64encode(ogg).decode()
    else:
        envio_mime, envio_b64 = mime, raw_b64  # fallback: original

    # O que fica guardado para reprodução no painel
    audio_id = uuid.uuid4().hex
    if mp3:
        audio_filename = f"{audio_id}.mp3"
        _guardar_blob(db, audio_filename, "audio", mp3, mime="audio/mpeg", subdir="audios")
    else:
        ext = "ogg" if ogg else ("ogg" if "ogg" in mime.lower() else "webm")
        audio_filename = f"{audio_id}.{ext}"
        _guardar_blob(db, audio_filename, "audio", (ogg or audio_bytes),
                      mime=("audio/ogg" if ogg else mime), subdir="audios")

    # Envia pelo Z-API com base64 diretamente (evita race condition de download de URL)
    zapi_ok = False
    zapi_erro = ""
    if settings.ZAPI_INSTANCE and settings.ZAPI_TOKEN:
        zapi_url = f"https://api.z-api.io/instances/{settings.ZAPI_INSTANCE}/token/{settings.ZAPI_TOKEN}/send-audio"
        headers_zapi = {"Client-Token": settings.ZAPI_CLIENT_TOKEN}
        # Monta data URI com o áudio convertido (ogg/opus) para virar nota de voz
        audio_data_uri = f"data:{envio_mime};base64,{envio_b64}"
        payload = {"phone": lead.telefone, "audio": audio_data_uri}
        try:
            async with httpx.AsyncClient() as client:
                resp = await client.post(zapi_url, headers=headers_zapi, json=payload, timeout=30)
            print(f"🎤 Z-API resposta: {resp.status_code} — {resp.text[:300]}")
            zapi_ok = resp.status_code == 200
            if not zapi_ok:
                zapi_erro = f"Z-API respondeu {resp.status_code}: {resp.text[:140]}"
                print(f"⚠️ {zapi_erro}")
        except Exception as e:
            zapi_erro = f"sem resposta do Z-API ({type(e).__name__}: {str(e)[:80]})"
            print(f"⚠️ Erro ao chamar Z-API: {e}")
    else:
        zapi_ok = True
        print(f"[Z-API SIMULADO] Áudio para {lead.telefone}")

    # Salva no histórico com referência ao arquivo (para reprodução no painel)
    # Mesmo se Z-API falhou, o arquivo está salvo — o painel pode reproduzir
    _salvar_msg_webhook(db, lead.telefone, f"[{usuario.nome}]: [AUDIO:{audio_filename}]", role="assistant")

    if not zapi_ok:
        raise HTTPException(status_code=502, detail=f"Não entregou pelo WhatsApp ({zapi_erro or 'falha desconhecida'}). O áudio ficou salvo no painel — tente reenviar.")

    return {"status": "enviado"}


def _buscar_midia(db, filename: str):
    """Retorna (bytes, mime, nome_original) da mídia — banco primeiro, disco como fallback."""
    reg = db.query(MidiaArquivo).filter(MidiaArquivo.filename == filename).first()
    if reg and reg.dados:
        return reg.dados, (reg.mime or None), (reg.nome_original or None)
    # Fallback: arquivos antigos ainda em disco (até o container reciclar)
    for sub in ("imagens", "documentos", "audios"):
        p = f"/app/{sub}/{filename}"
        if os.path.exists(p):
            try:
                with open(p, "rb") as f:
                    return f.read(), None, None
            except Exception:
                pass
    return None, None, None


def _resposta_midia_range(request, dados, media_type):
    """Resposta com suporte a HTTP Range (206) — necessário p/ o player do navegador descobrir a
    duração, permitir avançar e tocar o áudio COMPLETO (especialmente ogg/webm do WhatsApp)."""
    total = len(dados)
    range_header = request.headers.get("range")
    cabecalho_full = {"Accept-Ranges": "bytes", "Content-Length": str(total)}
    if not range_header:
        return Response(content=dados, media_type=media_type, headers=cabecalho_full)
    m = re.match(r"bytes=(\d*)-(\d*)", range_header.strip())
    if not m:
        return Response(content=dados, media_type=media_type, headers=cabecalho_full)
    ini_s, fim_s = m.group(1), m.group(2)
    if ini_s == "":   # sufixo: últimos N bytes
        n = int(fim_s) if fim_s else 0
        ini, fim = max(0, total - n), total - 1
    else:
        ini = int(ini_s)
        fim = int(fim_s) if fim_s else total - 1
    if ini >= total or ini > fim:
        return Response(status_code=416, headers={"Content-Range": f"bytes */{total}"})
    fim = min(fim, total - 1)
    chunk = dados[ini:fim + 1]
    return Response(content=chunk, status_code=206, media_type=media_type,
                    headers={"Accept-Ranges": "bytes",
                             "Content-Range": f"bytes {ini}-{fim}/{total}",
                             "Content-Length": str(len(chunk))})


@app.get("/api/audio/{filename}")
async def servir_audio(filename: str, request: Request, db: Session = Depends(get_db)):
    """Serve arquivos de áudio (Z-API baixa + reprodução no painel), agora com suporte a Range."""
    if not re.match(r'^[a-f0-9]{32}\.(webm|ogg|mp3|m4a|mp4|aac)$', filename):
        raise HTTPException(status_code=404)
    dados, mime, _ = _buscar_midia(db, filename)
    if dados is None:
        raise HTTPException(status_code=404)
    ext = filename.rsplit(".", 1)[-1]
    tipos = {"ogg": "audio/ogg", "webm": "audio/webm", "mp3": "audio/mpeg",
             "m4a": "audio/mp4", "mp4": "audio/mp4", "aac": "audio/aac"}
    return _resposta_midia_range(request, dados, mime or tipos.get(ext, "audio/ogg"))


@app.get("/api/imagem/{filename}")
async def servir_imagem(filename: str, db: Session = Depends(get_db)):
    """Serve arquivos de imagem salvos do WhatsApp."""
    if not re.match(r'^[a-f0-9]{32}\.(jpg|jpeg|png|webp|gif)$', filename):
        raise HTTPException(status_code=404)
    dados, mime, _ = _buscar_midia(db, filename)
    if dados is None:
        raise HTTPException(status_code=404)
    ext = filename.rsplit(".", 1)[-1].lower()
    tipos = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
             "webp": "image/webp", "gif": "image/gif"}
    return Response(content=dados, media_type=mime or tipos.get(ext, "image/jpeg"))


@app.get("/api/documento/{filename}")
async def servir_documento(filename: str, db: Session = Depends(get_db)):
    """Serve documentos/PDFs salvos do WhatsApp com header de download."""
    if not re.match(r'^[a-f0-9]{32}\.\w{2,5}$', filename):
        raise HTTPException(status_code=404)
    dados, mime, nome = _buscar_midia(db, filename)
    if dados is None:
        raise HTTPException(status_code=404)
    ext = filename.rsplit(".", 1)[-1].lower()
    media_type = mime or ("application/pdf" if ext == "pdf" else "application/octet-stream")
    # Nome de download amigável (preserva acentos via RFC 5987)
    nome_dl = nome or filename
    from urllib.parse import quote
    disp = f"attachment; filename=\"{filename}\"; filename*=UTF-8''{quote(nome_dl)}"
    return Response(content=dados, media_type=media_type, headers={"Content-Disposition": disp})


# ─── Contratos fechados + pasta de documentos do cliente (admin) ─────────────────

_MAX_DOC_BYTES = 25 * 1024 * 1024  # 25 MB por arquivo


@app.get("/api/contratos-fechados")
async def listar_contratos_fechados(
    busca: str = "",
    db: Session = Depends(get_db),
    admin: Usuario = Depends(requer_admin),
):
    """Lista os contratos fechados (admin) com dados do negócio e nº de documentos anexados."""
    q = db.query(Lead).filter(Lead.status == StatusLeadEnum.fechado)
    termo = (busca or "").strip().lower()
    leads = q.order_by(Lead.atualizado_em.desc()).all()
    # contagem de docs por lead
    from collections import defaultdict
    docs_count = defaultdict(int)
    for (lid,) in db.query(DocumentoCliente.lead_id).all():
        docs_count[lid] += 1
    resultado = []
    for l in leads:
        nome = l.nome or l.telefone or ""
        if termo and termo not in nome.lower() and termo not in (l.telefone or "").lower():
            continue
        resp = db.query(Usuario).filter(Usuario.id == l.atribuido_para).first() if l.atribuido_para else None
        # Placa: usa deal_placa; se vazio, tenta o dados_contrato (contrato digital)
        placa = l.deal_placa or ""
        if not placa and l.dados_contrato:
            try:
                placa = (json.loads(l.dados_contrato) or {}).get("vei_placa", "") or ""
            except Exception:
                placa = ""
        resultado.append({
            "id": l.id,
            "nome": nome,
            "telefone": l.telefone,
            "cpf": l.cpf or "",
            "deal_data": l.deal_data or "",
            "deal_veiculo": l.deal_veiculo or "",
            "deal_placa": placa,
            "deal_banco": l.deal_banco or "",
            "deal_operadora": l.deal_operadora or "",
            "deal_valor": l.deal_valor or "",
            "deal_comissao": l.deal_comissao or "",
            "responsavel": resp.nome if resp else "—",
            "docs": docs_count.get(l.id, 0),
            "fechado_em": _fmt_br(l.atualizado_em, "%d/%m/%Y") or "",
        })
    return resultado


@app.get("/api/leads/{lead_id}/documentos")
async def listar_documentos_cliente(
    lead_id: int, db: Session = Depends(get_db), usuario: Usuario = Depends(obter_usuario_atual),
):
    docs = (db.query(DocumentoCliente)
            .filter(DocumentoCliente.lead_id == lead_id)
            .order_by(DocumentoCliente.criado_em.desc()).all())
    return [{
        "id": d.id,
        "nome": d.nome,
        "filename": d.filename,
        "tamanho": d.tamanho,
        "tamanho_str": _tamanho_legivel(d.tamanho),
        "enviado_por": (d.usuario.nome if d.usuario else "—"),
        "em": _fmt_br(d.criado_em, "%d/%m/%Y %H:%M") or "",
    } for d in docs]


@app.post("/api/leads/{lead_id}/documentos")
async def anexar_documento_cliente(
    lead_id: int, arquivo: UploadFile = File(...),
    db: Session = Depends(get_db), usuario: Usuario = Depends(obter_usuario_atual),
):
    """Anexa um documento à pasta do cliente — bytes guardados no banco (durável)."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(404, "Cliente não encontrado")
    dados = await arquivo.read()
    if not dados:
        raise HTTPException(400, "Arquivo vazio")
    if len(dados) > _MAX_DOC_BYTES:
        raise HTTPException(400, "Arquivo muito grande (máx. 25 MB)")
    nome_orig = (arquivo.filename or "documento")[:250]
    ext = (nome_orig.rsplit(".", 1)[-1][:5] if "." in nome_orig else "bin").lower()
    ext = re.sub(r"[^a-z0-9]", "", ext) or "bin"
    filename = f"{uuid.uuid4().hex}.{ext}"
    mime = arquivo.content_type or "application/octet-stream"
    _guardar_blob(db, filename, "documento", dados, nome_original=nome_orig, mime=mime, subdir="documentos")
    doc = DocumentoCliente(lead_id=lead_id, nome=nome_orig, filename=filename,
                           mime=mime, tamanho=len(dados), enviado_por=usuario.id)
    db.add(doc)
    db.commit()
    db.refresh(doc)
    return {"status": "ok", "id": doc.id, "filename": filename}


@app.delete("/api/documentos-cliente/{doc_id}")
async def remover_documento_cliente(
    doc_id: int, db: Session = Depends(get_db), admin: Usuario = Depends(requer_admin),
):
    doc = db.query(DocumentoCliente).filter(DocumentoCliente.id == doc_id).first()
    if not doc:
        raise HTTPException(404, "Documento não encontrado")
    # Remove o blob associado (se nenhum outro registro usa o mesmo filename)
    outros = (db.query(DocumentoCliente)
              .filter(DocumentoCliente.filename == doc.filename,
                      DocumentoCliente.id != doc.id).first())
    if not outros:
        blob = db.query(MidiaArquivo).filter(MidiaArquivo.filename == doc.filename).first()
        if blob:
            db.delete(blob)
    db.delete(doc)
    db.commit()
    return {"status": "ok"}


@app.get("/api/cep/{cep}")
async def buscar_cep(cep: str, usuario: Usuario = Depends(obter_usuario_atual)):
    """Consulta o endereço de um CEP via ViaCEP (pelo servidor, sem bloqueio do navegador)."""
    cep_num = re.sub(r"\D", "", cep or "")
    if len(cep_num) != 8:
        raise HTTPException(400, "CEP inválido")
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.get(f"https://viacep.com.br/ws/{cep_num}/json/")
        d = r.json()
    except Exception:
        raise HTTPException(502, "Erro ao consultar o CEP")
    if not isinstance(d, dict) or d.get("erro"):
        raise HTTPException(404, "CEP não encontrado")
    return {
        "rua": d.get("logradouro", ""),
        "bairro": d.get("bairro", ""),
        "cidade": d.get("localidade", ""),
        "uf": d.get("uf", ""),
    }


@app.get("/api/placa/{placa}")
async def consultar_placa(placa: str, db: Session = Depends(get_db),
                          usuario: Usuario = Depends(obter_usuario_atual)):
    """Consulta dados do veículo pela placa via API Placas (apiplacas.com.br / wdapi2)."""
    placa_limpa = re.sub(r"[^A-Za-z0-9]", "", placa or "").upper()
    if len(placa_limpa) != 7:
        raise HTTPException(400, "Placa inválida (esperado 7 caracteres, ex: ABC1D23)")
    cfg = db.query(Configuracao).filter(Configuracao.chave == "apiplacas_token").first()
    token = (cfg.valor.strip() if cfg and cfg.valor else "") or os.getenv("APIPLACAS_TOKEN", "")
    if not token:
        raise HTTPException(400, "Token da API Placas não configurado. Cole-o em Config. do Bot.")
    try:
        async with httpx.AsyncClient(timeout=20) as client:
            r = await client.get(f"https://wdapi2.com.br/consulta/{placa_limpa}/{token}")
        d = r.json()
    except Exception:
        raise HTTPException(502, "Erro ao consultar a placa. Tente novamente.")
    if not isinstance(d, dict):
        raise HTTPException(502, "Resposta inesperada da consulta de placa")
    extra = d.get("extra") or {}
    def _g(*chaves):
        for ch in chaves:
            for fonte in (d, extra):
                v = fonte.get(ch)
                if v:
                    return str(v).strip()
        return ""
    marca       = _g("MARCA", "marca")
    modelo      = _g("MODELO", "modelo")
    ano_fab     = _g("ano", "anoFabricacao", "ano_fabricacao")
    ano_modelo  = _g("anoModelo", "ano_modelo", "anoModelo")
    chassi      = _g("chassi", "CHASSI").upper()
    cor         = _g("cor", "COR")
    if not (marca or modelo or chassi):
        raise HTTPException(404, str(d.get("mensagem") or d.get("message") or "Placa não encontrada"))
    modelo_completo = " ".join(x for x in [marca, modelo] if x).strip()
    return {"placa": placa_limpa, "marca": marca, "modelo": modelo,
            "modelo_completo": modelo_completo,
            "ano_fab": ano_fab, "ano_modelo": ano_modelo, "chassi": chassi, "cor": cor}


@app.get("/api/placa-debug/{placa}")
async def consultar_placa_debug(placa: str, db: Session = Depends(get_db),
                                admin: Usuario = Depends(requer_admin)):
    """Mostra a resposta CRUA da API Placas — para diagnosticar o preenchimento."""
    placa_limpa = re.sub(r"[^A-Za-z0-9]", "", placa or "").upper()
    cfg = db.query(Configuracao).filter(Configuracao.chave == "apiplacas_token").first()
    token = (cfg.valor.strip() if cfg and cfg.valor else "") or os.getenv("APIPLACAS_TOKEN", "")
    if not token:
        return {"erro": "Token da API Placas NÃO está configurado em Config. do Bot."}
    try:
        async with httpx.AsyncClient(timeout=20) as client:
            r = await client.get(f"https://wdapi2.com.br/consulta/{placa_limpa}/{token}")
        return {"status_http": r.status_code, "resposta": r.json()}
    except Exception as e:
        return {"erro": f"Falha ao consultar: {e}"}


@app.get("/api/admin/backup")
async def baixar_backup(admin: Usuario = Depends(requer_admin)):
    """Baixa um backup COMPLETO do banco (conversas, mídias, documentos — tudo). Admin only."""
    from models import engine
    import sqlite3
    from starlette.background import BackgroundTask
    db_path = engine.url.database
    if not db_path or not os.path.exists(db_path):
        raise HTTPException(400, "Backup disponível apenas para banco SQLite")
    fd, tmp_path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    try:
        src = sqlite3.connect(db_path)
        dst = sqlite3.connect(tmp_path)
        with dst:
            src.backup(dst)   # snapshot consistente mesmo com o sistema rodando
        src.close()
        dst.close()
    except Exception as e:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
        raise HTTPException(500, f"Erro ao gerar backup: {e}")
    fname = f"backup_facil_{_agora_br().strftime('%Y%m%d_%H%M')}.db"
    return FileResponse(tmp_path, filename=fname, media_type="application/octet-stream",
                        background=BackgroundTask(os.remove, tmp_path))


def _migrar_para_postgres_sync():
    """Copia TODOS os dados do banco atual para o PostgreSQL em POSTGRES_URL.
    NÃO mexe no banco atual (só lê). Recria o schema no Postgres e copia tabela a tabela."""
    pg_url = (os.environ.get("POSTGRES_URL") or "").strip()
    if not pg_url:
        raise HTTPException(400, "Defina a variável POSTGRES_URL no Railway antes (referência ao Postgres).")
    if pg_url.startswith("postgres://"):
        pg_url = pg_url.replace("postgres://", "postgresql://", 1)
    if not pg_url.startswith("postgresql"):
        raise HTTPException(400, "POSTGRES_URL inválida.")

    from sqlalchemy import create_engine as _ce, text as _text, Boolean as _Bool, String as _Str
    from models import Base, engine as src_engine

    # TRAVA DE SEGURANÇA: só migra de um SQLite (origem) para o Postgres (destino).
    # Depois da virada de chave o app JÁ roda no Postgres — rodar de novo apagaria tudo
    # (drop_all no destino + copiar da origem que virou o próprio destino vazio).
    if "sqlite" not in str(src_engine.url):
        raise HTTPException(400, "Migração bloqueada: o app já está rodando no PostgreSQL. "
                                 "Rodar agora apagaria os dados (origem = destino). Migração já foi concluída.")

    pg_engine = _ce(pg_url)

    # O SQLite NÃO respeita limites de VARCHAR, então há dados maiores que o tipo declarado
    # (ex.: telefone de GRUPO do WhatsApp "...@g.us" tem >20 chars). O Postgres recusa isso.
    # Solução: criar o schema no Postgres com os VARCHAR sem limite (ilimitado) e restaurar
    # os limites originais no fim — assim o modelo do app continua igual.
    _orig_lens = []
    for _t in Base.metadata.sorted_tables:
        for _c in _t.columns:
            if isinstance(_c.type, _Str) and getattr(_c.type, "length", None) is not None:
                _orig_lens.append((_c, _c.type.length))
                _c.type.length = None

    resultado = {}
    src = None
    dst = None
    try:
        # Schema limpo no Postgres (idempotente: pode rodar de novo sem problema)
        Base.metadata.drop_all(pg_engine)
        Base.metadata.create_all(pg_engine)

        src = src_engine.connect()
        dst = pg_engine.connect()
        # O SQLite NÃO valida chaves estrangeiras, então existem registros órfãos
        # (ex.: um contrato com lead_id de um lead já apagado). O Postgres valida e recusa.
        # Desliga a validação de FK nesta sessão p/ copiar os dados como estão hoje no SQLite.
        try:
            dst.execute(_text("SET session_replication_role = 'replica'"))
            dst.commit()
        except Exception:
            dst.rollback()
        for table in Base.metadata.sorted_tables:   # ordem respeita as chaves estrangeiras
            bool_cols = [c.name for c in table.columns if isinstance(c.type, _Bool)]
            total = 0
            batch = []
            for row in src.execute(table.select()):
                d = dict(row._mapping)
                for bc in bool_cols:               # SQLite guarda 0/1; Postgres precisa de bool
                    if d.get(bc) is not None:
                        d[bc] = bool(d[bc])
                batch.append(d)
                if len(batch) >= 200:
                    dst.execute(table.insert(), batch)
                    total += len(batch); batch = []
            if batch:
                dst.execute(table.insert(), batch)
                total += len(batch)
            dst.commit()
            # Reseta a sequência do id (auto-incremento) p/ o próximo valor correto
            try:
                dst.execute(_text(
                    f"SELECT setval(pg_get_serial_sequence('{table.name}', 'id'), "
                    f"GREATEST((SELECT COALESCE(MAX(id), 1) FROM {table.name}), 1))"
                ))
                dst.commit()
            except Exception:
                dst.rollback()
            resultado[table.name] = total
    finally:
        if src is not None:
            src.close()
        if dst is not None:
            dst.close()
        pg_engine.dispose()
        for _c, _ln in _orig_lens:      # restaura os limites originais do modelo
            _c.type.length = _ln
    return resultado


@app.get("/api/admin/migrar-postgres")
async def migrar_postgres(admin: Usuario = Depends(requer_admin)):
    """Migra os dados para o PostgreSQL (POSTGRES_URL). Não altera o banco atual. Admin only."""
    try:
        copiado = await asyncio.to_thread(_migrar_para_postgres_sync)
    except HTTPException:
        raise
    except Exception as e:
        import traceback as _tb
        return JSONResponse(status_code=500, content={
            "status": "erro",
            "erro": str(e),
            "tipo": type(e).__name__,
            "trace": _tb.format_exc()[-2500:],
        })
    return JSONResponse({"status": "ok", "copiado": copiado,
                         "total_tabelas": len(copiado),
                         "total_linhas": sum(copiado.values())})


def _inbox_sync(db):
    """Parte pesada do inbox (muitas consultas) — roda em thread p/ não travar o servidor."""
    leads_com_msg = (
        db.query(Lead)
        .filter(Lead.status != StatusLeadEnum.desqualificado)
        .order_by(Lead.atualizado_em.desc())
        .all()
    )
    resultado = []
    for l in leads_com_msg:
        ultima = (
            db.query(MensagemConversa)
            .filter(MensagemConversa.telefone == l.telefone)
            .order_by(MensagemConversa.id.desc())
            .first()
        )
        if not ultima:
            continue  # ignora leads sem nenhuma mensagem
        conteudo_limpo = re.sub(r'^\[[^\]]+\]:\s*', '', ultima.conteudo)
        # Não lido (compartilhado): marcada manualmente, OU última msg do cliente depois da última leitura
        nao_lido = bool(l.nao_lido_manual) or (
            (ultima.role == "user") and (l.lido_em is None or ultima.criado_em > l.lido_em))
        resultado.append({
            **_serial_lead(l, db),
            "ultima_mensagem": conteudo_limpo[:60],
            "ultima_hora": _fmt_br(ultima.criado_em, "%H:%M") if ultima and ultima.criado_em else "",
            "ultima_msg_ts": ultima.criado_em.timestamp() if ultima and ultima.criado_em else 0,
            "nao_lido": nao_lido,
        })
    resultado.sort(key=lambda x: x.get("ultima_msg_ts", 0), reverse=True)
    return resultado


@app.get("/api/inbox")
async def inbox(
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    """Retorna TODAS as conversas com mensagens, ordenadas pela mais recente.
    Roda em thread separada para NÃO travar o servidor (evita envios pendurados)."""
    return await asyncio.to_thread(_inbox_sync, db)


@app.post("/api/leads/{lead_id}/marcar-lida")
async def marcar_conversa_lida(lead_id: int, db: Session = Depends(get_db),
                               usuario: Usuario = Depends(obter_usuario_atual)):
    """Marca a conversa como lida para TODOS (leitura compartilhada)."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(404, "Lead não encontrado")
    lead.lido_em = datetime.utcnow()
    lead.nao_lido_manual = False   # abrir/ler limpa a marcação manual de "não lida"
    db.commit()
    return {"status": "ok"}


@app.post("/api/leads/{lead_id}/marcar-nao-lida")
async def marcar_conversa_nao_lida(lead_id: int, db: Session = Depends(get_db),
                                   usuario: Usuario = Depends(obter_usuario_atual)):
    """Marca a conversa como NÃO lida (manual) — pra voltar nela depois. Compartilhado."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(404, "Lead não encontrado")
    lead.nao_lido_manual = True
    db.commit()
    return {"status": "ok"}


@app.patch("/api/leads/{lead_id}")
async def editar_lead(
    lead_id: int,
    request: Request,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    """Edita dados do lead — disponível para admin e funcionários."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead não encontrado")
    body = await request.json()
    campos_editaveis = ["nome", "cpf", "data_nascimento", "carro_interesse", "modalidade", "observacoes", "cidade", "renda", "profissao", "tem_cnh", "email", "descadastrado"]
    # "Fora dos relatórios" só o ADMIN pode mexer (evita esconder conversa malfeita)
    if "ignorar_relatorios" in body and usuario.role == RoleEnum.admin:
        lead.ignorar_relatorios = bool(body["ignorar_relatorios"])
    for campo in campos_editaveis:
        if campo in body:
            valor = body[campo]
            if campo == "cpf" and valor:
                valor = re.sub(r"[^\d]", "", valor)
                if len(valor) == 11:
                    valor = f"{valor[:3]}.{valor[3:6]}.{valor[6:9]}-{valor[9:]}"
            setattr(lead, campo, valor if valor != "" else None)
    # Dados extras para o requerimento (JSON blob)
    if "dados_contrato" in body:
        import json as _json
        dc = body["dados_contrato"]
        lead.dados_contrato = _json.dumps(dc, ensure_ascii=False) if isinstance(dc, dict) else None
    # Carros em proposta (lista JSON)
    if "carros_proposta" in body:
        cp = body["carros_proposta"]
        lead.carros_proposta = json.dumps(cp, ensure_ascii=False) if isinstance(cp, list) else None
    lead.atualizado_em = datetime.utcnow()
    db.commit()
    db.refresh(lead)
    return _serial_lead(lead, db)


@app.post("/api/leads/{lead_id}/observacoes")
async def adicionar_observacao(
    lead_id: int,
    request: Request,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead não encontrado")
    body = await request.json()
    texto = (body.get("texto") or "").strip()
    if not texto:
        raise HTTPException(status_code=400, detail="Texto vazio")

    lista = _parse_observacoes(lead.observacoes)
    lista.append({
        "texto": texto,
        "usuario": usuario.nome,
        "em": datetime.utcnow().strftime("%d/%m/%Y %H:%M"),
    })
    lead.observacoes = json.dumps(lista, ensure_ascii=False)
    lead.atualizado_em = datetime.utcnow()
    db.commit()
    return {"status": "ok", "observacoes": lista}


@app.delete("/api/leads/{lead_id}/observacoes/{idx}")
async def deletar_observacao(
    lead_id: int,
    idx: int,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(requer_admin),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead não encontrado")
    lista = _parse_observacoes(lead.observacoes)
    if idx < 0 or idx >= len(lista):
        raise HTTPException(status_code=400, detail="Índice inválido")
    lista.pop(idx)
    lead.observacoes = json.dumps(lista, ensure_ascii=False)
    lead.atualizado_em = datetime.utcnow()
    db.commit()
    return {"status": "ok", "observacoes": lista}


@app.post("/api/leads/{lead_id}/fechar")
async def fechar_lead(
    lead_id: int,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead não encontrado")
    lead.status = StatusLeadEnum.fechado
    db.commit()
    return _serial_lead(lead, db)


@app.post("/api/sincronizar-chats")
async def sincronizar_chats(
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    """Importa todos os chats existentes do WhatsApp para o painel."""
    if not settings.ZAPI_INSTANCE or not settings.ZAPI_TOKEN:
        raise HTTPException(status_code=400, detail="Z-API não configurada.")

    base = f"https://api.z-api.io/instances/{settings.ZAPI_INSTANCE}/token/{settings.ZAPI_TOKEN}"
    headers = {"Client-Token": settings.ZAPI_CLIENT_TOKEN}

    criados = 0
    ignorados = 0
    page = 1

    async with httpx.AsyncClient() as http:
        while True:
            resp = await http.get(
                f"{base}/chats",
                headers=headers,
                params={"page": page, "pageSize": 100},
                timeout=30,
            )
            if resp.status_code != 200:
                raise HTTPException(status_code=502, detail=f"Erro Z-API: {resp.text}")

            chats = resp.json()
            if not chats:
                break

            for chat in chats:
                # Ignora grupos, broadcasts e spam
                if chat.get("isGroup") or chat.get("isMarkedSpam"):
                    ignorados += 1
                    continue

                telefone = re.sub(r"\D", "", chat.get("phone", ""))
                if not telefone or len(telefone) < 10:
                    ignorados += 1
                    continue

                # Só cria se ainda não existe
                existente = db.query(Lead).filter(Lead.telefone == telefone).first()
                if not existente:
                    nome = chat.get("name") or None
                    # Ignora se o nome for igual ao telefone (sem nome real)
                    if nome and re.sub(r"\D", "", nome) == telefone:
                        nome = None
                    lead = Lead(
                        telefone=telefone,
                        nome=nome,
                        status=StatusLeadEnum.em_atendimento,
                        estado_conversa=EstadoConversaEnum.transferido,
                    )
                    db.add(lead)
                    criados += 1
                else:
                    # Atualiza nome se ainda não tinha
                    nome = chat.get("name") or None
                    if nome and re.sub(r"\D", "", nome) == telefone:
                        nome = None
                    if nome and not existente.nome:
                        existente.nome = nome
                    ignorados += 1

            db.commit()

            if len(chats) < 100:
                break
            page += 1

    return {"criados": criados, "ignorados": ignorados, "total_paginas": page}


@app.delete("/api/leads/{lead_id}")
async def excluir_lead(
    lead_id: int,
    db: Session = Depends(get_db),
    admin: Usuario = Depends(requer_admin),
):
    """Exclui permanentemente um lead e todas as suas mensagens. Apenas admin."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead não encontrado")
    db.query(MensagemConversa).filter(MensagemConversa.telefone == lead.telefone).delete()
    db.delete(lead)
    db.commit()
    return {"status": "excluido"}


@app.put("/api/leads/{lead_id}/atribuir")
async def atribuir_lead(
    lead_id: int,
    request: Request,
    db: Session = Depends(get_db),
    admin: Usuario = Depends(requer_admin),
):
    """Troca o atendente responsável por um lead. Apenas admin."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Lead não encontrado")
    body = await request.json()
    usuario_id = body.get("usuario_id")
    if usuario_id:
        u = db.query(Usuario).filter(Usuario.id == usuario_id, Usuario.ativo == True).first()
        if not u:
            raise HTTPException(status_code=404, detail="Usuário não encontrado")
        lead.atribuido_para = u.id
        lead.assumido_em = lead.assumido_em or datetime.utcnow()
        if lead.status == StatusLeadEnum.em_atendimento or lead.status == StatusLeadEnum.qualificado:
            lead.status = StatusLeadEnum.assumido
    else:
        lead.atribuido_para = None
    lead.atualizado_em = datetime.utcnow()
    db.commit()
    return _serial_lead(lead, db)


@app.get("/api/dashboard-stats")
async def dashboard_stats(
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    """Métricas completas para a aba Dashboard."""
    agora = _agora_br()
    inicio_mes_utc = datetime(agora.year, agora.month, 1, tzinfo=_TZ_BR).astimezone(timezone.utc).replace(tzinfo=None)

    # ── Leads do mês por origem (exclui conversas internas) ────────────────
    leads_mes_q = db.query(Lead).filter(Lead.criado_em >= inicio_mes_utc,
                                        Lead.ignorar_relatorios.isnot(True))
    total_mes = leads_mes_q.count()

    origens_map = {
        "whatsapp":   "WhatsApp / Bot",
        "rede_social": "Rede Social",
        "parceiro":   "Parceiro",
        "indicacao":  "Indicação",
        "ex_cliente": "Ex-cliente",
    }
    por_origem = {}
    for chave, label in origens_map.items():
        por_origem[label] = leads_mes_q.filter(Lead.origem == chave).count()
    # Bot/sem origem: leads sem origem definida (chegaram pelo WhatsApp bot)
    sem_origem = leads_mes_q.filter(Lead.origem == None).count()
    por_origem["WhatsApp / Bot"] = por_origem.get("WhatsApp / Bot", 0) + sem_origem

    # ── Conversões do mês (exclui conversas internas) ──────────────────────
    _ign = Lead.ignorar_relatorios.isnot(True)
    propostas_mes = db.query(Lead).filter(
        Lead.criado_em >= inicio_mes_utc, _ign,
        Lead.status.in_([StatusLeadEnum.proposta_enviada, StatusLeadEnum.proposta_aprovada, StatusLeadEnum.fechado])
    ).count()
    aprovadas_mes = db.query(Lead).filter(
        Lead.criado_em >= inicio_mes_utc, _ign,
        Lead.status.in_([StatusLeadEnum.proposta_aprovada, StatusLeadEnum.fechado])
    ).count()
    fechados_mes = db.query(Lead).filter(
        Lead.fechado_em >= inicio_mes_utc, _ign,
        Lead.status == StatusLeadEnum.fechado
    ).count()

    conv_proposta  = round(propostas_mes / total_mes * 100, 1) if total_mes > 0 else 0
    conv_aprovada  = round(aprovadas_mes / propostas_mes * 100, 1) if propostas_mes > 0 else 0

    # ── Meta do mês ───────────────────────────────────────────────────────
    faixas_meta = _get_meta_faixas(db, agora.year, agora.month)
    meta = faixas_meta[0]["contratos"] if faixas_meta else 20  # menor faixa como base
    pct_meta = round(fechados_mes / meta * 100, 1) if meta > 0 else 0
    # faixa atingida = maior faixa cujo threshold <= fechados_mes
    faixa_atingida = None
    for f in faixas_meta:
        if fechados_mes >= f["contratos"]:
            faixa_atingida = f

    # ── Ranking do mês ────────────────────────────────────────────────────
    funcionarias = db.query(Usuario).filter(Usuario.role == RoleEnum.funcionario, Usuario.ativo == True).all()
    ranking = []
    for f in funcionarias:
        qtd = db.query(Lead).filter(
            Lead.atribuido_para == f.id,
            Lead.status == StatusLeadEnum.fechado,
            Lead.fechado_em >= inicio_mes_utc,
            _ign,
        ).count()
        ranking.append({"nome": f.nome, "contratos": qtd})
    ranking.sort(key=lambda x: x["contratos"], reverse=True)

    # ── Tarja financeira (admin) ───────────────────────────────────────────
    tarja = None
    if usuario.role == RoleEnum.admin:
        leads_fechados_mes = db.query(Lead).filter(
            Lead.fechado_em >= inicio_mes_utc,
            Lead.status == StatusLeadEnum.fechado,
            _ign,
        ).all()
        def _to_float(v):
            if not v: return 0.0
            try: return float(str(v).replace("R$","").replace(".","").replace(",",".").strip())
            except: return 0.0
        total_valor    = sum(_to_float(l.deal_valor) for l in leads_fechados_mes)
        total_comissao = sum(_to_float(l.deal_comissao) for l in leads_fechados_mes)
        tarja = {
            "contratos": len(leads_fechados_mes),
            "total_valor": f"R$ {total_valor:,.2f}".replace(",","X").replace(".",",").replace("X","."),
            "total_comissao": f"R$ {total_comissao:,.2f}".replace(",","X").replace(".",",").replace("X","."),
        }

    return {
        "mes": f"{agora.month:02d}/{agora.year}",
        "total_mes": total_mes,
        "por_origem": por_origem,
        "propostas_mes": propostas_mes,
        "aprovadas_mes": aprovadas_mes,
        "fechados_mes": fechados_mes,
        "conv_proposta": conv_proposta,
        "conv_aprovada": conv_aprovada,
        "meta": meta,
        "pct_meta": pct_meta,
        "faixas_meta": faixas_meta,
        "faixa_atingida": faixa_atingida,
        "ranking": ranking,
        "tarja": tarja,
    }


@app.get("/api/stats")
async def estatisticas(
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    total = db.query(Lead).count()
    em_atendimento    = db.query(Lead).filter(Lead.status == StatusLeadEnum.em_atendimento).count()
    qualificados      = db.query(Lead).filter(Lead.status == StatusLeadEnum.qualificado).count()
    assumidos         = db.query(Lead).filter(Lead.status == StatusLeadEnum.assumido).count()
    propostas         = db.query(Lead).filter(Lead.status == StatusLeadEnum.proposta_enviada).count()
    proposta_aprovada = db.query(Lead).filter(Lead.status == StatusLeadEnum.proposta_aprovada).count()
    fechados          = db.query(Lead).filter(Lead.status == StatusLeadEnum.fechado).count()
    perdidos          = db.query(Lead).filter(Lead.status == StatusLeadEnum.perdido).count()
    desqualificados   = db.query(Lead).filter(Lead.status == StatusLeadEnum.desqualificado).count()
    conv = qualificados + assumidos + propostas + proposta_aprovada + fechados

    # Leads inseridos neste mês calendário
    agora = _agora_br()
    inicio_mes = datetime(agora.year, agora.month, 1, tzinfo=_TZ_BR)
    leads_mes = db.query(Lead).filter(Lead.criado_em >= inicio_mes.astimezone(timezone.utc).replace(tzinfo=None)).count()

    # Taxas de conversão (base = total de leads)
    taxa_prop_aprovada = round(proposta_aprovada / total * 100, 1) if total > 0 else 0
    taxa_fechado       = round(fechados / total * 100, 1) if total > 0 else 0

    # Ranking de funcionárias por contratos fechados no mês
    inicio_mes_utc = inicio_mes.astimezone(timezone.utc).replace(tzinfo=None)
    funcionarias = db.query(Usuario).filter(
        Usuario.role == RoleEnum.funcionario, Usuario.ativo == True
    ).all()
    ranking_mes = []
    for f in funcionarias:
        qtd = db.query(Lead).filter(
            Lead.atribuido_para == f.id,
            Lead.status == StatusLeadEnum.fechado,
            Lead.fechado_em >= inicio_mes_utc,
        ).count()
        ranking_mes.append({"nome": f.nome, "contratos": qtd})
    ranking_mes.sort(key=lambda x: x["contratos"], reverse=True)

    return {
        "total": total,
        "leads_mes": leads_mes,
        "em_atendimento": em_atendimento,
        "qualificados": qualificados,
        "assumidos": assumidos,
        "propostas": propostas,
        "proposta_aprovada": proposta_aprovada,
        "fechados": fechados,
        "perdidos": perdidos,
        "desqualificados": desqualificados,
        "financiamento": db.query(Lead).filter(Lead.modalidade == "financiamento").count(),
        "refinanciamento": db.query(Lead).filter(Lead.modalidade == "refinanciamento").count(),
        "taxa_qualificacao":   round((conv / total * 100), 1) if total > 0 else 0,
        "taxa_prop_aprovada":  taxa_prop_aprovada,
        "taxa_fechado":        taxa_fechado,
        "ranking_mes":         ranking_mes,
    }


# ─── API Usuários (admin) ─────────────────────────────────────────────────────────

@app.get("/api/usuarios/opcoes")
async def listar_usuarios_opcoes(
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    """Lista id+nome de usuários ativos — acessível a qualquer usuária logada (para dropdowns)."""
    usuarios = db.query(Usuario).filter(Usuario.ativo == True).order_by(Usuario.nome).all()
    return [{"id": u.id, "nome": u.nome} for u in usuarios]


@app.get("/api/usuarios")
async def listar_usuarios(db: Session = Depends(get_db), admin: Usuario = Depends(requer_admin)):
    from sqlalchemy import func
    usuarios = db.query(Usuario).order_by(Usuario.criado_em).all()
    # leads em aberto por responsável — pra mostrar quem ainda tem atendimentos parados
    contagem = dict(
        db.query(Lead.atribuido_para, func.count(Lead.id))
        .filter(Lead.atribuido_para.isnot(None), Lead.status.notin_(_STATUS_FINALIZADOS))
        .group_by(Lead.atribuido_para)
        .all()
    )
    out = []
    for u in usuarios:
        d = _serial_usuario(u)
        d["leads_abertos"] = int(contagem.get(u.id, 0))
        out.append(d)
    return out


@app.post("/api/usuarios")
async def criar_usuario(request: Request, db: Session = Depends(get_db), admin: Usuario = Depends(requer_admin)):
    body = await request.json()
    email = body.get("email", "").strip().lower()
    if db.query(Usuario).filter(Usuario.email == email).first():
        raise HTTPException(status_code=400, detail="E-mail já cadastrado")
    u = Usuario(
        nome=body.get("nome", "").strip(),
        email=email,
        senha_hash=hash_senha(body.get("senha", "Senha@123")),
        role=body.get("role", RoleEnum.funcionario),
        ativo=True,
    )
    db.add(u)
    db.commit()
    db.refresh(u)
    return _serial_usuario(u)


@app.put("/api/usuarios/{uid}")
async def atualizar_usuario(uid: int, request: Request, db: Session = Depends(get_db), admin: Usuario = Depends(requer_admin)):
    u = db.query(Usuario).filter(Usuario.id == uid).first()
    if not u:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    body = await request.json()
    if "nome" in body:
        u.nome = body["nome"].strip()
    if "email" in body:
        novo_email = body["email"].strip().lower()
        if novo_email != u.email:
            if db.query(Usuario).filter(Usuario.email == novo_email, Usuario.id != uid).first():
                raise HTTPException(status_code=400, detail="E-mail já está em uso por outro usuário")
            u.email = novo_email
    if "role" in body:
        u.role = body["role"]
    if "ativo" in body:
        u.ativo = body["ativo"]
    if body.get("senha"):
        u.senha_hash = hash_senha(body["senha"])
    db.commit()
    db.refresh(u)
    return _serial_usuario(u)


@app.delete("/api/usuarios/{uid}")
async def desativar_usuario(uid: int, db: Session = Depends(get_db), admin: Usuario = Depends(requer_admin)):
    u = db.query(Usuario).filter(Usuario.id == uid).first()
    if not u:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    u.ativo = False
    db.commit()
    return {"status": "desativado"}


@app.delete("/api/usuarios/{uid}/excluir")
async def excluir_usuario_definitivo(uid: int, db: Session = Depends(get_db), admin: Usuario = Depends(requer_admin)):
    """Exclui DEFINITIVAMENTE um usuário — só se estiver desligado E sem histórico de negócio.
    Se tiver leads/ponto/contratos/etc., recusa: deve permanecer desligado p/ preservar os dados."""
    u = db.query(Usuario).filter(Usuario.id == uid).first()
    if not u:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    if u.id == admin.id:
        raise HTTPException(status_code=400, detail="Você não pode excluir a si mesma.")
    if u.ativo:
        raise HTTPException(status_code=400, detail="Só dá para excluir um usuário já desligado. Desligue primeiro.")
    nome = u.nome
    refs = {
        "leads atendidos": db.query(Lead).filter(Lead.atribuido_para == uid).count(),
        "parceiros na carteira": db.query(Parceiro).filter(Parceiro.operadora_id == uid).count(),
        "contratos": db.query(Contrato).filter(Contrato.criado_por_id == uid).count(),
        "registros de ponto": db.query(RegistroPonto).filter(RegistroPonto.usuario_id == uid).count(),
        "férias/folgas": db.query(AusenciaFuncionaria).filter(AusenciaFuncionaria.usuario_id == uid).count(),
        "agendamentos": db.query(Agendamento).filter(Agendamento.criado_por == uid).count(),
        "histórico de leads": db.query(HistoricoLead).filter(HistoricoLead.usuario_id == uid).count(),
        "justificativas de ponto": db.query(JustificativaPonto).filter(JustificativaPonto.usuario_id == uid).count(),
        "correções de ponto": db.query(CorrecaoPonto).filter(CorrecaoPonto.usuario_id == uid).count(),
    }
    bloqueios = {k: v for k, v in refs.items() if v}
    if bloqueios:
        detalhe = "; ".join(f"{k}: {v}" for k, v in bloqueios.items())
        raise HTTPException(
            status_code=400,
            detail=(f"Não dá para EXCLUIR: essa pessoa tem histórico no sistema ({detalhe}). "
                    f"Mantenha DESLIGADA — o acesso já está cortado e os dados ficam preservados."),
        )
    # Conta sem histórico de negócio → limpa logs de sessão/atividade e exclui
    try:
        db.query(SessaoUsuario).filter(SessaoUsuario.usuario_id == uid).delete(synchronize_session=False)
        db.query(AtividadePing).filter(AtividadePing.usuario_id == uid).delete(synchronize_session=False)
        db.delete(u)
        db.commit()
    except Exception as e:
        db.rollback()
        print(f"⚠️ Falha ao excluir usuário {uid}: {e}")
        raise HTTPException(status_code=400, detail="Não foi possível excluir com segurança (há registros ligados). Mantenha a pessoa desligada.")
    return {"status": "excluido", "nome": nome}


_STATUS_FINALIZADOS = [
    StatusLeadEnum.fechado.value,
    StatusLeadEnum.perdido.value,
    StatusLeadEnum.desqualificado.value,
]


@app.get("/api/usuarios/{uid}/desligamento")
async def preview_desligamento(uid: int, db: Session = Depends(get_db), admin: Usuario = Depends(requer_admin)):
    """Prévia do desligamento: quantos leads em aberto a pessoa tem e quem pode herdá-los."""
    u = db.query(Usuario).filter(Usuario.id == uid).first()
    if not u:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    if u.role == RoleEnum.admin:
        raise HTTPException(status_code=400, detail="Não é possível desligar um administrador por aqui.")
    abertos = db.query(Lead).filter(
        Lead.atribuido_para == uid,
        Lead.status.notin_(_STATUS_FINALIZADOS),
    ).count()
    equipe = db.query(Usuario).filter(
        Usuario.role == RoleEnum.funcionario, Usuario.ativo == True, Usuario.id != uid
    ).order_by(Usuario.nome).all()
    return {
        "nome": u.nome,
        "leads_abertos": abertos,
        "equipe": [{"id": x.id, "nome": x.nome} for x in equipe],
    }


@app.post("/api/usuarios/{uid}/desligar")
async def desligar_usuario(uid: int, request: Request, db: Session = Depends(get_db), admin: Usuario = Depends(requer_admin)):
    """Desliga a funcionária: redistribui os leads em aberto e desativa o acesso.
    Preserva todo o histórico (vendas fechadas, ponto, relatórios). modo:
    dividir (round-robin no time) | uma_pessoa (destino_id) | fila (sem dono) | admin."""
    u = db.query(Usuario).filter(Usuario.id == uid).first()
    if not u:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    if u.id == admin.id:
        raise HTTPException(status_code=400, detail="Você não pode desligar a si mesma.")
    if u.role == RoleEnum.admin:
        raise HTTPException(status_code=400, detail="Não é possível desligar um administrador por aqui. Use Editar para desativar.")
    try:
        body = await request.json()
    except Exception:
        body = {}
    modo = (body.get("modo") or "dividir").strip()
    destino_id = body.get("destino_id")

    abertos = db.query(Lead).filter(
        Lead.atribuido_para == uid,
        Lead.status.notin_(_STATUS_FINALIZADOS),
    ).all()
    distribuicao = {}

    def _add(nome, n=1):
        distribuicao[nome] = distribuicao.get(nome, 0) + n

    def _marcar(lead, destino_txt):
        """Registra nas observações internas que o lead veio da redistribuição."""
        lst = _parse_observacoes(lead.observacoes)
        lst.append({
            "texto": f"🔄 Redistribuído de {u.nome}{destino_txt}",
            "usuario": admin.nome,
            "em": datetime.utcnow().strftime("%d/%m/%Y %H:%M"),
        })
        lead.observacoes = json.dumps(lst, ensure_ascii=False)

    if modo == "fila":
        for lead in abertos:
            lead.atribuido_para = None
            _marcar(lead, " (devolvido à fila)")
        if abertos:
            _add("Fila (sem dono)", len(abertos))
    elif modo == "uma_pessoa":
        alvo = db.query(Usuario).filter(
            Usuario.id == destino_id, Usuario.id != uid,
            Usuario.ativo == True, Usuario.role == RoleEnum.funcionario
        ).first() if destino_id else None
        if not alvo:
            raise HTTPException(status_code=400, detail="Escolha uma operadora válida para receber os leads.")
        for lead in abertos:
            lead.atribuido_para = alvo.id
            _marcar(lead, f" → {alvo.nome}")
        if abertos:
            _add(alvo.nome, len(abertos))
    elif modo == "admin":
        for lead in abertos:
            lead.atribuido_para = admin.id
            _marcar(lead, f" → {admin.nome}")
        if abertos:
            _add(admin.nome, len(abertos))
    else:  # "dividir" — round-robin entre as operadoras ativas
        equipe = db.query(Usuario).filter(
            Usuario.role == RoleEnum.funcionario, Usuario.ativo == True, Usuario.id != uid
        ).order_by(Usuario.nome).all()
        if equipe:
            for i, lead in enumerate(abertos):
                alvo = equipe[i % len(equipe)]
                lead.atribuido_para = alvo.id
                _marcar(lead, f" → {alvo.nome}")
                _add(alvo.nome)
        else:  # ninguém no time → fica com quem está desligando
            for lead in abertos:
                lead.atribuido_para = admin.id
                _marcar(lead, f" → {admin.nome}")
            if abertos:
                _add(admin.nome, len(abertos))

    u.ativo = False
    db.commit()
    return {
        "status": "ok",
        "nome": u.nome,
        "leads_reatribuidos": len(abertos),
        "distribuicao": distribuicao,
    }


# ─── Ausências / Disponibilidade de Funcionárias ────────────────────────────────

@app.get("/api/usuarios/{uid}/ausencias")
async def listar_ausencias(uid: int, db: Session = Depends(get_db), admin: Usuario = Depends(requer_admin)):
    u = db.query(Usuario).filter(Usuario.id == uid).first()
    if not u:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    aus = (db.query(AusenciaFuncionaria)
             .filter(AusenciaFuncionaria.usuario_id == uid)
             .order_by(AusenciaFuncionaria.data_inicio)
             .all())
    return [
        {
            "id": a.id,
            "tipo": a.tipo,
            "data_inicio": a.data_inicio,
            "data_fim": a.data_fim,
            "observacao": a.observacao,
            "criado_em": a.criado_em.isoformat() if a.criado_em else None,
        }
        for a in aus
    ]


@app.post("/api/usuarios/{uid}/ausencias")
async def criar_ausencia(uid: int, request: Request, db: Session = Depends(get_db), admin: Usuario = Depends(requer_admin)):
    u = db.query(Usuario).filter(Usuario.id == uid).first()
    if not u:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    body = await request.json()
    tipo = body.get("tipo", "").strip()
    if tipo not in ("folga", "ferias", "afastamento"):
        raise HTTPException(status_code=400, detail="Tipo inválido")
    data_inicio = body.get("data_inicio", "").strip()
    data_fim = body.get("data_fim", "").strip()
    if not data_inicio or not data_fim:
        raise HTTPException(status_code=400, detail="Datas obrigatórias")
    if data_fim < data_inicio:
        raise HTTPException(status_code=400, detail="Data fim deve ser maior ou igual à data início")
    aus = AusenciaFuncionaria(
        usuario_id=uid,
        tipo=tipo,
        data_inicio=data_inicio,
        data_fim=data_fim,
        observacao=body.get("observacao", "").strip() or None,
    )
    db.add(aus)
    db.commit()
    db.refresh(aus)
    return {
        "id": aus.id,
        "tipo": aus.tipo,
        "data_inicio": aus.data_inicio,
        "data_fim": aus.data_fim,
        "observacao": aus.observacao,
        "criado_em": aus.criado_em.isoformat() if aus.criado_em else None,
    }


@app.delete("/api/ausencias/{aid}")
async def remover_ausencia(aid: int, db: Session = Depends(get_db), admin: Usuario = Depends(requer_admin)):
    aus = db.query(AusenciaFuncionaria).filter(AusenciaFuncionaria.id == aid).first()
    if not aus:
        raise HTTPException(status_code=404, detail="Ausência não encontrada")
    db.delete(aus)
    db.commit()
    return {"status": "removida"}


# ═══════════ RH: Férias/Folgas (pedido+aprovação+saldo) e Banco de Horas ═══════

def _parse_iso(s):
    try:
        return datetime.strptime((s or "").strip(), "%Y-%m-%d").date()
    except Exception:
        return None


def _dias_inclusivos(ini, fim):
    di, df = _parse_iso(ini), _parse_iso(fim)
    if not di or not df or df < di:
        return 0
    return (df - di).days + 1


def _meses_trabalhados(data_admissao):
    da = _parse_iso(data_admissao)
    if not da:
        return 0
    hoje = _agora_br().date()
    meses = (hoje.year - da.year) * 12 + (hoje.month - da.month)
    if hoje.day < da.day:
        meses -= 1
    return max(0, meses)


def _saldo_ferias(db, u):
    """Saldo de férias: 30 dias por período de 12 meses (CLT) − tirados + ajuste."""
    meses = _meses_trabalhados(u.data_admissao)
    periodos = meses // 12
    dias_direito = periodos * 30
    proporcional = round((meses % 12) * 2.5)
    aus = db.query(AusenciaFuncionaria).filter(
        AusenciaFuncionaria.usuario_id == u.id,
        AusenciaFuncionaria.status == "aprovada",
    ).all()
    tirados = 0
    for a in aus:
        if a.tipo == "ferias" or (a.tipo == "folga" and a.desconta_ferias):
            tirados += _dias_inclusivos(a.data_inicio, a.data_fim)
    ajuste = u.ferias_ajuste or 0
    return {
        "data_admissao": u.data_admissao,
        "meses_trabalhados": meses,
        "dias_direito": dias_direito,
        "proporcional_periodo": proporcional,
        "dias_tirados": tirados,
        "ajuste": ajuste,
        "saldo_disponivel": dias_direito - tirados + ajuste,
    }


def _serial_ausencia(a):
    return {
        "id": a.id,
        "usuario_id": a.usuario_id,
        "funcionaria": (a.usuario.nome if a.usuario else "—"),
        "tipo": a.tipo,
        "data_inicio": a.data_inicio,
        "data_fim": a.data_fim,
        "data_inicio_br": _iso_para_br(a.data_inicio),
        "data_fim_br": _iso_para_br(a.data_fim),
        "dias": _dias_inclusivos(a.data_inicio, a.data_fim),
        "observacao": a.observacao or "",
        "status": a.status or "aprovada",
        "desconta_ferias": bool(a.desconta_ferias),
        "obs_admin": a.obs_admin or "",
        "solicitante": (a.solicitante.nome if a.solicitante else None),
        "resolvido_por": (a.resolvedor.nome if a.resolvedor else None),
    }


@app.get("/api/rh/ausencias")
async def rh_listar_ausencias(db: Session = Depends(get_db),
                              usuario: Usuario = Depends(obter_usuario_atual)):
    """Todas as férias/folgas de quem está ATIVO — visível a todos (calendário compartilhado).
    Desligados não aparecem mais aqui (mas o histórico deles fica preservado)."""
    aus = (db.query(AusenciaFuncionaria)
           .join(Usuario, Usuario.id == AusenciaFuncionaria.usuario_id)
           .filter(Usuario.ativo == True)
           .order_by(AusenciaFuncionaria.data_inicio.desc()).all())
    return [_serial_ausencia(a) for a in aus]


@app.post("/api/rh/ausencias")
async def rh_pedir_ausencia(request: Request, db: Session = Depends(get_db),
                            usuario: Usuario = Depends(obter_usuario_atual)):
    """Funcionária PEDE folga/férias (fica pendente). Admin pode lançar direto p/ alguém (já aprovado)."""
    body = await request.json()
    tipo = (body.get("tipo") or "").strip()
    if tipo not in ("folga", "ferias", "afastamento"):
        raise HTTPException(400, "Tipo inválido")
    di = (body.get("data_inicio") or "").strip()
    df = (body.get("data_fim") or "").strip()
    if not _parse_iso(di) or not _parse_iso(df):
        raise HTTPException(400, "Datas inválidas")
    if df < di:
        raise HTTPException(400, "Data fim antes da data início")
    desconta = bool(body.get("desconta_ferias")) and tipo == "folga"
    alvo_id = usuario.id
    status = "pendente"
    if usuario.role == RoleEnum.admin and body.get("usuario_id"):
        try:
            alvo_id = int(body["usuario_id"])
            status = "aprovada"   # admin lançando para alguém já fica aprovado
        except (TypeError, ValueError):
            raise HTTPException(400, "Funcionária inválida")
    a = AusenciaFuncionaria(
        usuario_id=alvo_id, tipo=tipo, data_inicio=di, data_fim=df,
        observacao=(body.get("observacao") or "").strip() or None,
        status=status, solicitante_id=usuario.id, desconta_ferias=desconta,
        resolvido_por=(usuario.id if status == "aprovada" else None),
        resolvido_em=(datetime.utcnow() if status == "aprovada" else None),
    )
    db.add(a)
    db.commit()
    db.refresh(a)
    return _serial_ausencia(a)


@app.post("/api/rh/ausencias/{aid}/aprovar")
async def rh_aprovar_ausencia(aid: int, db: Session = Depends(get_db),
                              admin: Usuario = Depends(requer_admin)):
    a = db.query(AusenciaFuncionaria).filter(AusenciaFuncionaria.id == aid).first()
    if not a:
        raise HTTPException(404, "Pedido não encontrado")
    a.status = "aprovada"
    a.obs_admin = None
    a.resolvido_por = admin.id
    a.resolvido_em = datetime.utcnow()
    db.commit()
    db.refresh(a)
    return _serial_ausencia(a)


@app.post("/api/rh/ausencias/{aid}/rejeitar")
async def rh_rejeitar_ausencia(aid: int, request: Request, db: Session = Depends(get_db),
                               admin: Usuario = Depends(requer_admin)):
    body = await request.json()
    a = db.query(AusenciaFuncionaria).filter(AusenciaFuncionaria.id == aid).first()
    if not a:
        raise HTTPException(404, "Pedido não encontrado")
    a.status = "rejeitada"
    a.obs_admin = (body.get("obs") or "").strip()[:400] or None
    a.resolvido_por = admin.id
    a.resolvido_em = datetime.utcnow()
    db.commit()
    db.refresh(a)
    return _serial_ausencia(a)


@app.get("/api/rh/saldos")
async def rh_saldos(db: Session = Depends(get_db),
                    usuario: Usuario = Depends(obter_usuario_atual)):
    """Saldo de férias. Admin vê de todas; funcionária vê só o dela."""
    if usuario.role == RoleEnum.admin:
        funcs = (db.query(Usuario)
                 .filter(Usuario.ativo == True, Usuario.role == RoleEnum.funcionario)
                 .order_by(Usuario.nome).all())
    else:
        funcs = [usuario]
    return [{"usuario_id": u.id, "funcionaria": u.nome, **_saldo_ferias(db, u)} for u in funcs]


@app.put("/api/rh/usuario/{uid}")
async def rh_definir_usuario(uid: int, request: Request, db: Session = Depends(get_db),
                             admin: Usuario = Depends(requer_admin)):
    """Admin define data de admissão e ajuste manual do saldo de férias."""
    u = db.query(Usuario).filter(Usuario.id == uid).first()
    if not u:
        raise HTTPException(404, "Funcionária não encontrada")
    body = await request.json()
    if "data_admissao" in body:
        da = (body.get("data_admissao") or "").strip()
        u.data_admissao = da if _parse_iso(da) else None
    if "ferias_ajuste" in body:
        try:
            u.ferias_ajuste = int(body.get("ferias_ajuste") or 0)
        except (TypeError, ValueError):
            u.ferias_ajuste = 0
    db.commit()
    db.refresh(u)
    return {"usuario_id": u.id, "funcionaria": u.nome, **_saldo_ferias(db, u)}


def _segundos_esperados_dia(d):
    """Jornada esperada (líquida) num dia: 8h seg–sex, 4h sáb, 0 dom."""
    wd = d.weekday()   # 0=seg … 6=dom
    if wd < 5:
        return 8 * 3600
    if wd == 5:
        return 4 * 3600
    return 0


@app.get("/api/rh/banco-horas")
async def rh_banco_horas(meses: int = 6, db: Session = Depends(get_db),
                         admin: Usuario = Depends(requer_admin)):
    """Banco de horas por funcionária — conta só DIAS COMPLETOS (entrada+saída),
    a partir da data de início configurada (padrão: hoje, ou seja, de agora pra frente)."""
    # Data de início (config). Se não existir, começa HOJE.
    cfg = db.query(Configuracao).filter(Configuracao.chave == "banco_horas_inicio").first()
    inicio_cfg = _parse_iso(cfg.valor) if (cfg and _parse_iso(cfg.valor)) else None
    if inicio_cfg is None:
        inicio_cfg = _agora_br().date()
        if not cfg:
            db.add(Configuracao(chave="banco_horas_inicio",
                                valor=inicio_cfg.strftime("%Y-%m-%d"),
                                descricao="Data de início do banco de horas"))
            db.commit()
    meses = max(1, min(meses, 24))
    hoje = _agora_br().date()
    y, m = hoje.year, hoje.month - (meses - 1)
    while m <= 0:
        m += 12
        y -= 1
    janela = datetime(y, m, 1).date()
    inicio = max(inicio_cfg, janela)   # nunca conta antes da data de início
    agora_utc = datetime.utcnow()
    funcs = (db.query(Usuario)
             .filter(Usuario.ativo == True, Usuario.role == RoleEnum.funcionario)
             .order_by(Usuario.nome).all())
    nomes_mes = ["", "jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]
    resultado = []
    for u in funcs:
        saldo_s = 0
        por_mes = {}
        d = inicio
        while d < hoje:   # só dias COMPLETOS (não conta hoje, que está em andamento)
            pontos = _pontos_do_dia(db, u.id, d)
            if pontos:
                intervalos = _jornadas_de_pontos(pontos, agora_utc)
                # Só conta o dia se o ponto está FECHADO (entrada+saída).
                # Dia com ponto aberto (esqueceu de bater a saída) é ignorado — senão infla.
                if not any(aberto for _, _, aberto in intervalos):
                    trab = sum(int((f - i).total_seconds()) for i, f, _ in intervalos)
                    # Almoço não batido: se foi UM bloco contínuo > 6h (entrada→saída direto,
                    # sem registrar saída/volta do almoço), desconta 1h presumida — senão
                    # conta o almoço como trabalho e infla ~1h/dia.
                    if len(intervalos) == 1 and trab > 6 * 3600:
                        trab -= 3600
                    diff = trab - _segundos_esperados_dia(d)
                    saldo_s += diff
                    k = f"{d.year:04d}-{d.month:02d}"
                    por_mes[k] = por_mes.get(k, 0) + diff
            d += timedelta(days=1)
        # Alerta 6 meses: existe saldo positivo e o mês mais antigo da janela teve extra positivo
        k_antigo = f"{inicio.year:04d}-{inicio.month:02d}"
        alerta_6m = saldo_s > 0 and por_mes.get(k_antigo, 0) > 0 and meses >= 6
        meses_lst = []
        for k in sorted(por_mes.keys()):
            yy, mm = k.split("-")
            meses_lst.append({"mes": f"{nomes_mes[int(mm)]}/{yy}",
                              "saldo": _duracao_str_assinado(por_mes[k])})
        resultado.append({
            "usuario_id": u.id,
            "funcionaria": u.nome,
            "saldo": _duracao_str_assinado(saldo_s),
            "saldo_s": saldo_s,
            "positivo": saldo_s > 0,
            "alerta_6m": alerta_6m,
            "por_mes": meses_lst,
        })
    return {"desde": inicio.strftime("%d/%m/%Y"), "ate": hoje.strftime("%d/%m/%Y"),
            "inicio_iso": inicio_cfg.strftime("%Y-%m-%d"),
            "funcionarias": resultado}


@app.post("/api/rh/banco-horas/inicio")
async def rh_banco_horas_inicio(request: Request, db: Session = Depends(get_db),
                                admin: Usuario = Depends(requer_admin)):
    """Define a partir de quando o banco de horas é contado. Admin only."""
    body = await request.json()
    data = (body.get("data") or "").strip()
    if not _parse_iso(data):
        raise HTTPException(400, "Data inválida")
    cfg = db.query(Configuracao).filter(Configuracao.chave == "banco_horas_inicio").first()
    if cfg:
        cfg.valor = data
    else:
        db.add(Configuracao(chave="banco_horas_inicio", valor=data,
                            descricao="Data de início do banco de horas"))
    db.commit()
    return {"status": "ok", "inicio_iso": data}


@app.get("/api/admin/limpar-ponto")
async def limpar_ponto(confirmar: str = "", db: Session = Depends(get_db),
                       admin: Usuario = Depends(requer_admin)):
    """APAGA PERMANENTEMENTE todos os registros de ponto (e justificativas/correções).
    Exige ?confirmar=APAGAR-TUDO p/ não disparar por acidente. Admin only. Irreversível."""
    if confirmar != "APAGAR-TUDO":
        return JSONResponse(status_code=400, content={
            "status": "confirmacao_necessaria",
            "aviso": "Isso APAGA TODOS os registros de ponto PERMANENTEMENTE (sem volta). "
                     "Para confirmar, adicione ?confirmar=APAGAR-TUDO no final da URL."})
    n_pontos = db.query(RegistroPonto).delete()
    n_just = db.query(JustificativaPonto).delete()
    n_corr = db.query(CorrecaoPonto).delete()
    db.commit()
    return JSONResponse({"status": "ok", "pontos_apagados": n_pontos,
                         "justificativas_apagadas": n_just, "correcoes_apagadas": n_corr})


# ─── Perfil do próprio usuário ───────────────────────────────────────────────────

@app.put("/api/me")
async def atualizar_perfil(
    request: Request,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    """Permite que qualquer usuário logado atualize seu nome e/ou senha."""
    body = await request.json()
    nome       = body.get("nome", "").strip()
    senha_atual = body.get("senha_atual", "")
    nova_senha  = body.get("nova_senha", "")

    if nome:
        usuario.nome = nome

    if nova_senha:
        if not senha_atual:
            raise HTTPException(status_code=400, detail="Informe a senha atual para trocar a senha")
        if not verificar_senha(senha_atual, usuario.senha_hash):
            raise HTTPException(status_code=401, detail="Senha atual incorreta")
        if len(nova_senha) < 6:
            raise HTTPException(status_code=400, detail="A nova senha deve ter ao mínimo 6 caracteres")
        usuario.senha_hash = hash_senha(nova_senha)

    db.commit()
    db.refresh(usuario)
    return {"id": usuario.id, "nome": usuario.nome, "email": usuario.email, "role": usuario.role}


# ─── Configurações do bot (admin) ────────────────────────────────────────────────

@app.get("/api/config")
async def listar_config(db: Session = Depends(get_db), admin: Usuario = Depends(requer_admin)):
    configs = db.query(Configuracao).order_by(Configuracao.chave).all()
    return [{"chave": c.chave, "valor": c.valor, "descricao": c.descricao} for c in configs]


@app.put("/api/config/{chave}")
async def atualizar_config(
    chave: str, request: Request,
    db: Session = Depends(get_db),
    admin: Usuario = Depends(requer_admin),
):
    body = await request.json()
    config = db.query(Configuracao).filter(Configuracao.chave == chave).first()
    if not config:
        raise HTTPException(status_code=404, detail="Configuração não encontrada")
    config.valor = body.get("valor", config.valor)
    config.atualizado_em = datetime.utcnow()
    db.commit()
    return {"status": "ok", "chave": chave}


# ─── Meta mensal com faixas de bônus ─────────────────────────────────────────────

def _chave_meta(ano: int, mes: int) -> str:
    return f"meta_{ano}_{mes:02d}"

def _get_meta_faixas(db: Session, ano: int, mes: int) -> list:
    import json as _json
    cfg = db.query(Configuracao).filter(Configuracao.chave == _chave_meta(ano, mes)).first()
    if cfg:
        try:
            return _json.loads(cfg.valor)
        except Exception:
            pass
    # fallback: usa meta_contratos genérica
    meta_cfg = db.query(Configuracao).filter(Configuracao.chave == "meta_contratos").first()
    meta_num = int(meta_cfg.valor) if meta_cfg and meta_cfg.valor.isdigit() else 20
    return [{"contratos": meta_num, "bonus": 0}]

@app.get("/api/meta-mensal")
async def get_meta_mensal(
    ano: int = None, mes: int = None,
    db: Session = Depends(get_db),
    admin: Usuario = Depends(requer_admin),
):
    if not ano or not mes:
        agora = _agora_br()
        ano, mes = agora.year, agora.month
    return {"ano": ano, "mes": mes, "faixas": _get_meta_faixas(db, ano, mes)}

@app.put("/api/meta-mensal")
async def set_meta_mensal(
    request: Request,
    db: Session = Depends(get_db),
    admin: Usuario = Depends(requer_admin),
):
    import json as _json
    body = await request.json()
    ano    = body.get("ano")
    mes    = body.get("mes")
    faixas = body.get("faixas", [])
    if not ano or not mes or not isinstance(faixas, list) or not faixas:
        raise HTTPException(status_code=400, detail="ano, mes e faixas são obrigatórios")
    # ordena por contratos crescente
    faixas = sorted(faixas, key=lambda f: f["contratos"])
    chave = _chave_meta(ano, mes)
    cfg = db.query(Configuracao).filter(Configuracao.chave == chave).first()
    if cfg:
        cfg.valor = _json.dumps(faixas)
    else:
        db.add(Configuracao(chave=chave, valor=_json.dumps(faixas),
                            descricao=f"Meta mensal {mes:02d}/{ano} — faixas de bônus"))
    # mantém meta_contratos compatível com a menor faixa
    menor = min(f["contratos"] for f in faixas)
    meta_cfg = db.query(Configuracao).filter(Configuracao.chave == "meta_contratos").first()
    if meta_cfg:
        meta_cfg.valor = str(menor)
    db.commit()
    return {"ano": ano, "mes": mes, "faixas": faixas}


# ─── Bancos (lista gerenciável) ───────────────────────────────────────────────────

def _get_bancos_lista(db: Session) -> list:
    import json as _json
    cfg = db.query(Configuracao).filter(Configuracao.chave == "bancos_lista").first()
    if not cfg:
        return []
    try:
        return _json.loads(cfg.valor)
    except Exception:
        return []

def _set_bancos_lista(db: Session, bancos: list):
    import json as _json
    cfg = db.query(Configuracao).filter(Configuracao.chave == "bancos_lista").first()
    if not cfg:
        cfg = Configuracao(chave="bancos_lista", descricao="Lista de bancos para o formulário de contrato")
        db.add(cfg)
    cfg.valor = _json.dumps(bancos, ensure_ascii=False)
    cfg.atualizado_em = datetime.utcnow()
    db.commit()

@app.get("/api/bancos")
async def listar_bancos(db: Session = Depends(get_db), usuario: Usuario = Depends(obter_usuario_atual)):
    return _get_bancos_lista(db)

@app.post("/api/bancos")
async def adicionar_banco(
    request: Request,
    db: Session = Depends(get_db),
    admin: Usuario = Depends(requer_admin),
):
    body = await request.json()
    nome = (body.get("nome") or "").strip().upper()
    if not nome:
        raise HTTPException(status_code=400, detail="Nome do banco é obrigatório")
    bancos = _get_bancos_lista(db)
    if nome not in bancos:
        bancos.append(nome)
        bancos.sort()
        _set_bancos_lista(db, bancos)
    return bancos

@app.delete("/api/bancos/{nome}")
async def remover_banco(
    nome: str,
    db: Session = Depends(get_db),
    admin: Usuario = Depends(requer_admin),
):
    bancos = _get_bancos_lista(db)
    nome_upper = nome.strip().upper()
    bancos = [b for b in bancos if b.upper() != nome_upper]
    _set_bancos_lista(db, bancos)
    return bancos


# ─── Relatórios (admin) ───────────────────────────────────────────────────────────

def _duracao_str(segundos: int) -> str:
    """Converte segundos em string legível ex: '2h 15min'."""
    if segundos < 60:
        return f"{segundos}s"
    m = segundos // 60
    if m < 60:
        return f"{m}min"
    h, rm = divmod(m, 60)
    return f"{h}h {rm}min" if rm else f"{h}h"


def _duracao_str_assinado(segundos: int) -> str:
    """Como _duracao_str, mas com sinal (+/−) — para banco de horas."""
    if segundos == 0:
        return "0h"
    sinal = "+" if segundos > 0 else "−"
    return sinal + _duracao_str(abs(int(segundos)))


# ─── Agendamentos / lembretes dentro de cada lead ─────────────────────────────────

def _dt_local_para_utc_naive(s: str) -> datetime:
    """Converte 'YYYY-MM-DDTHH:MM' (horário BR) para datetime UTC naive."""
    s = (s or "").strip().replace(" ", "T")
    fmt = "%Y-%m-%dT%H:%M:%S" if s.count(":") == 2 else "%Y-%m-%dT%H:%M"
    dt_br = datetime.strptime(s, fmt).replace(tzinfo=_TZ_BR)
    return dt_br.astimezone(timezone.utc).replace(tzinfo=None)


def _serial_agendamento(a: Agendamento, lead: Lead | None = None) -> dict:
    lead = lead or a.lead
    return {
        "id": a.id,
        "lead_id": a.lead_id,
        "lead_nome": (lead.nome if lead and lead.nome and lead.nome != "—" else (lead.telefone if lead else "")),
        "titulo": a.titulo,
        "descricao": a.descricao or "",
        "quando": _fmt_br(a.quando, "%d/%m/%Y %H:%M"),
        "quando_iso": _fmt_br(a.quando, "%Y-%m-%dT%H:%M"),
        "concluido": bool(a.concluido),
        "resultado": a.resultado or "",
        "criado_por": a.criado_por,
        "criado_por_nome": (a.criador.nome if a.criador else ""),
        "vencido": (not a.concluido) and a.quando <= datetime.utcnow(),
    }


def _pode_ver_agendamentos_alerta(a: Agendamento, lead: Lead, usuario: Usuario) -> bool:
    """Notificação (sino/popup) só p/ criador do agendamento ou responsável pelo lead."""
    return a.criado_por == usuario.id or (lead and lead.atribuido_para == usuario.id)


@app.get("/api/leads/{lead_id}/agendamentos")
async def listar_agendamentos(lead_id: int, db: Session = Depends(get_db),
                              usuario: Usuario = Depends(obter_usuario_atual)):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(404, "Lead não encontrado")
    ags = (db.query(Agendamento)
           .filter(Agendamento.lead_id == lead_id)
           .order_by(Agendamento.concluido, Agendamento.quando).all())
    return [_serial_agendamento(a, lead) for a in ags]


@app.post("/api/leads/{lead_id}/agendamentos")
async def criar_agendamento(lead_id: int, request: Request, db: Session = Depends(get_db),
                            usuario: Usuario = Depends(obter_usuario_atual)):
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(404, "Lead não encontrado")
    body = await request.json()
    titulo = (body.get("titulo") or "").strip()
    quando = (body.get("quando") or "").strip()
    if not titulo:
        raise HTTPException(400, "Descreva a ação a fazer")
    if not quando:
        raise HTTPException(400, "Informe data e hora")
    try:
        quando_utc = _dt_local_para_utc_naive(quando)
    except ValueError:
        raise HTTPException(400, "Data/hora inválida")
    ag = Agendamento(
        lead_id=lead_id, criado_por=usuario.id, titulo=titulo[:200],
        descricao=(body.get("descricao") or "").strip() or None, quando=quando_utc,
    )
    db.add(ag)
    db.commit()
    db.refresh(ag)
    return _serial_agendamento(ag, lead)


@app.patch("/api/agendamentos/{ag_id}")
async def editar_agendamento(ag_id: int, request: Request, db: Session = Depends(get_db),
                             usuario: Usuario = Depends(obter_usuario_atual)):
    ag = db.query(Agendamento).filter(Agendamento.id == ag_id).first()
    if not ag:
        raise HTTPException(404, "Agendamento não encontrado")
    body = await request.json()
    if "titulo" in body:
        t = (body.get("titulo") or "").strip()
        if not t:
            raise HTTPException(400, "Descreva a ação a fazer")
        ag.titulo = t[:200]
    if "descricao" in body:
        ag.descricao = (body.get("descricao") or "").strip() or None
    if body.get("quando"):
        try:
            ag.quando = _dt_local_para_utc_naive(body["quando"])
        except ValueError:
            raise HTTPException(400, "Data/hora inválida")
    if "resultado" in body:
        ag.resultado = (body.get("resultado") or "").strip() or None
    if "concluido" in body:
        concluir = bool(body["concluido"])
        marcando_agora = concluir and not ag.concluido
        ag.concluido = concluir
        ag.concluido_em = datetime.utcnow() if concluir else None
        # Ao concluir com um relato, registra automaticamente uma observação no lead
        resultado_txt = (body.get("resultado") or "").strip()
        if marcando_agora and resultado_txt:
            lead = db.query(Lead).filter(Lead.id == ag.lead_id).first()
            if lead:
                lista = _parse_observacoes(lead.observacoes)
                lista.append({
                    "texto": f"✅ Follow-up concluído — {ag.titulo}: {resultado_txt}",
                    "usuario": usuario.nome,
                    "em": datetime.utcnow().strftime("%d/%m/%Y %H:%M"),
                })
                lead.observacoes = json.dumps(lista, ensure_ascii=False)
                lead.atualizado_em = datetime.utcnow()
    db.commit()
    db.refresh(ag)
    return _serial_agendamento(ag)


@app.delete("/api/agendamentos/{ag_id}")
async def remover_agendamento(ag_id: int, db: Session = Depends(get_db),
                              usuario: Usuario = Depends(obter_usuario_atual)):
    ag = db.query(Agendamento).filter(Agendamento.id == ag_id).first()
    if not ag:
        raise HTTPException(404, "Agendamento não encontrado")
    db.delete(ag)
    db.commit()
    return {"status": "ok"}


@app.get("/api/agendamentos/meus")
async def meus_agendamentos(db: Session = Depends(get_db),
                            usuario: Usuario = Depends(obter_usuario_atual)):
    """Agenda geral. Admin vê tudo; funcionária vê o que criou OU de leads que assumiu."""
    q = (db.query(Agendamento, Lead)
         .join(Lead, Agendamento.lead_id == Lead.id))
    if usuario.role != RoleEnum.admin:
        q = q.filter((Agendamento.criado_por == usuario.id) | (Lead.atribuido_para == usuario.id))
    pares = q.order_by(Agendamento.quando).all()

    agora = datetime.utcnow()
    hoje_br = _agora_br().date()
    atrasados, hoje, proximos, concluidos = [], [], [], []
    for a, lead in pares:
        item = _serial_agendamento(a, lead)
        if a.concluido:
            concluidos.append(item)
        elif a.quando <= agora:
            atrasados.append(item)
        elif _fmt_br(a.quando, "%Y-%m-%d") == hoje_br.strftime("%Y-%m-%d"):
            hoje.append(item)
        else:
            proximos.append(item)
    # concluídos: mais recentes primeiro, limitado
    concluidos = list(reversed(concluidos))[:30]
    return {"atrasados": atrasados, "hoje": hoje, "proximos": proximos, "concluidos": concluidos}


@app.get("/api/agendamentos/alertas")
async def alertas_agendamentos(db: Session = Depends(get_db),
                               usuario: Usuario = Depends(obter_usuario_atual)):
    """Itens vencidos (não concluídos) p/ o sino/pop-up — só criador ou responsável."""
    agora = datetime.utcnow()
    q = (db.query(Agendamento, Lead)
         .join(Lead, Agendamento.lead_id == Lead.id)
         .filter(Agendamento.concluido == False,
                 Agendamento.quando <= agora)
         .filter((Agendamento.criado_por == usuario.id) | (Lead.atribuido_para == usuario.id))
         .order_by(Agendamento.quando))
    return [_serial_agendamento(a, lead) for a, lead in q.all()]


@app.get("/api/debug/webhooks")
async def debug_webhooks(db: Session = Depends(get_db),
                        admin: Usuario = Depends(requer_admin)):
    """Mostra os últimos webhooks recebidos do Z-API (para depurar ligações)."""
    return list(reversed(_DEBUG_WEBHOOKS))


@app.get("/api/debug/painel", response_class=HTMLResponse)
async def debug_painel(db: Session = Depends(get_db),
                       admin: Usuario = Depends(requer_admin)):
    """Tela amigável dos últimos eventos do Z-API — para diagnosticar ligações."""
    calls = 0
    linhas = ""
    for w in reversed(_DEBUG_WEBHOOKS):
        notif = str(w.get("notification") or "")
        tipo = str(w.get("type") or "")
        is_call = notif.upper().startswith("CALL") or "call" in tipo.lower()
        if is_call:
            calls += 1
        cor = "#fee2e2" if is_call else "#ffffff"
        rotulo = "📞 LIGAÇÃO" if is_call else (notif or tipo or "mensagem")
        linhas += (f"<tr style='background:{cor}'>"
                   f"<td>{w.get('em','')}</td>"
                   f"<td>{w.get('phone') or '—'}</td>"
                   f"<td><b>{rotulo}</b></td>"
                   f"<td style='color:#64748b'>type={tipo or '—'} | notification={notif or '—'}</td>"
                   f"</tr>")
    if not linhas:
        linhas = "<tr><td colspan='4' style='text-align:center;padding:1rem;color:#888'>Nenhum evento recebido ainda. Mande uma mensagem ou ligue e atualize.</td></tr>"
    banner_cor = "#16a34a" if calls else "#64748b"
    html = f"""<!doctype html><html lang="pt-br"><head><meta charset="utf-8">
    <meta http-equiv="refresh" content="5">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Diagnóstico de Ligações</title></head>
    <body style="font-family:system-ui,Segoe UI,Arial;margin:0;background:#f1f5f9;padding:1.2rem">
      <h2 style="margin:.2rem 0;color:#0d2b4e">🔎 Diagnóstico de eventos do WhatsApp</h2>
      <p style="color:#64748b;margin:.2rem 0 1rem">Esta tela atualiza sozinha a cada 5 segundos. Ligue para o número e observe se aparece uma linha vermelha <b>📞 LIGAÇÃO</b>.</p>
      <div style="display:inline-block;background:{banner_cor};color:#fff;font-weight:800;padding:.5rem 1rem;border-radius:10px;margin-bottom:1rem">
        Ligações detectadas: {calls}
      </div>
      <table style="width:100%;border-collapse:collapse;background:#fff;border-radius:10px;overflow:hidden;font-size:.9rem">
        <thead><tr style="background:#0d2b4e;color:#fff;text-align:left">
          <th style="padding:.5rem">Hora</th><th style="padding:.5rem">Telefone</th>
          <th style="padding:.5rem">Evento</th><th style="padding:.5rem">Detalhe técnico</th>
        </tr></thead>
        <tbody>{linhas}</tbody>
      </table>
    </body></html>"""
    return HTMLResponse(html)


@app.get("/api/chamadas/alertas")
async def alertas_chamadas(minutos: int = 5, db: Session = Depends(get_db),
                           usuario: Usuario = Depends(obter_usuario_atual)):
    """Ligações recebidas pelo WhatsApp nos últimos N minutos — para o alerta no painel."""
    desde = datetime.utcnow() - timedelta(minutes=max(1, min(minutos, 60)))
    msgs = (db.query(MensagemConversa)
            .filter(MensagemConversa.criado_em >= desde,
                    MensagemConversa.role == "user")
            .filter((MensagemConversa.conteudo.like("📞%")) | (MensagemConversa.conteudo.like("📹%")))
            .order_by(MensagemConversa.criado_em.desc()).all())
    if not msgs:
        return []
    tels = {m.telefone for m in msgs}
    leads = {l.telefone: l for l in db.query(Lead).filter(Lead.telefone.in_(tels)).all()}
    out = []
    for m in msgs:
        lead = leads.get(m.telefone)
        out.append({
            "id": m.id,
            "lead_id": lead.id if lead else None,
            "nome": (lead.nome if lead and lead.nome else m.telefone),
            "telefone": m.telefone,
            "perdida": "perdida" in (m.conteudo or ""),
            "texto": m.conteudo,
            "hora": _fmt_br(m.criado_em, "%H:%M"),
        })
    return out


# ─── Ponto: marcação e relatório ─────────────────────────────────────────────────

_PONTO_LABELS = {
    "entrada":      "Início da jornada",
    "saida_almoco": "Saída p/ almoço",
    "volta_almoco": "Volta do almoço",
    "saida":        "Fim da jornada",
}


def _intervalo_dia_utc(data_br):
    """Retorna (inicio_utc, fim_utc) naive para uma data no fuso BR (BR 00:00 = UTC 03:00)."""
    inicio = datetime(data_br.year, data_br.month, data_br.day) + timedelta(hours=3)
    return inicio, inicio + timedelta(days=1)


def _br_para_utc_naive(data_str: str, hora_str: str) -> datetime:
    """Converte data (YYYY-MM-DD) + hora (HH:MM) no fuso BR para datetime UTC naive (como o banco guarda)."""
    dt_br = datetime.strptime(f"{data_str} {hora_str}", "%Y-%m-%d %H:%M").replace(tzinfo=_TZ_BR)
    return dt_br.astimezone(timezone.utc).replace(tzinfo=None)


def _pontos_do_dia(db, usuario_id, data_br):
    ini, fim = _intervalo_dia_utc(data_br)
    return (db.query(RegistroPonto)
            .filter(RegistroPonto.usuario_id == usuario_id,
                    RegistroPonto.timestamp >= ini,
                    RegistroPonto.timestamp < fim)
            .order_by(RegistroPonto.timestamp).all())


def _proximas_acoes_ponto(pontos):
    ultimo = pontos[-1].tipo if pontos else None
    return {
        None:           ["entrada"],
        "entrada":      ["saida_almoco", "saida"],
        "saida_almoco": ["volta_almoco"],
        "volta_almoco": ["saida"],
        "saida":        ["entrada"],
    }.get(ultimo, ["entrada"])


def _jornadas_de_pontos(pontos, agora_utc):
    """Pareia pontos em intervalos de trabalho [(inicio, fim, aberto), ...].
    START = entrada|volta_almoco ; STOP = saida_almoco|saida."""
    STARTS = {"entrada", "volta_almoco"}
    STOPS  = {"saida_almoco", "saida"}
    intervalos = []
    abertura = None
    for p in pontos:
        if p.tipo in STARTS:
            if abertura is None:
                abertura = p.timestamp
        elif p.tipo in STOPS:
            if abertura is not None:
                intervalos.append((abertura, p.timestamp, False))
                abertura = None
    if abertura is not None:
        intervalos.append((abertura, agora_utc, True))
    return intervalos


def _tempo_ativo_intervalos(db, usuario_id, intervalos):
    """Conta pings de atividade dentro dos intervalos. Cada ping = 60s."""
    total = 0
    for ini, fim, _ in intervalos:
        n = (db.query(AtividadePing)
             .filter(AtividadePing.usuario_id == usuario_id,
                     AtividadePing.timestamp >= ini,
                     AtividadePing.timestamp < fim).count())
        total += n * 60
    return total


@app.post("/api/ponto")
async def bater_ponto(request: Request, db: Session = Depends(get_db),
                      usuario: Usuario = Depends(obter_usuario_atual)):
    body = await request.json()
    tipo = (body.get("tipo") or "").strip()
    if tipo not in _PONTO_LABELS:
        raise HTTPException(400, "Tipo de ponto inválido")
    if _agora_br().hour < 9:
        raise HTTPException(400, "O ponto só pode ser batido a partir das 09h.")
    hoje = _agora_br().date()
    pontos = _pontos_do_dia(db, usuario.id, hoje)
    validos = _proximas_acoes_ponto(pontos)
    if tipo not in validos:
        raise HTTPException(400, "Ação não permitida agora.")
    reg = RegistroPonto(usuario_id=usuario.id, tipo=tipo, ip=_ip_da_requisicao(request))
    db.add(reg)
    db.commit()
    db.refresh(reg)
    return {"status": "ok", "tipo": tipo, "label": _PONTO_LABELS[tipo],
            "hora": _fmt_br(reg.timestamp, "%H:%M")}


# ─── Justificativas de horário (atestado) com aprovação do admin ──────────────

_MAX_ATESTADO_BYTES = 15 * 1024 * 1024


def _iso_para_br(s: str) -> str:
    try:
        a, m, d = s.split("-")
        return f"{d}/{m}/{a}"
    except Exception:
        return s or ""


def _serial_justificativa(j: JustificativaPonto) -> dict:
    return {
        "id": j.id,
        "usuario_id": j.usuario_id,
        "funcionaria": (j.usuario.nome if j.usuario else "—"),
        "data": j.data,
        "data_br": _iso_para_br(j.data),
        "texto": j.texto or "",
        "filename": j.filename,
        "nome_arquivo": j.nome_arquivo,
        "status": j.status,
        "obs_admin": j.obs_admin or "",
        "aprovador": (j.aprovador.nome if j.aprovador else None),
        "criado_em": _fmt_br(j.criado_em, "%d/%m/%Y %H:%M") or "",
    }


@app.post("/api/ponto/justificativa")
async def criar_justificativa(
    data: str = Form(...), texto: str = Form(""),
    arquivo: UploadFile = File(None),
    db: Session = Depends(get_db), usuario: Usuario = Depends(obter_usuario_atual),
):
    """A funcionária lança a justificativa do dia (com atestado). Fica 'pendente' até o admin aprovar."""
    data = (data or "").strip()
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", data):
        raise HTTPException(400, "Data inválida.")
    texto = (texto or "").strip()
    filename = None
    nome_arq = None
    if arquivo is not None and arquivo.filename:
        dados = await arquivo.read()
        if dados:
            if len(dados) > _MAX_ATESTADO_BYTES:
                raise HTTPException(400, "Arquivo muito grande (máx. 15 MB).")
            nome_arq = (arquivo.filename or "atestado")[:200]
            ext = (nome_arq.rsplit(".", 1)[-1][:5].lower() if "." in nome_arq else "bin")
            ext = re.sub(r"[^a-z0-9]", "", ext) or "bin"
            filename = f"{uuid.uuid4().hex}.{ext}"
            _guardar_blob(db, filename, "documento", dados, nome_original=nome_arq,
                          mime=(arquivo.content_type or "application/octet-stream"), subdir="documentos")
    if not texto and not filename:
        raise HTTPException(400, "Escreva o motivo ou anexe o atestado.")
    # Reaproveita uma justificativa não aprovada do mesmo dia (reenvio); senão cria nova
    j = (db.query(JustificativaPonto)
         .filter(JustificativaPonto.usuario_id == usuario.id,
                 JustificativaPonto.data == data,
                 JustificativaPonto.status != "aprovada")
         .order_by(JustificativaPonto.id.desc()).first())
    if not j:
        j = JustificativaPonto(usuario_id=usuario.id, data=data)
        db.add(j)
    j.texto = texto
    if filename:
        j.filename = filename
        j.nome_arquivo = nome_arq
    j.status = "pendente"
    j.obs_admin = None
    j.aprovado_por = None
    j.aprovado_em = None
    db.commit()
    db.refresh(j)
    return _serial_justificativa(j)


@app.get("/api/ponto/justificativa/minhas")
async def minhas_justificativas(db: Session = Depends(get_db),
                                usuario: Usuario = Depends(obter_usuario_atual)):
    js = (db.query(JustificativaPonto)
          .filter(JustificativaPonto.usuario_id == usuario.id)
          .order_by(JustificativaPonto.data.desc(), JustificativaPonto.id.desc()).all())
    return [_serial_justificativa(j) for j in js]


@app.get("/api/ponto/justificativas")
async def listar_justificativas(status: str = "", db: Session = Depends(get_db),
                                usuario: Usuario = Depends(requer_admin)):
    q = db.query(JustificativaPonto)
    if status in ("pendente", "aprovada", "rejeitada"):
        q = q.filter(JustificativaPonto.status == status)
    js = q.order_by(JustificativaPonto.criado_em.desc()).all()
    return [_serial_justificativa(j) for j in js]


@app.get("/api/ponto/justificativas/pendentes-count")
async def justificativas_pendentes_count(db: Session = Depends(get_db),
                                         usuario: Usuario = Depends(requer_admin)):
    n = db.query(JustificativaPonto).filter(JustificativaPonto.status == "pendente").count()
    return {"pendentes": n}


@app.post("/api/ponto/justificativa/{jid}/aprovar")
async def aprovar_justificativa(jid: int, db: Session = Depends(get_db),
                                usuario: Usuario = Depends(requer_admin)):
    j = db.query(JustificativaPonto).filter(JustificativaPonto.id == jid).first()
    if not j:
        raise HTTPException(404, "Justificativa não encontrada.")
    j.status = "aprovada"
    j.aprovado_por = usuario.id
    j.aprovado_em = datetime.utcnow()
    j.obs_admin = None
    db.commit()
    db.refresh(j)
    return _serial_justificativa(j)


@app.post("/api/ponto/justificativa/{jid}/rejeitar")
async def rejeitar_justificativa(jid: int, request: Request, db: Session = Depends(get_db),
                                 usuario: Usuario = Depends(requer_admin)):
    body = await request.json()
    j = db.query(JustificativaPonto).filter(JustificativaPonto.id == jid).first()
    if not j:
        raise HTTPException(404, "Justificativa não encontrada.")
    j.status = "rejeitada"
    j.aprovado_por = usuario.id
    j.aprovado_em = datetime.utcnow()
    j.obs_admin = (body.get("obs") or "").strip()[:400] or None
    db.commit()
    db.refresh(j)
    return _serial_justificativa(j)


@app.get("/api/ponto/hoje")
async def ponto_hoje(db: Session = Depends(get_db),
                     usuario: Usuario = Depends(obter_usuario_atual)):
    """Espelho do ponto do dia + próximas ações para a própria usuária."""
    hoje = _agora_br().date()
    pontos = _pontos_do_dia(db, usuario.id, hoje)
    return {
        "data": hoje.strftime("%d/%m/%Y"),
        "registros": [
            {"tipo": p.tipo, "label": _PONTO_LABELS[p.tipo], "hora": _fmt_br(p.timestamp, "%H:%M")}
            for p in pontos
        ],
        "proximas_acoes": [
            {"tipo": t, "label": _PONTO_LABELS[t]} for t in _proximas_acoes_ponto(pontos)
        ],
    }


@app.get("/api/ponto/espelho")
async def ponto_espelho(dias: int = 30, db: Session = Depends(get_db),
                        usuario: Usuario = Depends(obter_usuario_atual)):
    """Histórico do espelho de ponto da própria usuária (últimos N dias com registro)."""
    hoje = _agora_br().date()
    agora_utc = datetime.utcnow()
    resultado = []
    for i in range(max(1, min(dias, 90))):
        dia = hoje - timedelta(days=i)
        pontos = _pontos_do_dia(db, usuario.id, dia)
        if not pontos:
            continue
        intervalos = _jornadas_de_pontos(pontos, agora_utc)
        total_jornada = sum(int((f - ini).total_seconds()) for ini, f, _ in intervalos)
        resultado.append({
            "data": dia.strftime("%d/%m/%Y"),
            "registros": [
                {"tipo": p.tipo, "label": _PONTO_LABELS[p.tipo], "hora": _fmt_br(p.timestamp, "%H:%M")}
                for p in pontos
            ],
            "total_jornada": _duracao_str(total_jornada),
        })
    return resultado


@app.get("/api/ponto/relatorio")
async def ponto_relatorio(data: str = None, db: Session = Depends(get_db),
                          admin: Usuario = Depends(requer_admin)):
    """Relatório admin: jornada registrada × tempo ativo × ociosidade por funcionária num dia."""
    if data:
        try:
            d = datetime.strptime(data, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(400, "Data inválida (use YYYY-MM-DD)")
    else:
        d = _agora_br().date()
    agora_utc = datetime.utcnow()
    funcionarias = (db.query(Usuario)
                    .filter(Usuario.role == RoleEnum.funcionario)
                    .order_by(Usuario.nome).all())
    resultado = []
    _dia_iso = d.strftime("%Y-%m-%d")
    for u in funcionarias:
        pontos = _pontos_do_dia(db, u.id, d)
        _justs_dia = db.query(JustificativaPonto).filter(
            JustificativaPonto.usuario_id == u.id, JustificativaPonto.data == _dia_iso
        ).order_by(JustificativaPonto.id.desc()).all()
        _corrs_dia = db.query(CorrecaoPonto).filter(
            CorrecaoPonto.usuario_id == u.id, CorrecaoPonto.data == _dia_iso
        ).order_by(CorrecaoPonto.criado_em.desc()).all()
        # Mostra a funcionária se tem ponto OU justificativa OU correção no dia
        if not pontos and not _justs_dia and not _corrs_dia:
            continue
        intervalos = _jornadas_de_pontos(pontos, agora_utc)
        jornada_s = sum(int((f - ini).total_seconds()) for ini, f, _ in intervalos)
        ativo_s = min(_tempo_ativo_intervalos(db, u.id, intervalos), jornada_s)
        ocioso_s = max(0, jornada_s - ativo_s)
        resultado.append({
            "usuario": u.nome,
            "usuario_id": u.id,
            "registros": [
                {"id": p.id, "tipo": p.tipo, "label": _PONTO_LABELS[p.tipo],
                 "hora": _fmt_br(p.timestamp, "%H:%M"), "foto": p.foto_filename}
                for p in pontos
            ],
            "justificativas": [_serial_justificativa(j) for j in _justs_dia],
            "correcoes": [_serial_correcao(c) for c in _corrs_dia],
            "jornada": _duracao_str(jornada_s),
            "jornada_s": jornada_s,
            "ativo": _duracao_str(ativo_s),
            "ativo_s": ativo_s,
            "ocioso": _duracao_str(ocioso_s),
            "perc_ativo": round(ativo_s / jornada_s * 100) if jornada_s else 0,
            "em_andamento": any(a for _, _, a in intervalos),
        })
    return {
        "data": d.strftime("%d/%m/%Y"),
        "data_iso": d.strftime("%Y-%m-%d"),
        "funcionarias": resultado,
        "funcionarias_todas": [{"id": u.id, "nome": u.nome} for u in funcionarias],
    }


# ─── Ponto: correções pelo admin (criar / editar / remover) ───────────────────

_PONTO_TIPOS_VALIDOS = set(_PONTO_LABELS.keys())


def _serial_correcao(c: CorrecaoPonto) -> dict:
    return {
        "id": c.id,
        "usuario_id": c.usuario_id,
        "funcionaria": (c.usuario.nome if c.usuario else "—"),
        "solicitante": (c.solicitante.nome if c.solicitante else "—"),
        "data": c.data,
        "data_br": _iso_para_br(c.data),
        "acao": c.acao,
        "tipo_ponto": c.tipo_ponto,
        "tipo_label": _PONTO_LABELS.get(c.tipo_ponto, c.tipo_ponto or "—"),
        "hora_anterior": c.hora_anterior,
        "hora_nova": c.hora_nova,
        "registro_id": c.registro_id,
        "motivo": c.motivo or "",
        "status": c.status,
        "origem": c.origem,
        "obs_admin": c.obs_admin or "",
        "resolvido_por": (c.resolvedor.nome if c.resolvedor else None),
        "criado_em": _fmt_br(c.criado_em, "%d/%m/%Y %H:%M") or "",
        "resolvido_em": _fmt_br(c.resolvido_em, "%d/%m/%Y %H:%M") or "",
    }


def _registrar_correcao(db, *, usuario_id, solicitante_id, data, acao, tipo_ponto,
                        hora_anterior, hora_nova, registro_id, motivo, status,
                        origem, resolvido_por=None):
    """Grava a trilha de auditoria de uma correção de ponto."""
    c = CorrecaoPonto(
        usuario_id=usuario_id, solicitante_id=solicitante_id, data=data, acao=acao,
        tipo_ponto=tipo_ponto, hora_anterior=hora_anterior, hora_nova=hora_nova,
        registro_id=registro_id, motivo=(motivo or None), status=status, origem=origem,
        resolvido_por=resolvido_por,
        resolvido_em=(datetime.utcnow() if status in ("aplicada", "rejeitada") else None),
    )
    db.add(c)
    return c


@app.post("/api/ponto/admin")
async def admin_criar_ponto(request: Request, db: Session = Depends(get_db),
                            admin: Usuario = Depends(requer_admin)):
    """Admin registra/corrige um ponto em nome de uma funcionária."""
    body = await request.json()
    try:
        usuario_id = int(body.get("usuario_id"))
    except (TypeError, ValueError):
        raise HTTPException(400, "Funcionária inválida")
    tipo = (body.get("tipo") or "").strip()
    data = (body.get("data") or "").strip()
    hora = (body.get("hora") or "").strip()
    if tipo not in _PONTO_TIPOS_VALIDOS:
        raise HTTPException(400, "Tipo de ponto inválido")
    if not db.query(Usuario).filter(Usuario.id == usuario_id).first():
        raise HTTPException(404, "Funcionária não encontrada")
    try:
        ts = _br_para_utc_naive(data, hora)
    except ValueError:
        raise HTTPException(400, "Data ou hora inválida")
    reg = RegistroPonto(usuario_id=usuario_id, tipo=tipo, timestamp=ts, ip="admin")
    db.add(reg)
    db.flush()
    _registrar_correcao(db, usuario_id=usuario_id, solicitante_id=admin.id,
                        data=data, acao="adicionar", tipo_ponto=tipo,
                        hora_anterior=None, hora_nova=hora, registro_id=reg.id,
                        motivo=(body.get("motivo") or "").strip(), status="aplicada",
                        origem="admin", resolvido_por=admin.id)
    db.commit()
    return {"status": "ok"}


@app.patch("/api/ponto/{ponto_id}")
async def admin_editar_ponto(ponto_id: int, request: Request, db: Session = Depends(get_db),
                             admin: Usuario = Depends(requer_admin)):
    """Admin corrige o horário e/ou o tipo de um ponto existente."""
    reg = db.query(RegistroPonto).filter(RegistroPonto.id == ponto_id).first()
    if not reg:
        raise HTTPException(404, "Registro não encontrado")
    body = await request.json()
    # Valores ANTES da correção (para a trilha de auditoria)
    data_antes = _fmt_br(reg.timestamp, "%Y-%m-%d")
    hora_antes = _fmt_br(reg.timestamp, "%H:%M")
    tipo_antes = reg.tipo
    tipo = (body.get("tipo") or "").strip()
    if tipo:
        if tipo not in _PONTO_TIPOS_VALIDOS:
            raise HTTPException(400, "Tipo de ponto inválido")
        reg.tipo = tipo
    data = (body.get("data") or "").strip()
    hora = (body.get("hora") or "").strip()
    if data and hora:
        try:
            reg.timestamp = _br_para_utc_naive(data, hora)
        except ValueError:
            raise HTTPException(400, "Data ou hora inválida")
    _registrar_correcao(db, usuario_id=reg.usuario_id, solicitante_id=admin.id,
                        data=(data or data_antes), acao="editar", tipo_ponto=reg.tipo,
                        hora_anterior=hora_antes, hora_nova=(hora or hora_antes),
                        registro_id=reg.id,
                        motivo=(body.get("motivo") or "").strip(), status="aplicada",
                        origem="admin", resolvido_por=admin.id)
    db.commit()
    return {"status": "ok"}


@app.delete("/api/ponto/{ponto_id}")
async def admin_remover_ponto(ponto_id: int, motivo: str = "", db: Session = Depends(get_db),
                              admin: Usuario = Depends(requer_admin)):
    """Admin remove um ponto registrado por engano (fica registrado na auditoria)."""
    reg = db.query(RegistroPonto).filter(RegistroPonto.id == ponto_id).first()
    if not reg:
        raise HTTPException(404, "Registro não encontrado")
    _registrar_correcao(db, usuario_id=reg.usuario_id, solicitante_id=admin.id,
                        data=_fmt_br(reg.timestamp, "%Y-%m-%d"), acao="remover",
                        tipo_ponto=reg.tipo, hora_anterior=_fmt_br(reg.timestamp, "%H:%M"),
                        hora_nova=None, registro_id=reg.id, motivo=(motivo or "").strip(),
                        status="aplicada", origem="admin", resolvido_por=admin.id)
    db.delete(reg)
    db.commit()
    return {"status": "ok"}


# ─── Ponto: solicitação de correção pela funcionária + aprovação do admin ─────

@app.post("/api/ponto/correcao")
async def solicitar_correcao(request: Request, db: Session = Depends(get_db),
                             usuario: Usuario = Depends(obter_usuario_atual)):
    """A funcionária SOLICITA uma correção (não altera o ponto). Fica pendente para o admin."""
    body = await request.json()
    data = (body.get("data") or "").strip()
    acao = (body.get("acao") or "").strip()
    tipo_ponto = (body.get("tipo_ponto") or "").strip() or None
    hora = (body.get("hora") or "").strip() or None
    motivo = (body.get("motivo") or "").strip()
    registro_id = body.get("registro_id")
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", data):
        raise HTTPException(400, "Data inválida.")
    if acao not in ("adicionar", "editar", "remover"):
        raise HTTPException(400, "Ação inválida.")
    if not motivo:
        raise HTTPException(400, "Descreva o motivo da correção.")
    if acao in ("adicionar", "editar"):
        if tipo_ponto and tipo_ponto not in _PONTO_TIPOS_VALIDOS:
            raise HTTPException(400, "Tipo de ponto inválido.")
        if not hora or not re.match(r"^\d{2}:\d{2}$", hora):
            raise HTTPException(400, "Informe o horário (HH:MM).")
    # Para editar/remover, o registro precisa ser da própria funcionária
    if registro_id is not None:
        try:
            registro_id = int(registro_id)
        except (TypeError, ValueError):
            raise HTTPException(400, "Registro inválido.")
        alvo = db.query(RegistroPonto).filter(RegistroPonto.id == registro_id).first()
        if not alvo or alvo.usuario_id != usuario.id:
            raise HTTPException(404, "Ponto não encontrado.")
    hora_anterior = None
    if registro_id is not None and acao in ("editar", "remover"):
        alvo = db.query(RegistroPonto).filter(RegistroPonto.id == registro_id).first()
        if alvo:
            hora_anterior = _fmt_br(alvo.timestamp, "%H:%M")
            if not tipo_ponto:
                tipo_ponto = alvo.tipo
    c = CorrecaoPonto(
        usuario_id=usuario.id, solicitante_id=usuario.id, data=data, acao=acao,
        tipo_ponto=tipo_ponto, hora_anterior=hora_anterior,
        hora_nova=(hora if acao in ("adicionar", "editar") else None),
        registro_id=registro_id, motivo=motivo, status="pendente", origem="funcionaria",
    )
    db.add(c)
    db.commit()
    db.refresh(c)
    return _serial_correcao(c)


@app.get("/api/ponto/correcao/minhas")
async def minhas_correcoes(db: Session = Depends(get_db),
                           usuario: Usuario = Depends(obter_usuario_atual)):
    cs = (db.query(CorrecaoPonto)
          .filter(CorrecaoPonto.usuario_id == usuario.id)
          .order_by(CorrecaoPonto.criado_em.desc()).all())
    return [_serial_correcao(c) for c in cs]


@app.get("/api/ponto/meus")
async def meus_pontos_do_dia(data: str = "", db: Session = Depends(get_db),
                             usuario: Usuario = Depends(obter_usuario_atual)):
    """Pontos da própria funcionária num dia (para escolher qual corrigir/remover)."""
    try:
        d = datetime.strptime(data, "%Y-%m-%d").date() if data else _agora_br().date()
    except ValueError:
        raise HTTPException(400, "Data inválida")
    pontos = _pontos_do_dia(db, usuario.id, d)
    return [{"id": p.id, "tipo": p.tipo, "label": _PONTO_LABELS[p.tipo],
             "hora": _fmt_br(p.timestamp, "%H:%M")} for p in pontos]


@app.get("/api/ponto/correcoes")
async def listar_correcoes(status: str = "", db: Session = Depends(get_db),
                           admin: Usuario = Depends(requer_admin)):
    q = db.query(CorrecaoPonto)
    if status in ("pendente", "aplicada", "rejeitada"):
        q = q.filter(CorrecaoPonto.status == status)
    cs = q.order_by(CorrecaoPonto.criado_em.desc()).all()
    return [_serial_correcao(c) for c in cs]


@app.get("/api/ponto/correcoes/pendentes-count")
async def correcoes_pendentes_count(db: Session = Depends(get_db),
                                    admin: Usuario = Depends(requer_admin)):
    n = db.query(CorrecaoPonto).filter(CorrecaoPonto.status == "pendente",
                                       CorrecaoPonto.origem == "funcionaria").count()
    return {"pendentes": n}


@app.post("/api/ponto/correcao/{cid}/aplicar")
async def aplicar_correcao(cid: int, db: Session = Depends(get_db),
                           admin: Usuario = Depends(requer_admin)):
    """Admin aprova a solicitação E aplica a mudança no ponto (com trilha de auditoria)."""
    c = db.query(CorrecaoPonto).filter(CorrecaoPonto.id == cid).first()
    if not c:
        raise HTTPException(404, "Solicitação não encontrada.")
    if c.status != "pendente":
        raise HTTPException(400, "Esta solicitação já foi resolvida.")
    if c.acao == "adicionar":
        if not c.tipo_ponto or not c.hora_nova:
            raise HTTPException(400, "Solicitação incompleta para adicionar.")
        try:
            ts = _br_para_utc_naive(c.data, c.hora_nova)
        except ValueError:
            raise HTTPException(400, "Data/hora inválida na solicitação.")
        reg = RegistroPonto(usuario_id=c.usuario_id, tipo=c.tipo_ponto, timestamp=ts, ip="correcao")
        db.add(reg)
        db.flush()
        c.registro_id = reg.id
    elif c.acao == "editar":
        reg = db.query(RegistroPonto).filter(RegistroPonto.id == c.registro_id).first()
        if not reg:
            raise HTTPException(404, "O ponto a corrigir não existe mais.")
        if c.tipo_ponto:
            reg.tipo = c.tipo_ponto
        if c.hora_nova:
            try:
                reg.timestamp = _br_para_utc_naive(c.data, c.hora_nova)
            except ValueError:
                raise HTTPException(400, "Data/hora inválida na solicitação.")
    elif c.acao == "remover":
        reg = db.query(RegistroPonto).filter(RegistroPonto.id == c.registro_id).first()
        if reg:
            db.delete(reg)
    c.status = "aplicada"
    c.resolvido_por = admin.id
    c.resolvido_em = datetime.utcnow()
    db.commit()
    db.refresh(c)
    return _serial_correcao(c)


@app.post("/api/ponto/correcao/{cid}/rejeitar")
async def rejeitar_correcao(cid: int, request: Request, db: Session = Depends(get_db),
                            admin: Usuario = Depends(requer_admin)):
    body = await request.json()
    c = db.query(CorrecaoPonto).filter(CorrecaoPonto.id == cid).first()
    if not c:
        raise HTTPException(404, "Solicitação não encontrada.")
    if c.status != "pendente":
        raise HTTPException(400, "Esta solicitação já foi resolvida.")
    c.status = "rejeitada"
    c.obs_admin = (body.get("obs") or "").strip()[:400] or None
    c.resolvido_por = admin.id
    c.resolvido_em = datetime.utcnow()
    db.commit()
    db.refresh(c)
    return _serial_correcao(c)


@app.get("/api/relatorios")
async def relatorios(db: Session = Depends(get_db), admin: Usuario = Depends(requer_admin)):
    usuarios = db.query(Usuario).filter(Usuario.ativo == True, Usuario.role == RoleEnum.funcionario).all()
    _ign = Lead.ignorar_relatorios.isnot(True)
    por_funcionario = []
    for u in usuarios:
        total = db.query(Lead).filter(Lead.atribuido_para == u.id, _ign).count()
        fechados = db.query(Lead).filter(Lead.atribuido_para == u.id, Lead.status == StatusLeadEnum.fechado, _ign).count()
        por_funcionario.append({
            "nome": u.nome,
            "role": u.role,
            "total_assumidos": total,
            "fechados": fechados,
            "taxa": round((fechados / total * 100), 1) if total > 0 else 0,
        })
    return {
        "por_funcionario": por_funcionario,
        "por_modalidade": {
            "financiamento": db.query(Lead).filter(Lead.modalidade == "financiamento", _ign).count(),
            "refinanciamento": db.query(Lead).filter(Lead.modalidade == "refinanciamento", _ign).count(),
        },
    }


@app.get("/api/relatorio/produtividade")
async def relatorio_produtividade(periodo: str = "tudo",
                                  db: Session = Depends(get_db),
                                  admin: Usuario = Depends(requer_admin)):
    """Por operadora (responsável pelo lead): quantos dos leads dela JÁ TÊM e quantos
    AINDA NÃO TÊM observação e agendamento (follow-up manual). Conta individualmente.
    'periodo' filtra pelos leads que ENTRARAM no período (criado_em)."""
    _ign = Lead.ignorar_relatorios.isnot(True)
    desde = _inicio_periodo(periodo)
    # Leads que entram em relatório: (id, responsável, observações)
    q = db.query(Lead.id, Lead.atribuido_para, Lead.observacoes).filter(_ign)
    if desde is not None:
        q = q.filter(Lead.criado_em >= desde)
    leads = q.all()
    # IDs de leads com ao menos 1 agendamento (= follow-up inserido pela operadora)
    com_agendamento = {lid for (lid,) in db.query(Agendamento.lead_id).distinct().all()}
    # Operadoras (funcionárias) — inclui inativas só se tiverem leads
    operadoras = db.query(Usuario).filter(Usuario.role == RoleEnum.funcionario).all()

    stats = {u.id: {"id": u.id, "nome": u.nome, "_ativo": bool(u.ativo), "total": 0,
                    "com_obs": 0, "sem_obs": 0, "com_followup": 0, "sem_followup": 0}
             for u in operadoras}

    for lid, resp, obs in leads:
        s = stats.get(resp)
        if s is None:
            continue   # lead sem responsável (ou de admin/usuário fora da lista) → não conta
        s["total"] += 1
        if len(_parse_observacoes(obs)) > 0:
            s["com_obs"] += 1
        else:
            s["sem_obs"] += 1
        if lid in com_agendamento:
            s["com_followup"] += 1
        else:
            s["sem_followup"] += 1

    # Mostra operadoras ativas (mesmo com 0) e inativas só se tiverem leads
    lista = [s for s in stats.values() if s["_ativo"] or s["total"] > 0]
    for s in lista:
        s.pop("_ativo", None)
    lista.sort(key=lambda x: x["nome"].lower())

    totais = {"total": 0, "com_obs": 0, "sem_obs": 0, "com_followup": 0, "sem_followup": 0}
    for s in lista:
        for k in totais:
            totais[k] += s[k]

    return {"operadoras": lista, "totais": totais}


@app.get("/api/relatorio/eficiencia")
async def relatorio_eficiencia(periodo: str = "tudo",
                               db: Session = Depends(get_db),
                               admin: Usuario = Depends(requer_admin)):
    """Funil de conversão por operadora: recebidos → proposta → aprovados → fechados, + perdidos.
    Contagem CUMULATIVA por status atual: quem está (ou passou) num estágio à frente conta nos
    anteriores. Obs.: só temos o status ATUAL — um lead perdido conta só em 'perdidos' (não dá
    pra saber até onde avançou antes de ser perdido).
    'periodo' filtra pelos leads que ENTRARAM no período (criado_em)."""
    _ign = Lead.ignorar_relatorios.isnot(True)
    desde = _inicio_periodo(periodo)
    # "Recebidos" = leads atribuídos à operadora, fora os de teste e os desqualificados (spam/inválidos)
    q = (db.query(Lead.atribuido_para, Lead.status)
         .filter(_ign, Lead.status != StatusLeadEnum.desqualificado))
    if desde is not None:
        q = q.filter(Lead.criado_em >= desde)
    leads = q.all()
    operadoras = db.query(Usuario).filter(Usuario.role == RoleEnum.funcionario).all()

    # Estágios cumulativos (quem fechou também passou por proposta e aprovação)
    ATINGIU_PROPOSTA = {StatusLeadEnum.proposta_enviada.value,
                        StatusLeadEnum.proposta_aprovada.value,
                        StatusLeadEnum.fechado.value}
    ATINGIU_APROVADA = {StatusLeadEnum.proposta_aprovada.value,
                        StatusLeadEnum.fechado.value}

    stats = {u.id: {"id": u.id, "nome": u.nome, "_ativo": bool(u.ativo),
                    "recebidos": 0, "proposta": 0, "aprovados": 0,
                    "fechados": 0, "perdidos": 0}
             for u in operadoras}

    for resp, status in leads:
        s = stats.get(resp)
        if s is None:
            continue
        st = status.value if hasattr(status, "value") else status
        s["recebidos"] += 1
        if st in ATINGIU_PROPOSTA:
            s["proposta"] += 1
        if st in ATINGIU_APROVADA:
            s["aprovados"] += 1
        if st == StatusLeadEnum.fechado.value:
            s["fechados"] += 1
        if st == StatusLeadEnum.perdido.value:
            s["perdidos"] += 1

    lista = [s for s in stats.values() if s["_ativo"] or s["recebidos"] > 0]
    for s in lista:
        s.pop("_ativo", None)
        r = s["recebidos"]
        s["taxa_fechamento"] = round(s["fechados"] / r * 100, 1) if r > 0 else 0.0
    lista.sort(key=lambda x: x["recebidos"], reverse=True)

    totais = {"recebidos": 0, "proposta": 0, "aprovados": 0, "fechados": 0, "perdidos": 0}
    for s in lista:
        for k in totais:
            totais[k] += s[k]
    totais["taxa_fechamento"] = (round(totais["fechados"] / totais["recebidos"] * 100, 1)
                                 if totais["recebidos"] > 0 else 0.0)

    return {"operadoras": lista, "totais": totais}


def _dias_horas(seg) -> str:
    """Formata uma duração (segundos) de forma amigável: '12d 3h', '5h', '40min'."""
    seg = max(0, int(seg))
    dias = seg // 86400
    horas = (seg % 86400) // 3600
    if dias >= 1:
        return f"{dias}d {horas}h" if horas else f"{dias}d"
    if horas >= 1:
        return f"{horas}h"
    return f"{max(1, seg // 60)}min"


@app.get("/api/relatorio/fechamentos")
async def relatorio_fechamentos(periodo: str = "mes",
                                db: Session = Depends(get_db),
                                admin: Usuario = Depends(requer_admin)):
    """Quem REALMENTE fechou (pelo histórico) + tempo médio pra fechar.
    Conta pela DATA DE FECHAMENTO (evento -> 'fechado'), NÃO pela chegada: um lead que
    chegou em maio e fechou em junho conta no desempenho de junho. 'quem fechou' = quem
    registrou o fechamento (no passado/backfill = a operadora atribuída; exato daqui pra frente)."""
    from collections import defaultdict
    desde = _inicio_periodo(periodo)
    q = db.query(HistoricoLead).filter(HistoricoLead.para_status == StatusLeadEnum.fechado.value)
    if desde is not None:
        q = q.filter(HistoricoLead.quando >= desde)
    eventos = q.all()

    # Último evento de fechamento por lead (evita duplicar se reabriu e fechou de novo)
    por_lead = {}
    for e in eventos:
        cur = por_lead.get(e.lead_id)
        if cur is None or (e.quando and cur.quando and e.quando > cur.quando):
            por_lead[e.lead_id] = e
    if not por_lead:
        return {"periodo": periodo, "fechamentos": [], "total": {"qtd": 0, "tempo_medio": "—"}}

    leads = {l.id: l for l in db.query(Lead.id, Lead.criado_em, Lead.ignorar_relatorios)
             .filter(Lead.id.in_(list(por_lead.keys()))).all()}
    uids = {e.usuario_id for e in por_lead.values() if e.usuario_id}
    nomes = dict(db.query(Usuario.id, Usuario.nome).filter(Usuario.id.in_(uids)).all()) if uids else {}

    agg = defaultdict(lambda: {"qtd": 0, "soma_seg": 0, "com_tempo": 0})
    for lid, e in por_lead.items():
        l = leads.get(lid)
        if l is None or l.ignorar_relatorios:
            continue
        a = agg[e.usuario_id]
        a["qtd"] += 1
        if l.criado_em and e.quando and e.quando >= l.criado_em:
            a["soma_seg"] += (e.quando - l.criado_em).total_seconds()
            a["com_tempo"] += 1

    fechamentos = []
    for uid, a in agg.items():
        media = (a["soma_seg"] / a["com_tempo"]) if a["com_tempo"] else 0
        fechamentos.append({
            "id": uid,
            "nome": (nomes.get(uid) if uid else "Sistema/Bot") or f"Usuário {uid}",
            "qtd": a["qtd"],
            "tempo_medio": _dias_horas(media) if a["com_tempo"] else "—",
            "tempo_medio_seg": int(media),
        })
    fechamentos.sort(key=lambda x: x["qtd"], reverse=True)

    tot_qtd = sum(a["qtd"] for a in agg.values())
    tot_soma = sum(a["soma_seg"] for a in agg.values())
    tot_com = sum(a["com_tempo"] for a in agg.values())
    total = {"qtd": tot_qtd, "tempo_medio": _dias_horas(tot_soma / tot_com) if tot_com else "—"}
    return {"periodo": periodo, "fechamentos": fechamentos, "total": total}


@app.get("/api/relatorio/origem")
async def relatorio_origem(periodo: str = "tudo",
                           db: Session = Depends(get_db),
                           admin: Usuario = Depends(requer_admin)):
    """Por operadora: quantos leads vieram de PARCEIRO (cartela) x NOVOS (whatsapp/anúncio/etc).
    Ajuda a comparar volume de forma justa — quem tem cartela de parceiros maior recebe mais.
    Mesmos filtros do funil de eficiência, então 'Total' bate com 'Recebidos'."""
    _ign = Lead.ignorar_relatorios.isnot(True)
    desde = _inicio_periodo(periodo)
    q = (db.query(Lead.atribuido_para, Lead.parceiro_id, Lead.origem)
         .filter(_ign, Lead.status != StatusLeadEnum.desqualificado))
    if desde is not None:
        q = q.filter(Lead.criado_em >= desde)
    leads = q.all()

    operadoras = db.query(Usuario).filter(Usuario.role == RoleEnum.funcionario).all()
    stats = {u.id: {"id": u.id, "nome": u.nome, "_ativo": bool(u.ativo),
                    "total": 0, "de_parceiro": 0, "novos": 0}
             for u in operadoras}
    for resp, pid, origem in leads:
        s = stats.get(resp)
        if s is None:
            continue
        s["total"] += 1
        if pid is not None or origem == "parceiro":
            s["de_parceiro"] += 1
        else:
            s["novos"] += 1

    lista = [s for s in stats.values() if s["_ativo"] or s["total"] > 0]
    for s in lista:
        s.pop("_ativo", None)
    lista.sort(key=lambda x: x["total"], reverse=True)
    totais = {"total": 0, "de_parceiro": 0, "novos": 0}
    for s in lista:
        for k in totais:
            totais[k] += s[k]
    return {"operadoras": lista, "totais": totais}


@app.get("/api/relatorio/sem-proximo-passo")
async def relatorio_sem_proximo_passo(db: Session = Depends(get_db),
                                      admin: Usuario = Depends(requer_admin)):
    """Leads que ALGUÉM está trabalhando (assumido/pré-análise/proposta) e que NÃO têm
    nenhum agendamento PENDENTE = sem próximo passo. Por operadora — risco de esfriar."""
    ATIVOS = [StatusLeadEnum.assumido.value, StatusLeadEnum.pre_analise.value,
              StatusLeadEnum.proposta_enviada.value, StatusLeadEnum.proposta_aprovada.value]
    leads = (db.query(Lead)
             .filter(Lead.status.in_(ATIVOS),
                     Lead.ignorar_relatorios.isnot(True),
                     Lead.atribuido_para.isnot(None)).all())
    com_pendente = {lid for (lid,) in db.query(Agendamento.lead_id)
                    .filter(Agendamento.concluido.isnot(True)).distinct().all()}
    resp_ids = {l.atribuido_para for l in leads}
    nomes = ({u.id: u.nome for u in db.query(Usuario).filter(Usuario.id.in_(resp_ids)).all()}
             if resp_ids else {})

    agora = datetime.utcnow()
    por_op = {}
    for l in leads:
        if l.id in com_pendente:
            continue   # já tem próximo passo agendado
        op = por_op.setdefault(l.atribuido_para,
                               {"id": l.atribuido_para,
                                "nome": nomes.get(l.atribuido_para, "—"),
                                "qtd": 0, "leads": []})
        dias = (agora - l.atualizado_em).days if l.atualizado_em else 0
        op["qtd"] += 1
        op["leads"].append({
            "id": l.id,
            "nome": l.nome if (l.nome and l.nome != "—") else l.telefone,
            "status": l.status,
            "dias": dias,
        })
    lista = list(por_op.values())
    for op in lista:
        op["leads"].sort(key=lambda x: x["dias"], reverse=True)
    lista.sort(key=lambda x: x["qtd"], reverse=True)
    return {"operadoras": lista, "total": sum(op["qtd"] for op in lista)}


@app.get("/api/relatorio/leads-perdidos-ia")
async def leads_perdidos_ia(periodo: str = "30dias", db: Session = Depends(get_db),
                            admin: Usuario = Depends(requer_admin)):
    """IA faz um apanhado geral dos MOTIVOS de perda dos leads perdidos (sob demanda)."""
    desde = _inicio_periodo(periodo)
    q = db.query(Lead).filter(Lead.status == StatusLeadEnum.perdido.value,
                              Lead.ignorar_relatorios.isnot(True))
    if desde is not None:
        q = q.filter(Lead.atualizado_em >= desde)
    perdidos = q.order_by(Lead.atualizado_em.desc()).limit(35).all()
    if not perdidos:
        return {"analise": "Nenhum lead perdido no período.", "total": 0}

    ids = [l.id for l in perdidos]
    etapa_antes = {}
    for ev in (db.query(HistoricoLead)
               .filter(HistoricoLead.lead_id.in_(ids),
                       HistoricoLead.para_status == StatusLeadEnum.perdido.value)
               .order_by(HistoricoLead.quando).all()):
        etapa_antes[ev.lead_id] = ev.de_status   # o último sobrescreve (etapa imediatamente antes de perder)
    _LAB = {"em_atendimento": "em atendimento (bot)", "qualificado": "qualificado", "assumido": "assumido",
            "pre_analise": "pré-análise", "proposta_enviada": "proposta enviada",
            "proposta_aprovada": "proposta aprovada"}

    blocos = []
    for l in perdidos:
        ult = (db.query(MensagemConversa)
               .filter(MensagemConversa.telefone == l.telefone)
               .order_by(MensagemConversa.criado_em.desc()).limit(5).all())
        trecho = []
        for m in reversed(ult):
            c = (m.conteudo or "").strip()
            if not c:
                continue
            cu = c.upper()
            if cu.startswith("[IMAGE") or cu.startswith("[IMAGEM"):
                c = "[imagem]"
            elif cu.startswith("[AUDIO"):
                c = "[áudio]"
            elif cu.startswith("[DOC"):
                c = "[documento]"
            quem = "CLIENTE" if m.role == "user" else "ATENDIMENTO"
            trecho.append(f"{quem}: {c[:200]}")
        obs = "; ".join(o.get("texto", "") for o in _parse_observacoes(l.observacoes))[:300]
        etapa = _LAB.get(etapa_antes.get(l.id), etapa_antes.get(l.id) or "—")
        bloco = f"LEAD: {l.nome or l.telefone} | perdido na etapa: {etapa} | origem: {l.origem or 'whatsapp'}"
        if obs:
            bloco += f"\nObs da operadora: {obs}"
        if trecho:
            bloco += "\nFinal da conversa:\n" + "\n".join(trecho)
        blocos.append(bloco)
    contexto = ("\n\n---\n\n".join(blocos))[:14000]

    def _chamar():
        from bot import client, MODELO_IA
        prompt = (
            "Você é analista de uma operação de F&I (financiamento de veículos). Abaixo estão LEADS PERDIDOS "
            "(cada um com a etapa em que foi perdido, observações da operadora e o final da conversa). Faça um "
            "APANHADO GERAL dos MOTIVOS DE PERDA: agrupe por motivo (ex.: parcela/preço alto, crédito negado ou "
            "restrição, sumiu/parou de responder, comprou em outro lugar, desistiu da compra, faltou documentação, "
            "etc.), dê a proporção aproximada de cada motivo, aponte padrões (ex.: em que etapa mais se perde) e "
            "termine com 2 a 3 RECOMENDAÇÕES práticas pra perder menos. Baseie-se SÓ nos dados; quando o motivo não "
            "estiver claro, classifique como 'sem motivo claro'. Português, organizado, sem asteriscos.\n\n"
            "=== LEADS PERDIDOS ===\n" + contexto)
        resp = client.messages.create(model=MODELO_IA, max_tokens=700,
                                      messages=[{"role": "user", "content": prompt}])
        return resp.content[0].text.strip()

    try:
        analise = await asyncio.to_thread(_chamar)
    except Exception as e:
        raise HTTPException(503, f"IA indisponível agora ({type(e).__name__}). Tente de novo.")
    return {"analise": analise, "total": len(perdidos), "periodo": periodo}


@app.get("/api/leads/{lead_id}/historico")
async def lead_historico(lead_id: int, db: Session = Depends(get_db),
                         usuario: Usuario = Depends(obter_usuario_atual)):
    """Linha do tempo (jornada) de um lead: cada mudança de status, quem moveu e quando."""
    eventos = (db.query(HistoricoLead)
               .filter(HistoricoLead.lead_id == lead_id)
               .order_by(HistoricoLead.quando.asc(), HistoricoLead.id.asc()).all())
    uids = {e.usuario_id for e in eventos if e.usuario_id}
    nomes = dict(db.query(Usuario.id, Usuario.nome).filter(Usuario.id.in_(uids)).all()) if uids else {}
    out = [{
        "de": e.de_status,
        "para": e.para_status,
        "quem": (nomes.get(e.usuario_id) if e.usuario_id else "Sistema/Bot"),
        "quando": _fmt_br(e.quando, "%d/%m/%Y %H:%M"),
    } for e in eventos]
    return {"eventos": out}


def _backfill_historico_sync():
    """Recria a parte recuperável da linha do tempo p/ leads que ainda não têm histórico,
    a partir das datas já existentes (criado_em, assumido_em, fechado_em) e do status atual.
    Idempotente: só mexe em leads SEM histórico."""
    db = next(get_db())
    try:
        ja_tem = {lid for (lid,) in db.query(HistoricoLead.lead_id).distinct().all()}
        leads = db.query(Lead).all()
        criados = 0
        for l in leads:
            if l.id in ja_tem:
                continue
            atual = l.status.value if hasattr(l.status, "value") else l.status
            eventos = [(None, StatusLeadEnum.em_atendimento.value,
                        l.criado_em or datetime.utcnow(), None)]   # chegada
            if l.assumido_em:
                eventos.append((None, StatusLeadEnum.assumido.value, l.assumido_em, l.atribuido_para))
            if l.fechado_em:
                eventos.append((None, StatusLeadEnum.fechado.value, l.fechado_em, l.atribuido_para))
            cobertos = {e[1] for e in eventos}
            if atual not in cobertos:   # status atual (perdido, proposta, etc.) não coberto acima
                eventos.append((None, atual,
                                l.atualizado_em or l.criado_em or datetime.utcnow(),
                                l.atribuido_para))
            for de, para, quando, uid in eventos:
                db.add(HistoricoLead(lead_id=l.id, de_status=de, para_status=para,
                                     usuario_id=uid, quando=quando))
                criados += 1
        db.commit()
        return {"leads": len(leads), "eventos_criados": criados}
    finally:
        db.close()


@app.get("/api/admin/backfill-historico")
async def backfill_historico(admin: Usuario = Depends(requer_admin)):
    """Roda UMA vez p/ recriar a linha do tempo recuperável dos leads antigos. Idempotente."""
    res = await asyncio.to_thread(_backfill_historico_sync)
    return JSONResponse({"status": "ok", **res})


@app.get("/api/relatorio/parceiros")
async def relatorio_parceiros(
    db: Session = Depends(get_db),
    admin: Usuario = Depends(requer_admin),
):
    """Relatório de desempenho por parceiro (admin only)."""
    parceiros = db.query(Parceiro).order_by(Parceiro.nome).all()
    resultado = []
    for p in parceiros:
        leads = db.query(Lead).filter(Lead.parceiro_id == p.id).all()
        total        = len(leads)
        assumidos    = sum(1 for l in leads if l.status in (
            StatusLeadEnum.assumido, StatusLeadEnum.pre_analise, StatusLeadEnum.proposta_enviada,
            StatusLeadEnum.fechado, StatusLeadEnum.perdido))
        propostas    = sum(1 for l in leads if l.status in (
            StatusLeadEnum.proposta_enviada, StatusLeadEnum.fechado))
        fechados     = sum(1 for l in leads if l.status == StatusLeadEnum.fechado)
        # Contratos gerados para leads deste parceiro
        lead_ids     = [l.id for l in leads]
        contratos    = db.query(Contrato).filter(Contrato.lead_id.in_(lead_ids)).count() if lead_ids else 0
        resultado.append({
            "id":            p.id,
            "nome":          p.nome,
            "telefone":      p.telefone,
            "ativo":         p.ativo,
            "total_leads":   total,
            "assumidos":     assumidos,
            "propostas":     propostas,
            "fechados":      fechados,
            "contratos":     contratos,
            "taxa_fechamento": round(fechados / propostas * 100, 1) if propostas > 0 else 0,
        })
    # Ordena por mais leads enviados
    resultado.sort(key=lambda x: x["total_leads"], reverse=True)
    return resultado


def _sessoes_funcionarias(db: Session, limit: int = 1000) -> list:
    """Retorna sessões apenas de funcionárias (sem admin), ordenadas por login desc."""
    sessoes = (
        db.query(SessaoUsuario)
        .join(Usuario, Usuario.id == SessaoUsuario.usuario_id)
        .filter(Usuario.role == RoleEnum.funcionario)
        .order_by(SessaoUsuario.login_em.desc())
        .limit(limit)
        .all()
    )
    resultado = []
    for s in sessoes:
        fim = s.logout_em or s.ultimo_ativo_em
        tempo_s = max(0, int((fim - s.login_em).total_seconds())) if fim and s.login_em else 0
        # Garante que tempo ativo nunca ultrapasse o tempo logado (corrige dados inflados)
        tempo_ativo_corrigido = min(s.tempo_ativo_s or 0, tempo_s)
        resultado.append({
            "id": s.id,
            "usuario": s.usuario.nome if s.usuario else "—",
            "role": s.usuario.role if s.usuario else "—",
            "ip": s.ip or "—",
            "localizacao": s.localizacao or "—",
            "login_em": _fmt_br(s.login_em) or "—",
            "ultimo_ativo_em": _fmt_br(s.ultimo_ativo_em) or "—",
            "logout_em": _fmt_br(s.logout_em),
            "tempo_logado": _duracao_str(tempo_s),
            "tempo_ativo": _duracao_str(tempo_ativo_corrigido),
            "tempo_ativo_s": tempo_ativo_corrigido,
            "ativa": (
                s.logout_em is None
                and s.ultimo_ativo_em is not None
                and (datetime.utcnow() - s.ultimo_ativo_em).total_seconds() < 600  # 10 min = tolera 1 heartbeat perdido
            ),
        })
    return resultado


@app.get("/api/relatorio/sessoes")
async def relatorio_sessoes(
    db: Session = Depends(get_db),
    admin: Usuario = Depends(requer_admin),
):
    """Retorna histórico de sessões somente de funcionárias (admin only)."""
    return _sessoes_funcionarias(db)


# Horário comercial (BR): seg-sex 9–18h, sáb 9–13h, domingo fechado
_COMERCIAL = {0: (9, 18), 1: (9, 18), 2: (9, 18), 3: (9, 18), 4: (9, 18), 5: (9, 13)}


def _segundos_comerciais(ini_utc, fim_utc) -> int:
    """Segundos decorridos entre dois instantes, contando SÓ o horário comercial (BR)."""
    if not ini_utc or not fim_utc or fim_utc <= ini_utc:
        return 0
    ini = ini_utc.replace(tzinfo=timezone.utc).astimezone(_TZ_BR)
    fim = fim_utc.replace(tzinfo=timezone.utc).astimezone(_TZ_BR)
    total = 0.0
    dia = ini.date()
    while dia <= fim.date():
        janela = _COMERCIAL.get(dia.weekday())
        if janela:
            h0, h1 = janela
            d0 = datetime(dia.year, dia.month, dia.day, h0, 0, tzinfo=_TZ_BR)
            d1 = datetime(dia.year, dia.month, dia.day, h1, 0, tzinfo=_TZ_BR)
            s = max(ini, d0)
            e = min(fim, d1)
            if e > s:
                total += (e - s).total_seconds()
        dia += timedelta(days=1)
    return int(total)


def _tempos_resposta_por_func(db, desde, ate=None) -> dict:
    """Tempo de 1ª resposta humana por funcionária no período (horário comercial).
    Retorna {nome: {qtd, media, mediana}}. Atribui pela 1ª resposta '[Nome]:'."""
    from collections import defaultdict
    q = db.query(MensagemConversa).filter(MensagemConversa.criado_em >= desde)
    if ate:
        q = q.filter(MensagemConversa.criado_em < ate)
    msgs = q.order_by(MensagemConversa.telefone, MensagemConversa.criado_em).all()
    ignorados = {t for (t,) in db.query(Lead.telefone).filter(Lead.ignorar_relatorios == True).all()}
    grupos = defaultdict(list)
    for m in msgs:
        if m.telefone in ignorados:
            continue
        grupos[m.telefone].append(m)
    por_func = defaultdict(list)
    for tel, lst in grupos.items():
        idx_h = next((i for i, m in enumerate(lst)
                      if m.role == "assistant" and (m.conteudo or "").startswith("[")), None)
        if idx_h is None:
            continue
        h = lst[idx_h]
        cli = next((lst[j] for j in range(idx_h - 1, -1, -1) if lst[j].role == "user"), None)
        if cli is None:
            continue
        seg = _segundos_comerciais(cli.criado_em, h.criado_em)
        mobj = re.match(r"^\[([^\]]+)\]:", h.conteudo or "")
        por_func[mobj.group(1) if mobj else "—"].append(seg)

    def _ag(segs):
        if not segs:
            return {"qtd": 0, "media": "—", "mediana": "—"}
        o = sorted(segs); n = len(o)
        media = sum(o) // n
        mediana = o[n // 2] if n % 2 else (o[n // 2 - 1] + o[n // 2]) // 2
        return {"qtd": n, "media": _duracao_str(media), "mediana": _duracao_str(mediana)}
    return {nome: _ag(segs) for nome, segs in por_func.items()}


@app.get("/api/relatorio/conversas-paradas")
async def relatorio_conversas_paradas(db: Session = Depends(get_db),
                                      admin: Usuario = Depends(requer_admin)):
    """Conversas em que a ÚLTIMA mensagem é do cliente (esperando resposta agora)."""
    from sqlalchemy import func
    from collections import defaultdict
    ativos = [StatusLeadEnum.qualificado.value, StatusLeadEnum.assumido.value,
              StatusLeadEnum.proposta_enviada.value, StatusLeadEnum.proposta_aprovada.value]
    leads = db.query(Lead).filter(Lead.status.in_(ativos),
                                  Lead.ignorar_relatorios.isnot(True)).all()
    if not leads:
        return {"total": 0, "por_responsavel": [], "leads": []}

    # Última mensagem de cada telefone (1 só consulta)
    sub = (db.query(MensagemConversa.telefone, func.max(MensagemConversa.id).label("mid"))
           .group_by(MensagemConversa.telefone).subquery())
    ultimas = db.query(MensagemConversa).join(sub, MensagemConversa.id == sub.c.mid).all()
    last_by_tel = {m.telefone: m for m in ultimas}

    # Nomes das responsáveis
    resp_ids = {l.atribuido_para for l in leads if l.atribuido_para}
    nomes = {u.id: u.nome for u in db.query(Usuario).filter(Usuario.id.in_(resp_ids)).all()} if resp_ids else {}

    agora = datetime.utcnow()
    paradas = []
    for l in leads:
        m = last_by_tel.get(l.telefone)
        if not m or m.role != "user":
            continue  # última mensagem não é do cliente → não está esperando
        espera_s = max(0, int((agora - m.criado_em).total_seconds()))
        resp = nomes.get(l.atribuido_para) if l.atribuido_para else None
        paradas.append({
            "lead_id": l.id,
            "nome": l.nome if (l.nome and l.nome != "—") else l.telefone,
            "telefone": l.telefone,
            "responsavel": resp or "Aguardando atendente",
            "status": l.status,
            "espera_s": espera_s,
            "espera": _duracao_str(espera_s),
            "desde": _fmt_br(m.criado_em, "%d/%m %H:%M"),
        })
    paradas.sort(key=lambda x: x["espera_s"], reverse=True)

    # Resumo por responsável
    por_resp = defaultdict(lambda: {"qtd": 0, "max_s": 0})
    for p in paradas:
        r = por_resp[p["responsavel"]]
        r["qtd"] += 1
        r["max_s"] = max(r["max_s"], p["espera_s"])
    resumo = [{"nome": k, "qtd": v["qtd"], "max_espera": _duracao_str(v["max_s"])}
              for k, v in sorted(por_resp.items(), key=lambda x: -x[1]["max_s"])]

    return {"total": len(paradas), "por_responsavel": resumo, "leads": paradas}


@app.get("/api/templates-mensagem")
async def templates_mensagem(db: Session = Depends(get_db),
                             usuario: Usuario = Depends(obter_usuario_atual)):
    """Modelos de mensagem prontos para enviar ao cliente (proposta, pré-análise)."""
    chaves = [("template_proposta", "📋 Proposta completa"),
              ("template_pre_analise", "🔎 Pré-análise")]
    out = []
    for chave, titulo in chaves:
        c = db.query(Configuracao).filter(Configuracao.chave == chave).first()
        if c and c.valor and c.valor.strip():
            out.append({"chave": chave, "titulo": titulo, "texto": c.valor})
    return out


@app.get("/api/conversas-paradas/minhas")
async def conversas_paradas_minhas(db: Session = Depends(get_db),
                                   usuario: Usuario = Depends(obter_usuario_atual)):
    """Conversas paradas (esperando resposta) acima do limite — para o alerta no topo.
    Funcionária vê as dela; admin vê todas."""
    from sqlalchemy import func
    cfg = db.query(Configuracao).filter(Configuracao.chave == "parada_alerta_min").first()
    limite_min = int(cfg.valor) if (cfg and cfg.valor and cfg.valor.isdigit()) else 30
    ativos = [StatusLeadEnum.qualificado.value, StatusLeadEnum.assumido.value,
              StatusLeadEnum.proposta_enviada.value, StatusLeadEnum.proposta_aprovada.value]
    q = db.query(Lead).filter(Lead.status.in_(ativos), Lead.ignorar_relatorios.isnot(True))
    if usuario.role != RoleEnum.admin:
        q = q.filter(Lead.atribuido_para == usuario.id)
    leads = q.all()
    if not leads:
        return {"limite_min": limite_min, "itens": []}
    tels = [l.telefone for l in leads]
    sub = (db.query(MensagemConversa.telefone, func.max(MensagemConversa.id).label("mid"))
           .filter(MensagemConversa.telefone.in_(tels))
           .group_by(MensagemConversa.telefone).subquery())
    ultimas = db.query(MensagemConversa).join(sub, MensagemConversa.id == sub.c.mid).all()
    last_by_tel = {m.telefone: m for m in ultimas}
    resp_ids = {l.atribuido_para for l in leads if l.atribuido_para}
    nomes = {u.id: u.nome for u in db.query(Usuario).filter(Usuario.id.in_(resp_ids)).all()} if resp_ids else {}

    agora = datetime.utcnow()
    itens = []
    for l in leads:
        m = last_by_tel.get(l.telefone)
        if not m or m.role != "user":
            continue
        espera_s = max(0, int((agora - m.criado_em).total_seconds()))
        if espera_s < limite_min * 60:
            continue
        itens.append({
            "lead_id": l.id,
            "nome": l.nome if (l.nome and l.nome != "—") else l.telefone,
            "responsavel": (nomes.get(l.atribuido_para) if l.atribuido_para else "Aguardando atendente"),
            "espera_s": espera_s,
            "espera": _duracao_str(espera_s),
        })
    itens.sort(key=lambda x: x["espera_s"], reverse=True)
    return {"limite_min": limite_min, "itens": itens}


@app.get("/api/relatorio/tempo-resposta")
async def relatorio_tempo_resposta(
    dias: int = 30,
    db: Session = Depends(get_db),
    admin: Usuario = Depends(requer_admin),
):
    """Tempo até a 1ª resposta humana (após handoff da IA), contando só horário comercial.
    Mensagens da equipe são salvas com prefixo '[Nome]:' — é assim que distinguimos da IA."""
    from collections import defaultdict
    desde = datetime.utcnow() - timedelta(days=max(1, min(dias, 180)))
    msgs = (db.query(MensagemConversa)
            .filter(MensagemConversa.criado_em >= desde)
            .order_by(MensagemConversa.telefone, MensagemConversa.criado_em).all())
    ignorados = {t for (t,) in db.query(Lead.telefone).filter(Lead.ignorar_relatorios == True).all()}
    # Nomes de admins (dona/owner) — não entram no relatório de tempo de resposta
    admin_nomes = {u.nome for u in db.query(Usuario).filter(Usuario.role == RoleEnum.admin).all()}
    grupos = defaultdict(list)
    for m in msgs:
        if m.telefone in ignorados:
            continue
        grupos[m.telefone].append(m)

    tempos = []          # (segundos, nome_funcionaria)
    for tel, lst in grupos.items():
        idx_h = next((i for i, m in enumerate(lst)
                      if m.role == "assistant" and (m.conteudo or "").startswith("[")), None)
        if idx_h is None:
            continue
        h = lst[idx_h]
        cli = next((lst[j] for j in range(idx_h - 1, -1, -1) if lst[j].role == "user"), None)
        if cli is None:
            continue
        seg = _segundos_comerciais(cli.criado_em, h.criado_em)
        mobj = re.match(r"^\[([^\]]+)\]:", h.conteudo or "")
        nome = mobj.group(1) if mobj else "—"
        if nome in admin_nomes:
            continue   # admin/dona não conta (evita distorcer a média da equipe)
        tempos.append((seg, nome))

    # Leads qualificados que ainda não tiveram nenhuma 1ª resposta humana
    aguardando = 0
    quali = db.query(Lead).filter(Lead.status == StatusLeadEnum.qualificado,
                                  Lead.ignorar_relatorios.isnot(True)).all()
    for l in quali:
        tem_humano = (db.query(MensagemConversa)
                      .filter(MensagemConversa.telefone == l.telefone,
                              MensagemConversa.role == "assistant",
                              MensagemConversa.conteudo.like("[%"))
                      .first())
        if not tem_humano:
            aguardando += 1

    def _agg(lista_seg):
        if not lista_seg:
            return {"qtd": 0, "media_s": 0, "mediana_s": 0, "media": "—", "mediana": "—"}
        ordenada = sorted(lista_seg)
        n = len(ordenada)
        media = sum(ordenada) // n
        mediana = ordenada[n // 2] if n % 2 else (ordenada[n // 2 - 1] + ordenada[n // 2]) // 2
        return {"qtd": n, "media_s": media, "mediana_s": mediana,
                "media": _duracao_str(media), "mediana": _duracao_str(mediana)}

    geral = _agg([s for s, _ in tempos])

    por_func = defaultdict(list)
    for s, nome in tempos:
        por_func[nome].append(s)
    funcionarias = []
    for nome, segs in por_func.items():
        a = _agg(segs)
        funcionarias.append({"nome": nome, "qtd": a["qtd"], "media": a["media"],
                             "media_s": a["media_s"], "mediana": a["mediana"]})
    funcionarias.sort(key=lambda f: f["media_s"])

    return {"dias": dias, "geral": geral, "funcionarias": funcionarias, "aguardando": aguardando}


@app.get("/api/relatorio/volume-api")
async def relatorio_volume_api(db: Session = Depends(get_db), admin: Usuario = Depends(requer_admin)):
    """Mede o volume dos últimos 30 dias e estima o custo na API Oficial do WhatsApp.
    Regra: cliente chama + resposta em até 24h = grátis; mensagem fora das 24h = cobrada (template)."""
    from collections import defaultdict
    desde = datetime.utcnow() - timedelta(days=30)
    msgs = (db.query(MensagemConversa)
            .filter(MensagemConversa.criado_em >= desde)
            .order_by(MensagemConversa.telefone, MensagemConversa.criado_em).all())
    ignorados = {t for (t,) in db.query(Lead.telefone).filter(Lead.ignorar_relatorios == True).all()}
    grupos = defaultdict(list)
    for m in msgs:
        if m.telefone in ignorados:
            continue
        grupos[m.telefone].append(m)

    conversas_recebidas = recebidas = enviadas = proativas = 0
    for tel, lst in grupos.items():
        if any(m.role == "user" for m in lst):
            conversas_recebidas += 1
        last_user = None
        for m in lst:
            if m.role == "user":
                recebidas += 1
                last_user = m.criado_em
            else:
                enviadas += 1
                # fora da janela de 24h → seria mensagem-modelo (cobrada)
                if last_user is None or (m.criado_em - last_user) > timedelta(hours=24):
                    proativas += 1

    # Estimativa (faixas): plataforma BSP fixa + mensagens proativas por unidade
    BSP_MIN, BSP_MAX = 200, 400
    MSG_MIN, MSG_MAX = 0.10, 0.50
    custo_min = round(BSP_MIN + proativas * MSG_MIN)
    custo_max = round(BSP_MAX + proativas * MSG_MAX)
    return {
        "dias": 30,
        "conversas_recebidas": conversas_recebidas,
        "msgs_recebidas": recebidas,
        "msgs_enviadas": enviadas,
        "msgs_proativas": proativas,
        "custo_min": custo_min,
        "custo_max": custo_max,
    }


def _origem_label_curto(o):
    return {None: "Bot", "": "Bot", "whatsapp": "Bot",
            "parceiro": "Parceiro", "rede_social": "Rede Social",
            "indicacao": "Indicação", "ex_cliente": "Ex-cliente"}.get(o, o or "Bot")


def _resultado_label(status):
    if status == StatusLeadEnum.fechado.value:
        return "Fechado"
    if status == StatusLeadEnum.perdido.value:
        return "Perdido"
    return "Em andamento"


@app.get("/api/relatorio/leads-calendario")
async def relatorio_leads_calendario(
    ano: int = None, mes: int = None, funcionaria: int = None, escopo: str = "mes",
    db: Session = Depends(get_db), admin: Usuario = Depends(requer_admin),
):
    """Resumo de leads por funcionária. escopo='mes' (atendidos no mês, com calendário)
    ou 'geral' (todos os tempos — útil porque a venda demora a fechar)."""
    geral = (escopo == "geral")
    hoje = _agora_br().date()
    ano = ano or hoje.year
    mes = mes or hoje.month
    inicio = datetime(ano, mes, 1) + timedelta(hours=3)            # meia-noite BR em UTC naive
    if mes == 12:
        fim = datetime(ano + 1, 1, 1) + timedelta(hours=3)
    else:
        fim = datetime(ano, mes + 1, 1) + timedelta(hours=3)

    # Só funcionárias entram neste relatório (exclui leads atribuídos a admins)
    _ids_func = [u.id for u in db.query(Usuario).filter(Usuario.role == RoleEnum.funcionario).all()]
    q = db.query(Lead).filter(Lead.atribuido_para.isnot(None), Lead.ignorar_relatorios.isnot(True),
                              Lead.atribuido_para.in_(_ids_func))
    if not geral:
        q = q.filter(Lead.assumido_em >= inicio, Lead.assumido_em < fim)
    if funcionaria:
        q = q.filter(Lead.atribuido_para == funcionaria)
    leads = q.all()

    # Mapa de nomes das funcionárias
    func_ids = {l.atribuido_para for l in leads}
    nomes = {u.id: u.nome for u in db.query(Usuario).filter(Usuario.id.in_(func_ids)).all()} if func_ids else {}

    from collections import defaultdict
    # Status que indicam que o lead "virou proposta" (chegou a proposta ou além)
    _virou_proposta = {StatusLeadEnum.proposta_enviada.value,
                       StatusLeadEnum.proposta_aprovada.value,
                       StatusLeadEnum.fechado.value}
    dias = defaultdict(lambda: {"total": 0})            # dia -> total
    resumo = defaultdict(lambda: {"total": 0, "propostas": 0,
                                  "origens": defaultdict(int), "resultados": defaultdict(int)})

    for l in leads:
        if not geral and l.assumido_em:
            dias[int(_fmt_br(l.assumido_em, "%d"))]["total"] += 1
        nome = nomes.get(l.atribuido_para, "—")
        r = resumo[nome]
        r["total"] += 1
        if l.status in _virou_proposta:
            r["propostas"] += 1
        r["origens"][_origem_label_curto(l.origem)] += 1
        r["resultados"][_resultado_label(l.status)] += 1

    # Tempo de 1ª resposta (média/mediana) por funcionária — do mês ou geral
    if geral:
        tempos = _tempos_resposta_por_func(db, datetime(2020, 1, 1), None)
    else:
        tempos = _tempos_resposta_por_func(db, inicio, fim)

    resumo_lst = []
    for nome, d in sorted(resumo.items(), key=lambda x: -x[1]["total"]):
        t = tempos.get(nome, {"media": "—", "mediana": "—", "qtd": 0})
        resumo_lst.append({
            "funcionaria": nome,
            "total": d["total"],
            "propostas": d["propostas"],
            "origens": dict(d["origens"]),
            "resultados": dict(d["resultados"]),
            "tempo_media": t["media"],
            "tempo_mediana": t["mediana"],
            "tempo_qtd": t["qtd"],
        })

    # Lista de funcionárias para o filtro
    todas = db.query(Usuario).filter(Usuario.role == RoleEnum.funcionario).order_by(Usuario.nome).all()
    return {
        "ano": ano, "mes": mes,
        "escopo": "geral" if geral else "mes",
        "dias": {str(k): v for k, v in dias.items()},
        "resumo": resumo_lst,
        "funcionarias": [{"id": u.id, "nome": u.nome} for u in todas],
        "total_mes": len(leads),
    }


@app.get("/api/relatorio/leads-calendario/xlsx")
async def relatorio_leads_calendario_xlsx(
    ano: int = None, mes: int = None, funcionaria: int = None, escopo: str = "mes",
    db: Session = Depends(get_db), admin: Usuario = Depends(requer_admin),
):
    """Exporta o relatório de leads por funcionária em Excel (do mês ou geral)."""
    import io
    from collections import defaultdict
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    geral = (escopo == "geral")
    hoje = _agora_br().date()
    ano = ano or hoje.year
    mes = mes or hoje.month
    inicio = datetime(ano, mes, 1) + timedelta(hours=3)
    fim = (datetime(ano + 1, 1, 1) if mes == 12 else datetime(ano, mes + 1, 1)) + timedelta(hours=3)
    # Só funcionárias entram neste relatório (exclui leads atribuídos a admins)
    _ids_func = [u.id for u in db.query(Usuario).filter(Usuario.role == RoleEnum.funcionario).all()]
    q = db.query(Lead).filter(Lead.atribuido_para.isnot(None), Lead.ignorar_relatorios.isnot(True),
                              Lead.atribuido_para.in_(_ids_func))
    if not geral:
        q = q.filter(Lead.assumido_em >= inicio, Lead.assumido_em < fim)
    if funcionaria:
        q = q.filter(Lead.atribuido_para == funcionaria)
    leads = q.all()
    func_ids = {l.atribuido_para for l in leads}
    nomes = {u.id: u.nome for u in db.query(Usuario).filter(Usuario.id.in_(func_ids)).all()} if func_ids else {}
    _virou = {StatusLeadEnum.proposta_enviada.value, StatusLeadEnum.proposta_aprovada.value, StatusLeadEnum.fechado.value}

    resumo = defaultdict(lambda: {"total": 0, "propostas": 0, "fechado": 0, "perdido": 0,
                                  "andamento": 0, "origens": defaultdict(int)})
    origens_set = set()
    for l in leads:
        nome = nomes.get(l.atribuido_para, "—")
        r = resumo[nome]
        r["total"] += 1
        if l.status in _virou:
            r["propostas"] += 1
        if l.status == StatusLeadEnum.fechado.value:
            r["fechado"] += 1
        elif l.status == StatusLeadEnum.perdido.value:
            r["perdido"] += 1
        else:
            r["andamento"] += 1
        ol = _origem_label_curto(l.origem)
        r["origens"][ol] += 1
        origens_set.add(ol)
    origens_cols = sorted(origens_set)

    cor_cab = PatternFill("solid", fgColor="0D2B4E")
    fonte_cab = Font(bold=True, color="FFFFFF", size=10)
    fonte_norm = Font(size=9)
    centro = Alignment(horizontal="center", vertical="center")
    borda = Border(bottom=Side(style="thin", color="CCCCCC"), top=Side(style="thin", color="CCCCCC"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads por funcionária"
    _MESES = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho','Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
    cols = (["Funcionária", "Total atendidos", "Viraram proposta", "% Proposta",
             "Fechado", "% Conversão", "Perdido", "Em andamento", "% Prop→Contrato"]
            + [f"Origem: {o}" for o in origens_cols])
    for i, larg in enumerate([22, 14, 15, 11, 10, 12, 10, 14, 15] + [12] * len(origens_cols), 1):
        ws.column_dimensions[get_column_letter(i)].width = larg

    n = len(cols)
    titulo_periodo = "Geral (todos os tempos)" if geral else f"{_MESES[mes-1]} {ano}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    t = ws.cell(row=1, column=1, value=f"Leads por funcionária — {titulo_periodo}")
    t.font = Font(bold=True, color="FFFFFF", size=12); t.fill = cor_cab; t.alignment = centro
    ws.row_dimensions[1].height = 22

    for col, val in enumerate(cols, 1):
        c = ws.cell(row=2, column=col, value=val)
        c.font = fonte_cab; c.fill = cor_cab; c.alignment = centro; c.border = borda

    linha = 3
    def _pct(a, b):
        return f"{round(a/b*100)}%" if b else "0%"
    for nome, d in sorted(resumo.items(), key=lambda x: -x[1]["total"]):
        vals = [nome, d["total"], d["propostas"], _pct(d["propostas"], d["total"]),
                d["fechado"], _pct(d["fechado"], d["total"]), d["perdido"], d["andamento"],
                _pct(d["fechado"], d["propostas"])]
        vals += [d["origens"].get(o, 0) for o in origens_cols]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=linha, column=col, value=val)
            c.alignment = centro; c.border = borda; c.font = fonte_norm
        linha += 1
    if linha == 3:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n)
        ws.cell(row=3, column=1, value="Nenhum lead no período.").alignment = centro

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "leads_geral.xlsx" if geral else f"leads_{ano}_{mes:02d}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})


@app.get("/api/relatorio/ip-compartilhado")
async def relatorio_ip_compartilhado(
    dias: int = 30,
    db: Session = Depends(get_db),
    admin: Usuario = Depends(requer_admin),
):
    """Detecta IPs usados por mais de uma funcionária (possível acesso por terceiro).
    Sinaliza em vermelho quando houve acesso CONCOMITANTE (sobreposição no tempo)."""
    from collections import defaultdict
    desde = datetime.utcnow() - timedelta(days=max(1, min(dias, 180)))
    sessoes = (
        db.query(SessaoUsuario)
        .join(Usuario, Usuario.id == SessaoUsuario.usuario_id)
        .filter(Usuario.role == RoleEnum.funcionario)
        .filter(SessaoUsuario.login_em >= desde)
        .order_by(SessaoUsuario.login_em)
        .all()
    )
    por_ip = defaultdict(list)
    for s in sessoes:
        if s.ip and s.ip != "—":
            por_ip[s.ip].append(s)

    agora = datetime.utcnow()
    resultado = []
    for ip, lst in por_ip.items():
        # Resumo por usuária
        usuarios = {}
        for s in lst:
            nome = s.usuario.nome if s.usuario else "—"
            u = usuarios.setdefault(s.usuario_id, {
                "nome": nome, "sessoes": 0, "ultima": None, "localizacao": None,
            })
            u["sessoes"] += 1
            fim = s.ultimo_ativo_em or s.login_em
            if fim and (u["ultima"] is None or fim > u["ultima"]):
                u["ultima"] = fim
            if s.localizacao and s.localizacao != "—":
                u["localizacao"] = s.localizacao
        if len(usuarios) < 2:
            continue  # IP usado por uma só pessoa — normal

        # Detecta sobreposições no tempo entre usuárias diferentes
        intervalos = [
            (s.usuario_id, (s.usuario.nome if s.usuario else "—"),
             s.login_em, (s.logout_em or s.ultimo_ativo_em or s.login_em))
            for s in lst
        ]
        overlaps = []
        for i in range(len(intervalos)):
            for j in range(i + 1, len(intervalos)):
                a, b = intervalos[i], intervalos[j]
                if a[0] == b[0]:
                    continue
                if a[2] and b[2] and a[2] <= b[3] and b[2] <= a[3]:
                    ini = max(a[2], b[2])
                    fim = min(a[3], b[3])
                    overlaps.append({
                        "u1": a[1], "u2": b[1],
                        "inicio": _fmt_br(ini, "%d/%m/%Y %H:%M"),
                        "fim": _fmt_br(fim, "%H:%M"),
                        "_ts": ini,
                    })
        overlaps.sort(key=lambda o: o["_ts"], reverse=True)
        for o in overlaps:
            o.pop("_ts", None)

        usuarios_lst = sorted(usuarios.values(), key=lambda u: u["ultima"] or datetime.min, reverse=True)
        resultado.append({
            "ip": ip,
            "qtd_usuarias": len(usuarios),
            "concomitante": bool(overlaps),
            "usuarias": [{
                "nome": u["nome"],
                "sessoes": u["sessoes"],
                "ultima": _fmt_br(u["ultima"]) if u["ultima"] else "—",
                "localizacao": u["localizacao"] or "—",
            } for u in usuarios_lst],
            "sobreposicoes": overlaps[:5],
            "_recencia": max((u["ultima"] for u in usuarios.values() if u["ultima"]), default=datetime.min),
        })

    # Ordena: concomitantes primeiro, depois mais recentes
    resultado.sort(key=lambda r: (r["concomitante"], r["_recencia"]), reverse=True)
    for r in resultado:
        r.pop("_recencia", None)
    return {"dias": dias, "itens": resultado}


@app.delete("/api/relatorio/sessoes")
async def zerar_sessoes(
    db: Session = Depends(get_db),
    admin: Usuario = Depends(requer_admin),
):
    """Apaga todo o histórico de sessões (admin only)."""
    db.query(SessaoUsuario).delete()
    db.commit()
    return {"status": "zerado"}


@app.get("/api/relatorio/sessoes/csv")
async def relatorio_sessoes_csv(
    db: Session = Depends(get_db),
    admin: Usuario = Depends(requer_admin),
):
    """Exporta atividade das funcionárias como XLSX — uma aba por funcionária, agrupado por dia."""
    from fastapi.responses import StreamingResponse
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict

    sessoes = _sessoes_funcionarias(db)

    # ── Agrupa: funcionária → data → lista de sessões ──────────────────────
    grupos: dict[str, dict[str, list]] = defaultdict(lambda: defaultdict(list))
    for s in sessoes:
        login_raw = s["login_em"] or ""
        # login_em vem como "DD/MM/YYYY HH:MM" — pega só a data
        data_str = login_raw[:10] if len(login_raw) >= 10 else "Sem data"
        grupos[s["usuario"]][data_str].append(s)

    wb = Workbook()
    wb.remove(wb.active)  # remove aba padrão

    # Estilos
    cor_cabecalho   = PatternFill("solid", fgColor="0D2B4E")   # navy
    cor_dia         = PatternFill("solid", fgColor="E8F4FD")   # azul claro
    cor_resumo      = PatternFill("solid", fgColor="FEF9C3")   # amarelo claro
    fonte_cab       = Font(bold=True, color="FFFFFF", size=10)
    fonte_dia       = Font(bold=True, color="1A5276", size=10)
    fonte_resumo    = Font(bold=True, color="713F12", size=10)
    fonte_normal    = Font(size=9)
    alin_centro     = Alignment(horizontal="center", vertical="center")
    borda_fina      = Border(
        bottom=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
    )

    def _set_row(ws, row, values, fill=None, font=None, bold=False):
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = alin_centro
            cell.border = borda_fina
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            elif bold:
                cell.font = Font(bold=True, size=9)
            else:
                cell.font = fonte_normal

    for nome_func in sorted(grupos.keys()):
        dias = grupos[nome_func]
        # Nome da aba — máx 31 chars, sem chars inválidos
        aba = nome_func[:31].replace("/","").replace("\\","").replace("?","").replace("*","").replace("[","").replace("]","")
        ws = wb.create_sheet(title=aba)

        # Larguras das colunas
        larguras = [12, 10, 10, 10, 14, 14, 26, 18]
        cols_cabecalho = ["Data", "Sessões", "T. Ativo", "T. Logado", "Primeira Entrada", "Último Acesso", "Localização", "IP"]
        for i, larg in enumerate(larguras, 1):
            ws.column_dimensions[get_column_letter(i)].width = larg

        linha = 1

        # Título da aba
        ws.merge_cells(f"A{linha}:H{linha}")
        titulo = ws.cell(row=linha, column=1, value=f"Atividade — {nome_func}")
        titulo.font = Font(bold=True, color="FFFFFF", size=12)
        titulo.fill = cor_cabecalho
        titulo.alignment = alin_centro
        ws.row_dimensions[linha].height = 22
        linha += 1

        # Cabeçalho
        _set_row(ws, linha, cols_cabecalho, fill=cor_cabecalho, font=fonte_cab)
        ws.row_dimensions[linha].height = 18
        linha += 1

        for data_str in sorted(dias.keys(), reverse=True):
            sessoes_dia = dias[data_str]

            # Calcula resumo do dia
            total_ativo_s = sum(s["tempo_ativo_s"] or 0 for s in sessoes_dia)
            total_logado_s = 0
            primeira_entrada = None
            ultimo_acesso = None
            localizacoes = set()
            ips = set()
            for s in sessoes_dia:
                if s["login_em"] and s["login_em"] != "—":
                    if not primeira_entrada or s["login_em"] < primeira_entrada:
                        primeira_entrada = s["login_em"]
                if s["ultimo_ativo_em"] and s["ultimo_ativo_em"] != "—":
                    if not ultimo_acesso or s["ultimo_ativo_em"] > ultimo_acesso:
                        ultimo_acesso = s["ultimo_ativo_em"]
                if s["localizacao"] and s["localizacao"] != "—":
                    localizacoes.add(s["localizacao"])
                if s["ip"] and s["ip"] != "—":
                    ips.add(s["ip"])

            # Linha de resumo do dia
            ativo_str = _duracao_str(total_ativo_s)
            _set_row(ws, linha, [
                data_str,
                len(sessoes_dia),
                ativo_str,
                "—",
                primeira_entrada or "—",
                ultimo_acesso or "—",
                ", ".join(sorted(localizacoes)) or "—",
                ", ".join(sorted(ips)) or "—",
            ], fill=cor_dia, font=fonte_dia)
            ws.row_dimensions[linha].height = 16
            linha += 1

            # Sessões individuais do dia (indentadas)
            for s in sorted(sessoes_dia, key=lambda x: x["login_em"] or ""):
                _set_row(ws, linha, [
                    "",                         # data (já na linha do dia)
                    1,                          # sessões
                    s["tempo_ativo"],
                    s["tempo_logado"],
                    s["login_em"] or "—",
                    s["ultimo_ativo_em"] or "—",
                    s["localizacao"] or "—",
                    s["ip"] or "—",
                ])
                ws.row_dimensions[linha].height = 15
                linha += 1

        # Linha de total geral da funcionária
        total_geral_s = sum(s["tempo_ativo_s"] or 0 for dias_list in dias.values() for s in dias_list)
        total_sessoes = sum(len(v) for v in dias.values())
        _set_row(ws, linha, [
            "TOTAL GERAL",
            total_sessoes,
            _duracao_str(total_geral_s),
            "—", "—", "—", "—", "—",
        ], fill=cor_resumo, font=fonte_resumo)
        ws.row_dimensions[linha].height = 18

    # ── Aba resumo geral (primeira) ────────────────────────────────────────
    ws_res = wb.create_sheet(title="Resumo Geral", index=0)
    ws_res.column_dimensions["A"].width = 24
    ws_res.column_dimensions["B"].width = 10
    ws_res.column_dimensions["C"].width = 12
    ws_res.column_dimensions["D"].width = 20
    ws_res.column_dimensions["E"].width = 20

    ws_res.merge_cells("A1:E1")
    t = ws_res.cell(row=1, column=1, value="Resumo de Atividade das Funcionárias")
    t.font = Font(bold=True, color="FFFFFF", size=12)
    t.fill = cor_cabecalho
    t.alignment = alin_centro
    ws_res.row_dimensions[1].height = 22

    _set_row(ws_res, 2, ["Funcionária", "Sessões", "Tempo Ativo", "Primeira Entrada", "Último Acesso"],
             fill=cor_cabecalho, font=fonte_cab)
    ws_res.row_dimensions[2].height = 18

    linha_res = 3
    for nome_func in sorted(grupos.keys()):
        dias = grupos[nome_func]
        total_s = sum(s["tempo_ativo_s"] or 0 for dl in dias.values() for s in dl)
        total_sess = sum(len(v) for v in dias.values())
        todas = [s for dl in dias.values() for s in dl]
        p_entrada = min((s["login_em"] for s in todas if s["login_em"] and s["login_em"] != "—"), default="—")
        u_acesso  = max((s["ultimo_ativo_em"] for s in todas if s["ultimo_ativo_em"] and s["ultimo_ativo_em"] != "—"), default="—")
        _set_row(ws_res, linha_res, [nome_func, total_sess, _duracao_str(total_s), p_entrada, u_acesso])
        ws_res.row_dimensions[linha_res].height = 15
        linha_res += 1

    # Salva em buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    headers = {"Content-Disposition": "attachment; filename=atividade_funcionarias.xlsx"}
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@app.get("/api/relatorio/ponto/xlsx")
async def relatorio_ponto_xlsx(
    inicio: str = None, fim: str = None,
    db: Session = Depends(get_db),
    admin: Usuario = Depends(requer_admin),
):
    """Exporta ponto × tempo ativo das funcionárias num intervalo de datas como XLSX.
    Abas: Resumo (totais do período por funcionária) + Detalhado (uma linha por dia)."""
    from fastapi.responses import StreamingResponse
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Intervalo de datas (BR) ────────────────────────────────────────────
    hoje = _agora_br().date()
    try:
        d_fim = datetime.strptime(fim, "%Y-%m-%d").date() if fim else hoje
        d_ini = datetime.strptime(inicio, "%Y-%m-%d").date() if inicio else (d_fim - timedelta(days=29))
    except ValueError:
        raise HTTPException(400, "Datas inválidas (use YYYY-MM-DD)")
    if d_ini > d_fim:
        d_ini, d_fim = d_fim, d_ini
    if (d_fim - d_ini).days > 366:
        raise HTTPException(400, "Intervalo máximo de 366 dias")

    agora_utc = datetime.utcnow()
    funcionarias = (db.query(Usuario)
                    .filter(Usuario.role == RoleEnum.funcionario)
                    .order_by(Usuario.nome).all())

    # ── Coleta linhas detalhadas + acumula totais por funcionária ───────────
    detalhe = []            # uma linha por (funcionária, dia com registro)
    totais = {}             # usuario_id -> {nome, dias, jornada_s, ativo_s, ocioso_s}
    dia = d_ini
    while dia <= d_fim:
        for u in funcionarias:
            pontos = _pontos_do_dia(db, u.id, dia)
            if not pontos:
                continue
            intervalos = _jornadas_de_pontos(pontos, agora_utc)
            jornada_s = sum(int((f - i).total_seconds()) for i, f, _ in intervalos)
            ativo_s = min(_tempo_ativo_intervalos(db, u.id, intervalos), jornada_s)
            ocioso_s = max(0, jornada_s - ativo_s)
            # Primeiro horário de cada tipo no dia
            horas = {}
            for p in pontos:
                horas.setdefault(p.tipo, _fmt_br(p.timestamp, "%H:%M"))
            detalhe.append({
                "nome": u.nome,
                "data": dia.strftime("%d/%m/%Y"),
                "entrada": horas.get("entrada", "—"),
                "saida_almoco": horas.get("saida_almoco", "—"),
                "volta_almoco": horas.get("volta_almoco", "—"),
                "saida": horas.get("saida", "—"),
                "jornada_s": jornada_s,
                "ativo_s": ativo_s,
                "ocioso_s": ocioso_s,
                "perc": round(ativo_s / jornada_s * 100) if jornada_s else 0,
            })
            t = totais.setdefault(u.id, {"nome": u.nome, "dias": 0, "jornada_s": 0, "ativo_s": 0, "ocioso_s": 0})
            t["dias"] += 1
            t["jornada_s"] += jornada_s
            t["ativo_s"] += ativo_s
            t["ocioso_s"] += ocioso_s
        dia += timedelta(days=1)

    # ── Estilos ─────────────────────────────────────────────────────────────
    cor_cab   = PatternFill("solid", fgColor="0D2B4E")
    cor_resumo= PatternFill("solid", fgColor="FEF9C3")
    fonte_cab = Font(bold=True, color="FFFFFF", size=10)
    fonte_res = Font(bold=True, color="713F12", size=10)
    fonte_norm= Font(size=9)
    centro    = Alignment(horizontal="center", vertical="center")
    borda     = Border(bottom=Side(style="thin", color="CCCCCC"), top=Side(style="thin", color="CCCCCC"))

    def _set(ws, row, vals, fill=None, font=None):
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = centro
            c.border = borda
            if fill: c.fill = fill
            c.font = font or fonte_norm

    wb = Workbook()
    wb.remove(wb.active)
    periodo_label = f"{d_ini.strftime('%d/%m/%Y')} a {d_fim.strftime('%d/%m/%Y')}"

    # ── Aba Resumo ───────────────────────────────────────────────────────────
    ws = wb.create_sheet(title="Resumo")
    for i, larg in enumerate([24, 8, 13, 13, 13, 10, 15, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = larg
    ws.merge_cells("A1:H1")
    t = ws.cell(row=1, column=1, value=f"Ponto × Tempo Ativo — {periodo_label}")
    t.font = Font(bold=True, color="FFFFFF", size=12); t.fill = cor_cab; t.alignment = centro
    ws.row_dimensions[1].height = 22
    _set(ws, 2, ["Funcionária", "Dias", "Jornada", "Tempo Ativo", "Ocioso", "% Ativo",
                 "Méd. Jornada/dia", "Méd. Ativo/dia"], fill=cor_cab, font=fonte_cab)
    linha = 3
    for u in funcionarias:
        if u.id not in totais:
            continue
        t = totais[u.id]
        perc = round(t["ativo_s"] / t["jornada_s"] * 100) if t["jornada_s"] else 0
        dias = t["dias"] or 1
        _set(ws, linha, [t["nome"], t["dias"], _duracao_str(t["jornada_s"]),
                         _duracao_str(t["ativo_s"]), _duracao_str(t["ocioso_s"]), f"{perc}%",
                         _duracao_str(t["jornada_s"] // dias), _duracao_str(t["ativo_s"] // dias)])
        linha += 1
    if linha == 3:
        ws.merge_cells(f"A3:H3")
        ws.cell(row=3, column=1, value="Nenhum ponto registrado no período.").alignment = centro

    # ── Aba Detalhado ──────────────────────────────────────────────────────────
    wsd = wb.create_sheet(title="Detalhado")
    for i, larg in enumerate([20, 12, 10, 11, 11, 10, 12, 12, 12, 10], 1):
        wsd.column_dimensions[get_column_letter(i)].width = larg
    wsd.merge_cells("A1:J1")
    t = wsd.cell(row=1, column=1, value=f"Detalhamento diário — {periodo_label}")
    t.font = Font(bold=True, color="FFFFFF", size=12); t.fill = cor_cab; t.alignment = centro
    wsd.row_dimensions[1].height = 22
    _set(wsd, 2, ["Funcionária", "Data", "Entrada", "S. Almoço", "V. Almoço", "Saída",
                  "Jornada", "T. Ativo", "Ocioso", "% Ativo"], fill=cor_cab, font=fonte_cab)
    linha = 3
    for r in detalhe:
        _set(wsd, linha, [r["nome"], r["data"], r["entrada"], r["saida_almoco"],
                          r["volta_almoco"], r["saida"], _duracao_str(r["jornada_s"]),
                          _duracao_str(r["ativo_s"]), _duracao_str(r["ocioso_s"]), f"{r['perc']}%"])
        linha += 1
    if linha == 3:
        wsd.merge_cells("A3:J3")
        wsd.cell(row=3, column=1, value="Nenhum ponto registrado no período.").alignment = centro

    # ── Aba Correções de ponto (trilha de auditoria) ───────────────────────────
    _ini_iso, _fim_iso = d_ini.strftime("%Y-%m-%d"), d_fim.strftime("%Y-%m-%d")
    _acao_lbl = {"adicionar": "Adicionar", "editar": "Editar", "remover": "Remover"}
    _st_lbl = {"pendente": "Pendente", "aplicada": "Aplicada", "aprovada": "Aprovada", "rejeitada": "Rejeitada"}
    corrs = (db.query(CorrecaoPonto)
             .filter(CorrecaoPonto.data >= _ini_iso, CorrecaoPonto.data <= _fim_iso)
             .order_by(CorrecaoPonto.data, CorrecaoPonto.criado_em).all())
    wsc = wb.create_sheet(title="Correções (auditoria)")
    for i, larg in enumerate([20, 12, 11, 14, 9, 9, 30, 12, 11, 18], 1):
        wsc.column_dimensions[get_column_letter(i)].width = larg
    wsc.merge_cells("A1:J1")
    t = wsc.cell(row=1, column=1, value=f"Correções de ponto — {periodo_label}")
    t.font = Font(bold=True, color="FFFFFF", size=12); t.fill = cor_cab; t.alignment = centro
    wsc.row_dimensions[1].height = 22
    _set(wsc, 2, ["Funcionária", "Data", "Ação", "Tipo", "De", "Para", "Motivo",
                  "Origem", "Status", "Resolvido por"], fill=cor_cab, font=fonte_cab)
    linha = 3
    for c in corrs:
        _set(wsc, linha, [
            (c.usuario.nome if c.usuario else "—"),
            _iso_para_br(c.data),
            _acao_lbl.get(c.acao, c.acao),
            _PONTO_LABELS.get(c.tipo_ponto, c.tipo_ponto or "—"),
            c.hora_anterior or "—",
            c.hora_nova or "—",
            c.motivo or "—",
            ("Funcionária" if c.origem == "funcionaria" else "Admin"),
            _st_lbl.get(c.status, c.status),
            (c.resolvedor.nome if c.resolvedor else "—"),
        ])
        linha += 1
    if linha == 3:
        wsc.merge_cells("A3:J3")
        wsc.cell(row=3, column=1, value="Nenhuma correção no período.").alignment = centro

    # ── Aba Justificativas de horário ──────────────────────────────────────────
    justs = (db.query(JustificativaPonto)
             .filter(JustificativaPonto.data >= _ini_iso, JustificativaPonto.data <= _fim_iso)
             .order_by(JustificativaPonto.data, JustificativaPonto.criado_em).all())
    wsj = wb.create_sheet(title="Justificativas")
    for i, larg in enumerate([20, 12, 42, 11, 11, 18], 1):
        wsj.column_dimensions[get_column_letter(i)].width = larg
    wsj.merge_cells("A1:F1")
    t = wsj.cell(row=1, column=1, value=f"Justificativas de horário — {periodo_label}")
    t.font = Font(bold=True, color="FFFFFF", size=12); t.fill = cor_cab; t.alignment = centro
    wsj.row_dimensions[1].height = 22
    _set(wsj, 2, ["Funcionária", "Data", "Motivo", "Atestado", "Status", "Aprovado por"],
         fill=cor_cab, font=fonte_cab)
    linha = 3
    for j in justs:
        _set(wsj, linha, [
            (j.usuario.nome if j.usuario else "—"),
            _iso_para_br(j.data),
            j.texto or "—",
            ("Sim" if j.filename else "Não"),
            _st_lbl.get(j.status, j.status),
            (j.aprovador.nome if j.aprovador else "—"),
        ])
        linha += 1
    if linha == 3:
        wsj.merge_cells("A3:F3")
        wsj.cell(row=3, column=1, value="Nenhuma justificativa no período.").alignment = centro

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"ponto_{d_ini.strftime('%Y%m%d')}_{d_fim.strftime('%Y%m%d')}.xlsx"
    headers = {"Content-Disposition": f"attachment; filename={fname}"}
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


# ─── Fila rotativa de atendimento ────────────────────────────────────────────────

def _get_fila_cfg(db: Session):
    """Retorna (ordem: list[int], posicao: int) da configuração."""
    import json
    ordem_cfg = db.query(Configuracao).filter(Configuracao.chave == "fila_ordem").first()
    pos_cfg   = db.query(Configuracao).filter(Configuracao.chave == "fila_posicao").first()
    ordem = json.loads(ordem_cfg.valor) if ordem_cfg else []
    posicao = int(pos_cfg.valor) if pos_cfg else 0
    return ordem, posicao

def _set_fila_cfg(db: Session, ordem: list, posicao: int, alterado_por: str = None):
    import json
    from datetime import datetime
    for chave, valor in [("fila_ordem", json.dumps(ordem)), ("fila_posicao", str(posicao))]:
        cfg = db.query(Configuracao).filter(Configuracao.chave == chave).first()
        if cfg:
            cfg.valor = valor
        else:
            db.add(Configuracao(chave=chave, valor=valor, descricao="Fila rotativa de atendimento"))
    if alterado_por:
        agora = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
        for chave, valor in [("fila_alterado_por", alterado_por), ("fila_alterado_em", agora)]:
            cfg = db.query(Configuracao).filter(Configuracao.chave == chave).first()
            if cfg:
                cfg.valor = valor
            else:
                db.add(Configuracao(chave=chave, valor=valor, descricao="Fila rotativa — log de alteração"))
    db.commit()

def _fila_snapshot(db: Session):
    """Retorna a fila com dados dos usuários para exibição."""
    ordem, posicao = _get_fila_cfg(db)
    if not ordem:
        return {"ordem": [], "posicao": 0, "proxima": None}
    # Filtra apenas usuários ativos
    usuarios = {u.id: u for u in db.query(Usuario).filter(
        Usuario.id.in_(ordem), Usuario.ativo == True
    ).all()}
    # Remove ids inválidos da ordem
    ordem = [uid for uid in ordem if uid in usuarios]
    if not ordem:
        return {"ordem": [], "posicao": 0, "proxima": None}
    posicao = posicao % len(ordem)

    # Contagem de leads assumidos hoje por usuário — apenas os que chegaram pela plataforma (origem NULL)
    # Usa assumido_em (quando a operadora pegou o lead) e datetimes naive UTC para compatibilidade com SQLite
    hoje_br = _agora_br().date()
    inicio_hoje = datetime(hoje_br.year, hoje_br.month, hoje_br.day) + timedelta(hours=3)  # meia-noite BR em UTC naive
    fim_hoje = inicio_hoje + timedelta(days=1)
    leads_hoje_raw = (
        db.query(Lead.atribuido_para, Lead.id)
        .filter(
            Lead.atribuido_para.in_(ordem),
            Lead.assumido_em >= inicio_hoje,
            Lead.assumido_em < fim_hoje,
            Lead.origem.is_(None),  # exclui leads inseridos manualmente
            Lead.ignorar_relatorios.isnot(True),  # exclui conversas internas
        )
        .all()
    )
    leads_hoje: dict[int, int] = {}
    for uid, _ in leads_hoje_raw:
        leads_hoje[uid] = leads_hoje.get(uid, 0) + 1

    # ── PRÓXIMA = quem tem MENOS leads hoje (empate: ordem da lista) ──────────
    # Assim a distribuição se equilibra sozinha — no fim do dia todas com a
    # mesma quantidade. Não depende mais de "avançar" um ponteiro (que travava).
    idx_proxima = min(range(len(ordem)), key=lambda i: (leads_hoje.get(ordem[i], 0), i))

    alt_por_cfg = db.query(Configuracao).filter(Configuracao.chave == "fila_alterado_por").first()
    alt_em_cfg  = db.query(Configuracao).filter(Configuracao.chave == "fila_alterado_em").first()
    return {
        "ordem": [
            {"id": uid, "nome": usuarios[uid].nome, "leads_hoje": leads_hoje.get(uid, 0)}
            for uid in ordem
        ],
        "posicao": idx_proxima,
        "proxima": {"id": ordem[idx_proxima], "nome": usuarios[ordem[idx_proxima]].nome, "leads_hoje": leads_hoje.get(ordem[idx_proxima], 0)},
        "alterado_por": alt_por_cfg.valor if alt_por_cfg else None,
        "alterado_em":  alt_em_cfg.valor  if alt_em_cfg  else None,
    }

@app.get("/api/fila")
async def get_fila(db: Session = Depends(get_db), usuario: Usuario = Depends(obter_usuario_atual)):
    return _fila_snapshot(db)

@app.put("/api/fila/ordem")
async def set_fila_ordem(request: Request, db: Session = Depends(get_db), usuario: Usuario = Depends(obter_usuario_atual)):
    """Qualquer usuário autenticado pode reordenar a fila."""
    body = await request.json()
    ordem = body.get("ordem", [])
    if not isinstance(ordem, list):
        raise HTTPException(status_code=400, detail="ordem deve ser lista de IDs")
    _, posicao = _get_fila_cfg(db)
    _set_fila_cfg(db, ordem, posicao % len(ordem) if ordem else 0, alterado_por=usuario.nome)
    return _fila_snapshot(db)

@app.post("/api/fila/avancar")
async def avancar_fila(db: Session = Depends(get_db), usuario: Usuario = Depends(obter_usuario_atual)):
    """Avança a fila para a próxima posição."""
    ordem, posicao = _get_fila_cfg(db)
    if not ordem:
        return {"status": "fila vazia"}
    nova_posicao = (posicao + 1) % len(ordem)
    _set_fila_cfg(db, ordem, nova_posicao)
    return _fila_snapshot(db)


# ─── Dashboard ────────────────────────────────────────────────────────────────────

_NO_CACHE_HEADERS = {
    "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
    "Pragma": "no-cache",
    "Expires": "0",
}


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard():
    html_path = Path(__file__).parent / "templates" / "dashboard.html"
    if html_path.exists():
        return HTMLResponse(html_path.read_text(encoding="utf-8"), headers=_NO_CACHE_HEADERS)
    return HTMLResponse("<h1>dashboard.html não encontrado</h1>")


@app.get("/")
async def root():
    return {"app": "Fácil Financiamentos Bot v2", "status": "online", "dashboard": "/dashboard"}


# ─── Contratos / Assinatura Digital ─────────────────────────────────────────────

@app.post("/api/leads/{lead_id}/contrato")
async def gerar_contrato_endpoint(
    lead_id: int,
    request: Request,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    import traceback as _tb
    try:
        # ── 1. importar módulo ────────────────────────────────────────────
        from gerar_contrato import gerar_pdf_contrato, salvar_pdf

        # ── 2. lead ──────────────────────────────────────────────────────
        lead = db.query(Lead).filter(Lead.id == lead_id).first()
        if not lead:
            raise HTTPException(404, "Lead não encontrado")

        body = await request.json()
        dados = body.get("dados", {})

        dados.setdefault("req_nome",    lead.nome     or "")
        dados.setdefault("req_cpf",     lead.cpf      or "")
        dados.setdefault("req_celular", lead.telefone or "")
        dados.setdefault("modalidade",  lead.modalidade or "refinanciamento")
        dados.setdefault("data_contrato",
            _agora_br().strftime("%d de %B de %Y")
            .replace("January","Janeiro").replace("February","Fevereiro")
            .replace("March","Marco").replace("April","Abril")
            .replace("May","Maio").replace("June","Junho")
            .replace("July","Julho").replace("August","Agosto")
            .replace("September","Setembro").replace("October","Outubro")
            .replace("November","Novembro").replace("December","Dezembro"))

        # ── 3. gerar PDF ─────────────────────────────────────────────────
        doc_id = secrets.token_hex(8).upper()
        pdf_bytes, hash_doc = gerar_pdf_contrato(dados, doc_id)

        # ── 4. salvar arquivo ────────────────────────────────────────────
        tok = secrets.token_hex(32)
        nome_arquivo = f"contrato_{lead_id}_{tok[:8]}.pdf"
        caminho = salvar_pdf(pdf_bytes, nome_arquivo)

        # ── 5. persistir no banco ────────────────────────────────────────
        tok_prop = secrets.token_hex(32)
        contrato = Contrato(
            lead_id=lead_id,
            criado_por_id=usuario.id,
            token=tok,
            token_prop=tok_prop,
            hash_doc=hash_doc,
            pdf_original=caminho,
            dados_contrato=json.dumps(dados, ensure_ascii=False),
            status="pendente",
            status_prop="pendente",
        )
        db.add(contrato)
        db.commit()
        db.refresh(contrato)

        base_url = str(request.base_url).rstrip("/")
        link_req  = f"{base_url}/assinar/{tok}"
        link_prop = f"{base_url}/assinar/{tok_prop}"

        # ── 6. Enviar link de assinatura para o vendedor/proprietário via WhatsApp ──
        prop_tel = dados.get("prop_telefone", "").strip()
        prop_nome = dados.get("prop_nome", "Proprietário").strip() or "Proprietário"
        if prop_tel:
            msg_prop = (
                f"Olá, {prop_nome}! 👋\n\n"
                f"A Fácil Financiamentos gerou um contrato que requer sua assinatura digital.\n\n"
                f"🔗 Clique no link abaixo para assinar:\n{link_prop}\n\n"
                f"O processo é rápido e 100% online. Qualquer dúvida, entre em contato conosco."
            )
            try:
                await enviar_zapi(prop_tel, msg_prop)
            except Exception:
                pass  # não bloqueia se falhar o envio

        return {
            "contrato_id": contrato.id,
            "hash": hash_doc,
            "doc_id": doc_id,
            "link_requerente":   link_req,
            "link_proprietario": link_prop,
            # compat retroativa
            "link": link_req,
        }

    except HTTPException:
        raise  # re-lança 404 normalmente
    except Exception as exc:
        tb_str = _tb.format_exc()
        print(f"❌ Erro em gerar_contrato_endpoint:\n{tb_str}")
        raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {exc}")


@app.get("/assinar/{token}", response_class=HTMLResponse)
async def pagina_assinar(token: str):
    caminho = Path("templates/assinar.html")
    return HTMLResponse(caminho.read_text(encoding="utf-8"))


def _detectar_role_contrato(token: str, db):
    """Retorna (contrato, role) onde role = 'requerente' | 'proprietario'."""
    c = db.query(Contrato).filter(Contrato.token == token).first()
    if c:
        return c, "requerente"
    c = db.query(Contrato).filter(Contrato.token_prop == token).first()
    if c:
        return c, "proprietario"
    return None, None


@app.get("/assinar/{token}/pdf-original")
async def pdf_preview_contrato(token: str, db: Session = Depends(get_db)):
    c, _ = _detectar_role_contrato(token, db)
    if not c:
        raise HTTPException(404, "Contrato não encontrado")
    if not c.pdf_original:
        raise HTTPException(404, "PDF não disponível")
    p = Path(c.pdf_original)
    if not p.exists():
        raise HTTPException(404, "Arquivo não encontrado no servidor")
    return FileResponse(str(p), media_type="application/pdf",
                        headers={"Content-Disposition": "inline"})


@app.get("/assinar/{token}/conteudo")
async def conteudo_contrato(token: str, db: Session = Depends(get_db)):
    contrato, role = _detectar_role_contrato(token, db)
    if not contrato:
        raise HTTPException(404, "Contrato não encontrado")
    if role == "requerente"   and contrato.status      == "assinado":
        raise HTTPException(410, "Contrato ja assinado")
    if role == "proprietario" and contrato.status_prop == "assinado":
        raise HTTPException(410, "Contrato ja assinado")

    d = json.loads(contrato.dados_contrato or "{}") if contrato.dados_contrato else {}
    nome_req  = d.get("req_nome",  "-")
    nome_prop = d.get("prop_nome", "-")

    # Mascara o telefone do lead para exibir na tela de confirmação
    lead = db.query(Lead).filter(Lead.id == contrato.lead_id).first()
    tel_mascarado = ""
    if lead and lead.telefone:
        t = re.sub(r'\D', '', lead.telefone)
        if len(t) >= 8:
            tel_mascarado = f"(*****){t[-4:]}"

    return {
        "hash":              contrato.hash_doc,
        "role":              role,
        "nome_req":          nome_req,
        "nome_prop":         nome_prop,
        "doc_id":            d.get("doc_id", ""),
        "data":              d.get("data_contrato", ""),
        "telefone_mascarado": tel_mascarado,
    }


@app.post("/assinar/{token}/enviar-codigo")
async def enviar_codigo_otp(token: str, db: Session = Depends(get_db)):
    """Gera e envia código OTP de 6 dígitos via WhatsApp para confirmação de assinatura."""
    import random
    contrato, role = _detectar_role_contrato(token, db)
    if not contrato:
        raise HTTPException(404, "Contrato não encontrado")

    codigo = str(random.randint(100000, 999999))
    expira = datetime.utcnow() + timedelta(minutes=10)

    if role == "requerente":
        contrato.codigo_req        = codigo
        contrato.codigo_req_expira = expira
    else:
        contrato.codigo_prop        = codigo
        contrato.codigo_prop_expira = expira
    db.commit()

    lead = db.query(Lead).filter(Lead.id == contrato.lead_id).first()
    telefone = lead.telefone if lead else None

    if telefone:
        msg = (
            f"🔐 *Código de confirmação — Fácil Financiamentos*\n\n"
            f"Olá! Seu código para confirmar a assinatura do contrato é:\n\n"
            f"*{codigo}*\n\n"
            f"⏱ Válido por 10 minutos.\n"
            f"Não compartilhe este código com ninguém."
        )
        await enviar_zapi(telefone, msg)

    tel_mascarado = ""
    if telefone:
        t = re.sub(r'\D', '', telefone)
        if len(t) >= 4:
            tel_mascarado = f"(*****){t[-4:]}"

    return {"ok": True, "telefone_mascarado": tel_mascarado}


@app.post("/assinar/{token}/verificar-codigo")
async def verificar_codigo_otp(token: str, request: Request, db: Session = Depends(get_db)):
    """Valida o código OTP digitado pelo assinante."""
    body = await request.json()
    codigo_digitado = str(body.get("codigo", "")).strip()

    contrato, role = _detectar_role_contrato(token, db)
    if not contrato:
        raise HTTPException(404, "Contrato não encontrado")

    if role == "requerente":
        codigo_salvo = contrato.codigo_req
        expira       = contrato.codigo_req_expira
    else:
        codigo_salvo = contrato.codigo_prop
        expira       = contrato.codigo_prop_expira

    if not codigo_salvo:
        raise HTTPException(400, detail="Nenhum código foi enviado. Solicite um novo código.")
    if not expira or datetime.utcnow() > expira:
        raise HTTPException(400, detail="Código expirado. Solicite um novo código.")
    if codigo_digitado != codigo_salvo:
        raise HTTPException(400, detail="Código incorreto. Verifique e tente novamente.")

    # Invalida o código após uso bem-sucedido
    if role == "requerente":
        contrato.codigo_req        = None
        contrato.codigo_req_expira = None
    else:
        contrato.codigo_prop        = None
        contrato.codigo_prop_expira = None
    db.commit()

    return {"ok": True}


@app.post("/assinar/{token}")
async def submeter_assinatura(token: str, request: Request, db: Session = Depends(get_db)):
    import traceback as _tb
    try:
        from gerar_contrato import (
            base64_para_imagem, gerar_pdf_final_completo,
            salvar_pdf, CONTRATOS_DIR,
        )
        from pathlib import Path as Pt

        contrato, role = _detectar_role_contrato(token, db)
        if not contrato:
            raise HTTPException(404, "Contrato nao encontrado")
        if role == "requerente"   and contrato.status      == "assinado":
            raise HTTPException(410, "Contrato ja assinado")
        if role == "proprietario" and contrato.status_prop == "assinado":
            raise HTTPException(410, "Contrato ja assinado")

        body = await request.json()
        selfie_b64     = body.get("selfie", "")
        assin_b64      = body.get("assinatura", "")
        doc_frente_b64 = body.get("doc_frente", "")
        doc_verso_b64  = body.get("doc_verso", "")
        geo            = body.get("geo", "")
        ip             = request.client.host if request.client else "desconhecido"
        agora          = _agora_br().strftime("%d/%m/%Y %H:%M:%S")

        base = CONTRATOS_DIR / f"contrato_{contrato.lead_id}_{token[:8]}"
        d_contrato = json.loads(contrato.dados_contrato or "{}") if contrato.dados_contrato else {}
        base_url = str(request.base_url).rstrip("/")

        if role == "requerente":
            selfie_path = str(base) + "_selfie_req.jpg"
            assin_path  = str(base) + "_assin_req.png"
            frente_path = str(base) + "_doc_frente_req.jpg"
            verso_path  = str(base) + "_doc_verso_req.jpg"
            base64_para_imagem(selfie_b64,     Pt(selfie_path))
            base64_para_imagem(assin_b64,      Pt(assin_path))
            base64_para_imagem(doc_frente_b64, Pt(frente_path))
            base64_para_imagem(doc_verso_b64,  Pt(verso_path))

            contrato.status              = "assinado"
            contrato.selfie_path         = selfie_path
            contrato.assinatura_path     = assin_path
            contrato.doc_frente_req_path = frente_path
            contrato.doc_verso_req_path  = verso_path
            contrato.ip_cliente          = ip
            contrato.geolocalizacao      = geo
            contrato.assinado_em         = datetime.utcnow()

        else:  # proprietario
            selfie_path = str(base) + "_selfie_prop.jpg"
            assin_path  = str(base) + "_assin_prop.png"
            frente_path = str(base) + "_doc_frente_prop.jpg"
            verso_path  = str(base) + "_doc_verso_prop.jpg"
            base64_para_imagem(selfie_b64,     Pt(selfie_path))
            base64_para_imagem(assin_b64,      Pt(assin_path))
            base64_para_imagem(doc_frente_b64, Pt(frente_path))
            base64_para_imagem(doc_verso_b64,  Pt(verso_path))

            contrato.status_prop          = "assinado"
            contrato.selfie_prop_path     = selfie_path
            contrato.assinatura_prop_path = assin_path
            contrato.doc_frente_prop_path = frente_path
            contrato.doc_verso_prop_path  = verso_path
            contrato.ip_prop              = ip
            contrato.geo_prop             = geo
            contrato.assinado_prop_em     = datetime.utcnow()

        db.commit()
        db.refresh(contrato)

        # ── Gera PDF final unificado com todas as assinaturas disponíveis ──────
        if not contrato.pdf_original or not Pt(contrato.pdf_original).exists():
            raise HTTPException(500, "PDF original nao encontrado")

        dados_req  = None
        dados_prop = None

        if contrato.status == "assinado":
            dados_req = {
                "assinado_em": contrato.assinado_em.strftime("%d/%m/%Y %H:%M:%S") if contrato.assinado_em else agora,
                "ip": contrato.ip_cliente or ip,
                "geo": contrato.geolocalizacao or "nao fornecida",
                "hash_doc": contrato.hash_doc,
                "nome": d_contrato.get("req_nome") or (contrato.lead.nome if contrato.lead else "-") or "-",
                "cpf":  d_contrato.get("req_cpf")  or (contrato.lead.cpf  if contrato.lead else "-") or "-",
            }

        if contrato.status_prop == "assinado":
            dados_prop = {
                "assinado_em": contrato.assinado_prop_em.strftime("%d/%m/%Y %H:%M:%S") if contrato.assinado_prop_em else agora,
                "ip": contrato.ip_prop or ip,
                "geo": contrato.geo_prop or "nao fornecida",
                "hash_doc": contrato.hash_doc,
                "nome": d_contrato.get("prop_nome", "-"),
                "cpf":  d_contrato.get("prop_cpf",  "-"),
            }

        pdf_final = gerar_pdf_final_completo(
            contrato.pdf_original,
            assin_req_path=contrato.assinatura_path,
            selfie_req_path=contrato.selfie_path,
            dados_req=dados_req,
            doc_frente_req_path=contrato.doc_frente_req_path,
            doc_verso_req_path=contrato.doc_verso_req_path,
            assin_prop_path=contrato.assinatura_prop_path,
            selfie_prop_path=contrato.selfie_prop_path,
            dados_prop=dados_prop,
            doc_frente_prop_path=contrato.doc_frente_prop_path,
            doc_verso_prop_path=contrato.doc_verso_prop_path,
            doc_id=d_contrato.get("doc_id", ""),
            verificacao_url_req=f"{base_url}/verificar/{contrato.token}",
            verificacao_url_prop=f"{base_url}/verificar/{contrato.token_prop}" if contrato.token_prop else "",
        )

        pdf_final_path = str(base) + "_assinado_final.pdf"
        Pt(pdf_final_path).write_bytes(pdf_final)
        contrato.pdf_assinado = pdf_final_path
        db.commit()

        return {"status": "ok", "assinado_em": agora, "role": role}

    except HTTPException:
        raise
    except Exception as exc:
        print(f"Erro em submeter_assinatura: {_tb.format_exc()}")
        raise HTTPException(500, f"{type(exc).__name__}: {exc}")


# ── Verificação pública de assinatura ────────────────────────────────────────
@app.get("/verificar/{token}", response_class=HTMLResponse)
async def verificar_assinatura(token: str, db: Session = Depends(get_db)):
    c = db.query(Contrato).filter(Contrato.token == token).first()
    if not c:
        return HTMLResponse("<h1>Contrato não encontrado</h1>", status_code=404)

    d = json.loads(c.dados_contrato or "{}") if c.dados_contrato else {}
    nome = d.get("req_nome") or (c.lead.nome if c.lead else "-") or "-"
    cpf  = d.get("req_cpf")  or (c.lead.cpf  if c.lead else "-") or "-"
    assinado_em_br = _fmt_br(c.assinado_em, "%d/%m/%Y %H:%M:%S") or "-"
    criado_em_br   = _fmt_br(c.criado_em) or "-"

    status_badge = (
        '<span style="background:#16a34a;color:#fff;padding:.3rem .9rem;border-radius:99px;font-weight:700;font-size:.9rem">✅ ASSINADO</span>'
        if c.status == "assinado" else
        '<span style="background:#f59e0b;color:#fff;padding:.3rem .9rem;border-radius:99px;font-weight:700;font-size:.9rem">⏳ PENDENTE</span>'
    )

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Verificação de Assinatura — Fácil Financiamentos</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Segoe UI',system-ui,sans-serif;background:#f4f6fb;color:#1a202c}}
  .header{{background:#0d2b4e;color:#fff;padding:1.25rem;text-align:center}}
  .header h1{{font-size:1.1rem;font-weight:700}}
  .header p{{font-size:.78rem;opacity:.65;margin-top:3px}}
  .container{{max-width:560px;margin:1.5rem auto;padding:0 1rem 2rem}}
  .card{{background:#fff;border-radius:12px;box-shadow:0 4px 20px rgba(0,0,0,.08);padding:1.5rem;margin-bottom:1rem}}
  .card h2{{font-size:.95rem;color:#0d2b4e;margin-bottom:1rem;display:flex;align-items:center;gap:.4rem}}
  .row{{display:flex;justify-content:space-between;align-items:flex-start;padding:.55rem 0;border-bottom:1px solid #f1f5f9;font-size:.85rem}}
  .row:last-child{{border-bottom:none}}
  .row .lbl{{color:#64748b;font-weight:500}}
  .row .val{{color:#1a202c;font-weight:600;text-align:right;max-width:60%;word-break:break-all}}
  .hash{{font-size:.68rem;color:#64748b;word-break:break-all;line-height:1.5;margin-top:.5rem;padding:.5rem;background:#f8fafc;border-radius:6px;border:1px solid #e2e8f0}}
  .status-area{{text-align:center;padding:.5rem 0 1rem}}
  .footer{{text-align:center;font-size:.72rem;color:#94a3b8;margin-top:1.5rem}}
</style>
</head>
<body>
<div class="header">
  <h1>Fácil Financiamentos</h1>
  <p>Verificação de Assinatura Eletrônica</p>
</div>
<div class="container">
  <div class="card">
    <div class="status-area">{status_badge}</div>
    <h2>📄 Dados do Documento</h2>
    <div class="row"><span class="lbl">Nº do documento</span><span class="val">{d.get("doc_id", c.id)}</span></div>
    <div class="row"><span class="lbl">Gerado em</span><span class="val">{criado_em_br}</span></div>
    <div class="row"><span class="lbl">Assinado em</span><span class="val">{assinado_em_br}</span></div>
    <div class="row"><span class="lbl">IP do assinante</span><span class="val">{c.ip_cliente or "-"}</span></div>
  </div>
  <div class="card">
    <h2>👤 Assinante</h2>
    <div class="row"><span class="lbl">Nome</span><span class="val">{nome.upper()}</span></div>
    <div class="row"><span class="lbl">CPF</span><span class="val">{cpf}</span></div>
  </div>
  <div class="card">
    <h2>🔒 Integridade</h2>
    <p style="font-size:.8rem;color:#64748b;margin-bottom:.5rem">Hash SHA-256 do documento original:</p>
    <div class="hash">{c.hash_doc}</div>
  </div>
  <div class="footer">
    Assinado eletronicamente nos termos da Lei 14.063/2020 e MP 2.200-2/2001.<br>
    Fácil Financiamentos · Rua Lauro Ignacio Ponte, 08, Sala 202 · BH/MG
  </div>
</div>
</body></html>"""
    return HTMLResponse(html)


@app.get("/api/leads/{lead_id}/contratos")
async def listar_contratos(
    lead_id: int,
    request: Request,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    contratos = db.query(Contrato).filter(Contrato.lead_id == lead_id).order_by(Contrato.criado_em.desc()).all()
    base_url = str(request.base_url).rstrip("/")
    return [
        {
            "id": c.id,
            "criado_em":       _fmt_br(c.criado_em) or "-",
            # Requerente
            "status_req":      c.status or "pendente",
            "assinado_req_em": _fmt_br(c.assinado_em),
            "link_req":        f"{base_url}/assinar/{c.token}",
            # Proprietário
            "status_prop":     c.status_prop or "pendente",
            "assinado_prop_em": _fmt_br(c.assinado_prop_em),
            "link_prop":       f"{base_url}/assinar/{c.token_prop}" if c.token_prop else None,
            # PDF disponível assim que qualquer parte assinar
            "pdf_id":          c.id if c.pdf_assinado else None,
            "ambos_assinaram": c.status == "assinado" and c.status_prop == "assinado",
        }
        for c in contratos
    ]


@app.get("/api/contratos/{contrato_id}/pdf")
async def baixar_pdf_assinado(
    contrato_id: int,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual),
):
    c = db.query(Contrato).filter(Contrato.id == contrato_id).first()
    if not c:
        raise HTTPException(404, "Contrato não encontrado")
    if not c.pdf_assinado:
        raise HTTPException(400, "Contrato ainda não assinado")
    p = Path(c.pdf_assinado)
    if not p.exists():
        raise HTTPException(404, "Arquivo não encontrado")
    return FileResponse(str(p), media_type="application/pdf", filename=f"contrato_assinado_{contrato_id}.pdf")


# ─── Helpers ─────────────────────────────────────────────────────────────────────

def _parse_observacoes(raw: str | None) -> list:
    """Retorna observações como lista de dicts {texto, usuario, em}.
    Compatível com formato antigo (texto puro) e novo (JSON array)."""
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, list):
            return parsed
        # Era um objeto único, converte
        return [{"texto": str(parsed), "usuario": "—", "em": "—"}]
    except Exception:
        # Formato legado: texto puro → converte sem autor
        return [{"texto": raw, "usuario": "—", "em": "—"}]


def _safe_json(raw, padrao):
    """json.loads que nunca quebra — retorna o padrão se o conteúdo estiver inválido/vazio."""
    if not raw:
        return padrao
    try:
        return json.loads(raw)
    except Exception:
        return padrao


def _serial_lead(l: Lead, db: Session) -> dict:
    responsavel = None
    if l.atribuido_para:
        u = db.query(Usuario).filter(Usuario.id == l.atribuido_para).first()
        if u:
            responsavel = {"id": u.id, "nome": u.nome}
    return {
        "id": l.id,
        "telefone": l.telefone,
        "nome": l.nome or "—",
        "cpf": l.cpf or "—",
        "data_nascimento": l.data_nascimento or "—",
        "carro_interesse": l.carro_interesse or "—",
        "modalidade": l.modalidade,
        "status": l.status,
        "estado_conversa": l.estado_conversa,
        "responsavel": responsavel,
        "assumido_em": _fmt_br(l.assumido_em),
        "criado_em": _fmt_br(l.criado_em) or "—",
        "atualizado_em": _fmt_br(l.atualizado_em) or "—",
        "observacoes": _parse_observacoes(l.observacoes),
        "origem": l.origem or "whatsapp",
        "origem_detalhe": l.origem_detalhe or "",
        "parceiro_id": l.parceiro_id,
        "parceiro_nome": l.parceiro.nome if l.parceiro else "",
        # Dados do contrato fechado
        "deal_data":     l.deal_data or "",
        "deal_veiculo":  l.deal_veiculo or "",
        "deal_placa":    l.deal_placa or "",
        "deal_retorno":  l.deal_retorno or "",
        "deal_valor":    l.deal_valor or "",
        "deal_comissao": l.deal_comissao or "",
        "deal_banco":      l.deal_banco or "",
        "deal_conta_pg":   l.deal_conta_pg or "",
        "deal_operadora":  l.deal_operadora or "",
        # Dados extras p/ requerimento
        "dados_contrato": _safe_json(l.dados_contrato, {}),
        # Perfil do cliente
        "cidade":   l.cidade   or "",
        "email":    l.email    or "",
        "renda":    l.renda    or "",
        "profissao": l.profissao or "",
        "tem_cnh":  l.tem_cnh,   # None=não informado | True=sim | False=não
        "oculto_funil": bool(l.oculto_funil),
        "descadastrado": bool(l.descadastrado),
        "ignorar_relatorios": bool(l.ignorar_relatorios),
        "carros_proposta": _safe_json(l.carros_proposta, []),
    }


def _serial_usuario(u: Usuario) -> dict:
    return {
        "id": u.id, "nome": u.nome, "email": u.email,
        "role": u.role, "ativo": u.ativo,
        "criado_em": _fmt_br(u.criado_em, "%d/%m/%Y") or "—",
    }


def _criar_config_padrao(db: Session):
    """Cria configurações padrão do bot se não existirem."""
    configs_padrao = [
        {
            "chave": "regras_financiamento",
            "descricao": "Regras específicas para leads de financiamento de veículo",
            "valor": (
                "REGRAS PARA FINANCIAMENTO:\n"
                "- O fechamento do contrato é PRESENCIAL em Belo Horizonte/MG (exigência do banco)\n"
                "- Após o cliente informar a cidade, verifique se está a até 200km de BH\n"
                "- Cidades dentro de ~200km de BH: Contagem, Betim, Sete Lagoas, Montes Claros, Governador Valadares, Ipatinga, Coronel Fabriciano, Juiz de Fora, Uberlândia, Uberaba, Divinópolis, Itabira, João Monlevade, Conselheiro Lafaiete, Ouro Preto, Poços de Caldas, Pouso Alegre, Varginha, Lavras, São João del-Rei, Barbacena, Viçosa, Muriaé\n"
                "- Se o cliente for de cidade FORA desse raio, pergunte: 'Para o financiamento, o fechamento do contrato é feito presencialmente aqui em BH (exigência do banco). Você teria disponibilidade de vir até nós?'\n"
                "- Se o cliente NÃO puder vir: informe gentilmente que para financiamento precisamos do fechamento presencial e sugira o refinanciamento caso ele já tenha um veículo. Marque como desqualificado.\n"
                "- Se o cliente PUDER vir: continue normalmente com a coleta de dados"
            ),
        },
        {
            "chave": "regras_refinanciamento",
            "descricao": "Regras específicas para leads de refinanciamento/CGI",
            "valor": (
                "REGRAS PARA REFINANCIAMENTO:\n"
                "- Atendimento pode ser 100% ONLINE, de qualquer lugar do Brasil\n"
                "- Não há restrição geográfica\n"
                "- O cliente pode ter o veículo quitado ou semi-quitado\n"
                "- Continue normalmente com a coleta de dados independente da cidade"
            ),
        },
        {
            "chave": "mensagem_boas_vindas",
            "descricao": "Mensagem inicial de boas-vindas (enviada quando o cliente entra em contato)",
            "valor": (
                "Olá, seja bem vindo a Fácil Financiamentos. Meu nome é Maria e sou sua atendente virtual, estou aqui para ajudá-lo!\n\n"
                "Qual o seu nome?"
            ),
        },
        {
            "chave": "mensagem_finalizacao",
            "descricao": "Mensagem de encerramento após coletar todos os dados",
            "valor": "Obrigado pelas confirmações, em breve uma de nossas consultoras, entrará em contato. 🤝",
        },
        {
            "chave": "mensagem_followup",
            "descricao": "1º follow-up: enviado após X horas sem resposta do cliente",
            "valor": "Oi! 😊 Vi que nossa conversa ficou parada...\nQuando quiser continuar, estou aqui! Gostaria de retomar?",
        },
        {
            "chave": "mensagem_followup_2",
            "descricao": "2º follow-up: enviado 24h após o 1º sem resposta",
            "valor": "Olá! 👋 Passando para saber se ainda tem interesse em financiar ou refinanciar seu veículo.\nEstamos com ótimas condições e podemos te ajudar! Me conta, ficou alguma dúvida?",
        },
        {
            "chave": "mensagem_followup_3",
            "descricao": "3º follow-up: última tentativa, enviado 48h após o 2º. Após isso o lead é marcado como Perdido.",
            "valor": "Oi! Última tentativa de contato por aqui. 😊\nSe mudar de ideia sobre o financiamento ou refinanciamento, pode nos chamar a qualquer momento!\nFicamos à disposição. 🤝",
        },
        {
            "chave": "followup_horas",
            "descricao": "Horas de inatividade antes de enviar o 1º follow-up (padrão: 4)",
            "valor": "4",
        },
        {
            "chave": "followup_ativo",
            "descricao": "Liga/desliga TODOS os follow-ups automáticos. Use 1 para ligado, 0 para desligado.",
            "valor": "1",
        },
        {
            "chave": "followup_max_rodada",
            "descricao": "Máximo de follow-ups enviados por rodada (anti-spam/ban). Padrão: 20",
            "valor": "20",
        },
        {
            "chave": "followup_max_dias",
            "descricao": "Não inicia follow-up em lead cuja última mensagem é mais antiga que X dias. Padrão: 15",
            "valor": "15",
        },
        {
            "chave": "followup_optout",
            "descricao": "Inclui a frase de descadastro (SAIR) nos follow-ups — reduz risco de denúncia/ban. 1=sim, 0=não",
            "valor": "1",
        },
        {
            "chave": "followup_optout_texto",
            "descricao": "Frase de descadastro adicionada ao fim dos follow-ups",
            "valor": "\n\n_Se não quiser mais receber, responda SAIR._",
        },
        {
            "chave": "apiplacas_token",
            "descricao": "Token da API Placas (apiplacas.com.br) para buscar dados do veículo pela placa",
            "valor": "",
        },
        {
            "chave": "parada_alerta_min",
            "descricao": "Minutos sem resposta para alertar 'conversa parada' no topo do painel. Padrão: 30",
            "valor": "30",
        },
        {
            "chave": "template_proposta",
            "descricao": "Modelo de mensagem: Proposta completa (enviado ao cliente pelo botão Modelos)",
            "valor": (
                "PROPOSTA PARA FINANCIAMENTO\n\n"
                "NOME :\nCPF:\nRG:\nDATA NASCIMENTO:\nNATURALIDADE (cidade):\n\n"
                "POSSUI CNH? (   ) SIM   (    ) NÃO\n\n"
                "PAI:\nMÃE:\n\n"
                "RUA/AV:\nN°:\nBAIRRO:\nCEP:\nN° CELULAR:\nEMAIL:\n\n"
                "*DADOS COMERCIAIS*\n\n"
                "EMPRESA ONDE TRABALHA:\nCARGO:\nTEMPO DE SERVIÇO:\n"
                "ENDEREÇO DE TRABALHO RUA/AV:\nN°:\nBAIRRO:\nN° TELEFONE EMPRESA:\n\n"
                "RENDA MENSAL:\n\nOUTRAS RENDAS:"
            ),
        },
        {
            "chave": "template_pre_analise",
            "descricao": "Modelo de mensagem: Pré-análise (enviado ao cliente pelo botão Modelos)",
            "valor": (
                "Dados necessários para pré análise:\n"
                "• CPF:\n• Data de nascimento:\n• Modelo e ano do veículo:\n• Possui CNH?"
            ),
        },
        {
            "chave": "meta_contratos",
            "descricao": "Meta de contratos fechados no mês (número inteiro)",
            "valor": "20",
        },
    ]

    for c in configs_padrao:
        existe = db.query(Configuracao).filter(Configuracao.chave == c["chave"]).first()
        if not existe:
            db.add(Configuracao(chave=c["chave"], valor=c["valor"], descricao=c["descricao"]))
    db.commit()

    # Inicializa fila rotativa se ainda não existir
    fila_existe = db.query(Configuracao).filter(Configuracao.chave == "fila_ordem").first()
    if not fila_existe:
        import json as _json
        # Busca funcionárias na ordem: Camila, Larissa, Luana
        nomes_ordem = ["Camila", "Larissa", "Luana"]
        ids_fila = []
        for nome in nomes_ordem:
            u = db.query(Usuario).filter(Usuario.nome.ilike(f"%{nome}%"), Usuario.ativo == True).first()
            if u:
                ids_fila.append(u.id)
        # Completa com demais funcionárias ativas não incluídas
        todos = db.query(Usuario).filter(Usuario.ativo == True, Usuario.role == RoleEnum.funcionario).all()
        for u in todos:
            if u.id not in ids_fila:
                ids_fila.append(u.id)
        if ids_fila:
            db.add(Configuracao(chave="fila_ordem", valor=_json.dumps(ids_fila), descricao="Fila rotativa de atendimento"))
            db.add(Configuracao(chave="fila_posicao", valor="0", descricao="Posição atual da fila rotativa"))
            db.commit()
            print(f"✅ Fila rotativa inicializada com {len(ids_fila)} funcionária(s)")


def _salvar_msg_webhook(db: Session, telefone: str, texto: str, role: str = "user"):
    from models import MensagemConversa
    msg = MensagemConversa(telefone=telefone, role=role, conteudo=texto)
    db.add(msg)
    db.commit()


async def _notificar_equipe(telefone: str, db: Session):
    resumo = obter_resumo_lead(telefone, db)
    if resumo:
        print("\n🎉 NOVO LEAD QUALIFICADO!")
        print(json.dumps(resumo, ensure_ascii=False, indent=2))
        print("─" * 40)
